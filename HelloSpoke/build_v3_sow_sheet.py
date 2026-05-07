"""Build HelloSpoke v3 SOW project sheet — multi-tab Excel workbook.

Tabs:
  1. Cover & Pricing
  2. Workstream Hours
  3. Task Breakdown (workstream-by-workstream)
  4. Phase Matrix
  5. Phase Requirements
  6. Phase Success Criteria
  7. Risk Register
  8. Scope Assumptions
  9. Out of Scope
 10. Open Items

Sayer brand colors per ~/.claude/rules and brand guide:
  Yellow #FEC700, Grey700 #2E2E2E, Black #000000
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Brand
YELLOW = "FEC700"
GREY = "2E2E2E"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"

HEADER_FILL = PatternFill("solid", fgColor=GREY)
HEADER_FONT = Font(bold=True, color=WHITE, size=11)
ACCENT_FILL = PatternFill("solid", fgColor=YELLOW)
ACCENT_FONT = Font(bold=True, color="000000", size=11)
TOTAL_FILL = PatternFill("solid", fgColor=YELLOW)
TOTAL_FONT = Font(bold=True, color="000000", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor=LIGHT_GREY)

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="top")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_body(ws, start_row, end_row, ncols, zebra=True):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
            if zebra and (r - start_row) % 2 == 1:
                cell.fill = ZEBRA_FILL


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()


# =========================================================================
# Tab 1 — Cover & Pricing
# =========================================================================
ws = wb.active
ws.title = "Cover & Pricing"

ws["A1"] = "HelloSpoke — HubSpot CRM Implementation"
ws["A1"].font = Font(bold=True, size=18)
ws["A2"] = "v3 SOW Project Sheet · 2026-05-07"
ws["A2"].font = Font(italic=True, size=11)

ws["A4"] = "Engagement Summary"
ws["A4"].font = ACCENT_FONT
ws["A4"].fill = ACCENT_FILL

cover_rows = [
    ("Client", "HelloSpoke"),
    ("Sayer Lead", "Cameron Taggart (deal owner) · Kyle Harbuck (scope/delivery)"),
    ("Engagement Type", "HubSpot CRM Implementation — configuration, automation, UX, enablement"),
    ("Source CRM", "Salesforce (contract expires October 2026)"),
    ("Target CRM", "HubSpot Sales Hub Professional + Commerce Hub Professional"),
    ("Timeline", "12 weeks"),
    ("Rate", "$175/hr"),
    ("Median Hours", "164"),
    ("Median Investment", "$28,700"),
    ("Investment Range", "$25,375 (low) – $32,900 (high)"),
    ("Optional Add-On (QuotaPath)", "$2,450 median (decision pending Jeremy 2026-05-08)"),
    ("Payment Terms", "Equal monthly installments of $7,175 over 4 months · Net-15"),
    ("Tech & Admin Fee", "5% applied to each invoice"),
    ("Source Call (Rescope)", "Fireflies 01KQSVGHKNQ1PM7X5ZPWJ4GJJK · 2026-05-04 walkthrough"),
    ("Decisions Log", "HelloSpoke_decisions.md (5/4 + 5/7 entries)"),
    ("Supersedes", "4/24 proposal (anchor $44,000) and Gamma deck erumqig7pyx7rhm"),
]
for i, (k, v) in enumerate(cover_rows, start=5):
    ws.cell(row=i, column=1, value=k).font = Font(bold=True)
    ws.cell(row=i, column=2, value=v).alignment = WRAP
    ws.cell(row=i, column=1).border = BORDER
    ws.cell(row=i, column=2).border = BORDER

ws["A22"] = "Phase Reweighting"
ws["A22"].font = ACCENT_FONT
ws["A22"].fill = ACCENT_FILL

phase_table = [
    ("Phase", "Weeks", "Workstreams", "Median Hours"),
    ("Phase 1 — Foundation + Audit", "Wk 1–4", "W1, W2, W3", 46),
    ("Phase 2 — Build", "Wk 5–9", "W4, W5, W6, W7, W8", 80),
    ("Phase 3 — Enable + Go-Live", "Wk 10–12", "W9, W10", 22),
    ("Cross-phase", "All", "PM", 16),
    ("TOTAL", "12 weeks", "", 164),
]
for i, row in enumerate(phase_table, start=23):
    for j, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.border = BORDER
        cell.alignment = WRAP
        if i == 23:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        if i == 28:  # TOTAL row
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT

set_widths(ws, [32, 70])


# =========================================================================
# Tab 2 — Workstream Hours
# =========================================================================
ws = wb.create_sheet("Workstream Hours")

cols = ["#", "Workstream", "Low", "Median", "High",
        "Low $", "Median $", "High $", "Phase"]
ws.append(cols)
style_header(ws, 1, len(cols))

ws_data = [
    ("W1", "HubSpot Foundational Architecture", 14, 18, 22, "Phase 1"),
    ("W2", "Data QA + Pre-Production Audit", 14, 18, 22, "Phase 1"),
    ("W3", "Sales Stage Definitions + Gating", 8, 10, 14, "Phase 1"),
    ("W4", "Quoting & CPQ Workflow (Commerce Hub)", 20, 24, 30, "Phase 2"),
    ("W5", "DocuSign Replacement Evaluation + Setup", 4, 6, 12, "Phase 2"),
    ("W6", "Automations & Workflows", 16, 20, 26, "Phase 2"),
    ("W7", "Reporting & Dashboards", 14, 18, 24, "Phase 2"),
    ("W8", "ClickUp Workflow Refinement", 8, 12, 16, "Phase 2"),
    ("W9", "Training & Enablement", 10, 12, 16, "Phase 3"),
    ("W10", "UAT & Go-Live", 8, 10, 14, "Phase 3"),
    ("PM", "Project Management", 12, 16, 20, "Cross-phase"),
]
RATE = 175
for r, (wid, name, lo, med, hi, phase) in enumerate(ws_data, start=2):
    ws.cell(row=r, column=1, value=wid)
    ws.cell(row=r, column=2, value=name)
    ws.cell(row=r, column=3, value=lo)
    ws.cell(row=r, column=4, value=med)
    ws.cell(row=r, column=5, value=hi)
    ws.cell(row=r, column=6, value=lo * RATE)
    ws.cell(row=r, column=7, value=med * RATE)
    ws.cell(row=r, column=8, value=hi * RATE)
    ws.cell(row=r, column=9, value=phase)
style_body(ws, 2, len(ws_data) + 1, len(cols))

# Totals row
total_row = len(ws_data) + 2
totals = [sum(d[2] for d in ws_data),
          sum(d[3] for d in ws_data),
          sum(d[4] for d in ws_data)]
ws.cell(row=total_row, column=2, value="TOTAL (raw sum)")
ws.cell(row=total_row, column=3, value=totals[0])
ws.cell(row=total_row, column=4, value=totals[1])
ws.cell(row=total_row, column=5, value=totals[2])
ws.cell(row=total_row, column=6, value=totals[0] * RATE)
ws.cell(row=total_row, column=7, value=totals[1] * RATE)
ws.cell(row=total_row, column=8, value=totals[2] * RATE)
for c in range(1, len(cols) + 1):
    ws.cell(row=total_row, column=c).fill = TOTAL_FILL
    ws.cell(row=total_row, column=c).font = TOTAL_FONT
    ws.cell(row=total_row, column=c).border = BORDER

# Proposal range row
proposal_row = total_row + 1
ws.cell(row=proposal_row, column=2, value="PROPOSAL RANGE (tightened)")
ws.cell(row=proposal_row, column=3, value=145)
ws.cell(row=proposal_row, column=4, value=164)
ws.cell(row=proposal_row, column=5, value=188)
ws.cell(row=proposal_row, column=6, value=145 * RATE)
ws.cell(row=proposal_row, column=7, value=164 * RATE)
ws.cell(row=proposal_row, column=8, value=188 * RATE)
for c in range(1, len(cols) + 1):
    ws.cell(row=proposal_row, column=c).fill = TOTAL_FILL
    ws.cell(row=proposal_row, column=c).font = TOTAL_FONT
    ws.cell(row=proposal_row, column=c).border = BORDER

# Optional QuotaPath
opt_row = proposal_row + 2
ws.cell(row=opt_row, column=1, value="QP")
ws.cell(row=opt_row, column=2, value="QuotaPath Integration (OPTIONAL — pending 5/8)")
ws.cell(row=opt_row, column=3, value=10)
ws.cell(row=opt_row, column=4, value=14)
ws.cell(row=opt_row, column=5, value=18)
ws.cell(row=opt_row, column=6, value=10 * RATE)
ws.cell(row=opt_row, column=7, value=14 * RATE)
ws.cell(row=opt_row, column=8, value=18 * RATE)
ws.cell(row=opt_row, column=9, value="Optional")
for c in range(1, len(cols) + 1):
    ws.cell(row=opt_row, column=c).fill = ACCENT_FILL
    ws.cell(row=opt_row, column=c).font = Font(bold=True)
    ws.cell(row=opt_row, column=c).border = BORDER

# $ formatting
for r in range(2, opt_row + 1):
    for c in (6, 7, 8):
        ws.cell(row=r, column=c).number_format = '"$"#,##0'

set_widths(ws, [6, 44, 8, 10, 8, 12, 14, 12, 14])


# =========================================================================
# Tab 3 — Task Breakdown
# =========================================================================
ws = wb.create_sheet("Task Breakdown")
cols = ["WS", "Task ID", "Task", "Phase", "Owner",
        "Est. Hrs (median)", "Notes / Dependencies"]
ws.append(cols)
style_header(ws, 1, len(cols))

# Comprehensive task breakdown — sized to the median hours per workstream
tasks = [
    # W1 — 18 hrs
    ("W1", "ARC-T01", "Validate HubSpot tier (Sales Hub Pro + Commerce Hub Pro requirement)", "Phase 1", "Sayer", 2, "Blocker for W4; surface upgrade decision early"),
    ("W1", "ARC-T02", "Configure standard objects (Companies, Contacts, Deals)", "Phase 1", "Sayer", 3, "—"),
    ("W1", "ARC-T03", "Review + validate custom objects from Jeremy's sandbox", "Phase 1", "Sayer", 2, "Sandbox access required"),
    ("W1", "ARC-T04", "Build sales pipeline with Sara's finalized stages", "Phase 1", "Sayer", 3, "Depends on Sara delivering stage definitions wk 1"),
    ("W1", "ARC-T05", "Configure lead status + lifecycle stage progression", "Phase 1", "Sayer", 2, "—"),
    ("W1", "ARC-T06", "Property strategy + custom property creation (snake_case, group org)", "Phase 1", "Sayer", 3, "Archive redundant SF-era properties"),
    ("W1", "ARC-T07", "Team / permission structure (sales, ops, leadership)", "Phase 1", "Sayer", 2, "—"),
    ("W1", "ARC-T08", "Document architecture decisions in admin playbook", "Phase 1", "Sayer", 1, "—"),

    # W2 — 18 hrs
    ("W2", "QA-T01", "Sandbox access + initial inventory (record counts, object coverage)", "Phase 1", "Sayer", 3, "Blocker: Jeremy sandbox access"),
    ("W2", "QA-T02", "Data integrity audit (completeness, dedupe, null rates, formatting)", "Phase 1", "Sayer", 4, "—"),
    ("W2", "QA-T03", "Object/property mapping validation (source → target correctness)", "Phase 1", "Sayer", 3, "—"),
    ("W2", "QA-T04", "Association validation (Contact↔Company↔Deal + custom objects)", "Phase 1", "Sayer", 3, "—"),
    ("W2", "QA-T05", "Custom-object structure review (schema, association labels, pipeline alignment)", "Phase 1", "Sayer", 2, "—"),
    ("W2", "QA-T06", "Error handling + rollback plan documentation", "Phase 1", "Sayer", 2, "—"),
    ("W2", "QA-T07", "Production-push green-light memo to Jeremy", "Phase 1", "Sayer", 1, "Closes Phase 1 data side"),

    # W3 — 10 hrs
    ("W3", "STG-T01", "Sales-stage workshop with Sara + Christina", "Phase 1", "Joint", 2, "Sara delivers definitions wk 1"),
    ("W3", "STG-T02", "Build stages in HubSpot pipeline (probabilities, entry/exit criteria)", "Phase 1", "Sayer", 2, "—"),
    ("W3", "STG-T03", "Configure required-field gating: sales → implementation (up to 10 fields)", "Phase 1", "Sayer", 2, "—"),
    ("W3", "STG-T04", "Configure required-field gating: implementation → billing (up to 10 fields)", "Phase 1", "Sayer", 2, "—"),
    ("W3", "STG-T05", "Build stuck-deal HubSpot report for leadership", "Phase 1", "Sayer", 2, "—"),

    # W4 — 24 hrs
    ("W4", "CPQ-T01", "Commerce Hub Pro setup + Commerce Hub seat assignments", "Phase 2", "Sayer", 2, "Tier validated in W1"),
    ("W4", "CPQ-T02", "Product library architecture (SKU, pricing rules, groups)", "Phase 2", "Sayer", 4, "Tier-priced products require manual creation"),
    ("W4", "CPQ-T03", "CSV product upload pattern (import + bulk-update by product ID)", "Phase 2", "Sayer", 4, "Solves Jeremy's #1 named pain"),
    ("W4", "CPQ-T04", "Quote template configuration with HelloSpoke branding", "Phase 2", "Sayer", 3, "—"),
    ("W4", "CPQ-T05", "Validation rules for quotes above pricing thresholds", "Phase 2", "Sayer", 2, "—"),
    ("W4", "CPQ-T06", "Approval workflows on quote creation/submission", "Phase 2", "Sayer", 2, "—"),
    ("W4", "CPQ-T07", "Native e-signature configuration on quotes (Dropbox Sign-powered)", "Phase 2", "Sayer", 3, "Caps: 25/user/mo Pro, 50/user Enterprise"),
    ("W4", "CPQ-T08", "Quote-to-deal workflow automation", "Phase 2", "Sayer", 2, "Quote acceptance → stage advance + tasks"),
    ("W4", "CPQ-T09", "Quoting documentation + handoff into W9 training", "Phase 2", "Sayer", 2, "—"),

    # W5 — 6 hrs
    ("W5", "SIG-T01", "Decision-support workshop with Jeremy (catalog non-quote signing needs)", "Phase 2", "Joint", 1, "Decision deadline aligned with Phase 1 kickoff"),
    ("W5", "SIG-T02", "If Commerce Hub native: signature delegation rules + branded experience", "Phase 2", "Sayer", 3, "Default path; folds into W4 config"),
    ("W5", "SIG-T03", "If PandaDoc selected: account setup + native integration install", "Phase 2", "Sayer", 0, "Alternate; +6 hrs at high end of W5 range"),
    ("W5", "SIG-T04", "If PandaDoc selected: template migration (top 5 doc types)", "Phase 2", "Sayer", 0, "Alternate"),
    ("W5", "SIG-T05", "Audit trail + signing experience verification", "Phase 2", "Sayer", 2, "—"),

    # W6 — 20 hrs
    ("W6", "AUT-T01", "Deal-stage notifications (4 stages min)", "Phase 2", "Sayer", 3, "—"),
    ("W6", "AUT-T02", "Deal-stage task creation on advance", "Phase 2", "Sayer", 3, "—"),
    ("W6", "AUT-T03", "Lifecycle progression: lead → MQL → SQL → opp → customer", "Phase 2", "Sayer", 4, "—"),
    ("W6", "AUT-T04", "Customer-journey: lead capture → company assoc → quote → order → kickoff", "Phase 2", "Sayer", 4, "—"),
    ("W6", "AUT-T05", "Sales ownership routing (round-robin or rule-based)", "Phase 2", "Sayer", 3, "—"),
    ("W6", "AUT-T06", "Stuck-deal alerts to managers", "Phase 2", "Sayer", 2, "—"),
    ("W6", "AUT-T07", "Internal email templates + automated send-on-trigger", "Phase 2", "Sayer", 1, "—"),

    # W7 — 18 hrs
    ("W7", "DASH-T01", "Executive dashboard (revenue, pipeline health, conversion, velocity)", "Phase 2", "Sayer", 4, "—"),
    ("W7", "DASH-T02", "Sales rep dashboard (my deals, my activity, my targets, alerts)", "Phase 2", "Sayer", 4, "—"),
    ("W7", "DASH-T03", "Ops dashboard (onboarding, ClickUp status sync, customer health)", "Phase 2", "Sayer", 4, "—"),
    ("W7", "DASH-T04", "Custom views + saved filters per role (4–6 views)", "Phase 2", "Sayer", 3, "—"),
    ("W7", "DASH-T05", "Report templates + scheduled exports", "Phase 2", "Sayer", 3, "—"),

    # W8 — 12 hrs
    ("W8", "CLK-T01", "Current state review of Jeremy-built integration", "Phase 2", "Sayer", 2, "Validated during W2"),
    ("W8", "CLK-T02", "Workflow correctness audit (statuses, routing, visibility)", "Phase 2", "Sayer", 3, "—"),
    ("W8", "CLK-T03", "Status alignment (canonical onboarding statuses mirrored in HubSpot)", "Phase 2", "Sayer", 3, "—"),
    ("W8", "CLK-T04", "Gating + handoff rules between sales↔impl↔ops", "Phase 2", "Sayer", 2, "—"),
    ("W8", "CLK-T05", "Who-does-what SOPs for implementation team", "Phase 2", "Sayer", 2, "—"),

    # W9 — 12 hrs
    ("W9", "TRN-T01", "Admin training session 1: CRM structure + workflows + reports", "Phase 3", "Sayer", 3, "Recorded"),
    ("W9", "TRN-T02", "Admin training session 2: report building + user management", "Phase 3", "Sayer", 2, "Recorded"),
    ("W9", "TRN-T03", "End-user sales training (pipeline, quoting, dashboard usage)", "Phase 3", "Sayer", 3, "Recorded + live Q&A"),
    ("W9", "TRN-T04", "End-user ops training (onboarding, ClickUp sync, handoffs)", "Phase 3", "Sayer", 2, "Recorded"),
    ("W9", "TRN-T05", "Admin playbook + end-user quick-reference cards", "Phase 3", "Sayer", 2, "—"),

    # W10 — 10 hrs
    ("W10", "UAT-T01", "Test scenario design (top 10 sales, 5 ops, 5 reporting)", "Phase 3", "Sayer", 3, "—"),
    ("W10", "UAT-T02", "Test matrix execution + bug triage", "Phase 3", "Joint", 3, "—"),
    ("W10", "UAT-T03", "UAT facilitation with sales + ops users", "Phase 3", "Sayer", 1, "—"),
    ("W10", "UAT-T04", "Go-live cutover + communications plan", "Phase 3", "Joint", 1, "Single cutover event; phased = change order"),
    ("W10", "UAT-T05", "30-day hypercare (bandwidth allocation)", "Phase 3", "Sayer", 2, "Bugs + clarifications only"),

    # PM — 16 hrs
    ("PM", "PM-T01", "Weekly standup facilitation × 12 weeks", "Cross-phase", "Sayer", 6, "30 min/week"),
    ("PM", "PM-T02", "Status reporting + governance artifact maintenance", "Cross-phase", "Sayer", 3, "—"),
    ("PM", "PM-T03", "Stakeholder alignment + escalation handling", "Cross-phase", "Sayer", 3, "—"),
    ("PM", "PM-T04", "Phase reviews × 3 (kickoff/foundation, build, UAT/go-live)", "Cross-phase", "Sayer", 4, "—"),

    # Optional QuotaPath — 14 hrs
    ("QP*", "QP-T01", "HubSpot deal/property configuration to feed QuotaPath", "Optional", "Sayer", 4, "Pending Jeremy 5/8"),
    ("QP*", "QP-T02", "QuotaPath account setup + native integration install", "Optional", "Joint", 3, "HelloSpoke owns QuotaPath subscription"),
    ("QP*", "QP-T03", "Commission rule definition workshop", "Optional", "Joint", 3, "Sara/leadership input required"),
    ("QP*", "QP-T04", "Sales-engineer alignment session (per Cameron's caveat)", "Optional", "Joint", 2, "—"),
    ("QP*", "QP-T05", "Validation testing across 1 commission cycle", "Optional", "Sayer", 2, "—"),
]

for r, t in enumerate(tasks, start=2):
    for c, val in enumerate(t, start=1):
        ws.cell(row=r, column=c, value=val)
style_body(ws, 2, len(tasks) + 1, len(cols))

# Verify per-workstream task hours sum to median (sanity)
ws_sum_row = len(tasks) + 3
ws.cell(row=ws_sum_row, column=2, value="Per-workstream task-hour totals (verification):").font = Font(italic=True, bold=True)

verify_table = [
    ("W1", 8, "Architecture", "tasks sum"),
    ("W2", 7, "Data QA", "tasks sum"),
    ("W3", 5, "Sales Stages", "tasks sum"),
    ("W4", 9, "Quoting/CPQ", "tasks sum"),
    ("W5", 5, "DocuSign", "tasks sum (median path = 6 hrs)"),
    ("W6", 7, "Automations", "tasks sum"),
    ("W7", 5, "Dashboards", "tasks sum"),
    ("W8", 5, "ClickUp", "tasks sum"),
    ("W9", 5, "Training", "tasks sum"),
    ("W10", 5, "UAT/Go-Live", "tasks sum"),
    ("PM", 4, "PM", "tasks sum"),
    ("QP*", 5, "QuotaPath (optional)", "tasks sum"),
]
# We'll inject a cross-check based on the actual data
from collections import defaultdict
sums = defaultdict(int)
for t in tasks:
    sums[t[0]] += t[5]

note_row = ws_sum_row + 1
for wid, ntasks, label, _ in verify_table:
    ws.cell(row=note_row, column=2,
            value=f"{wid} ({label}): {ntasks} tasks, {sums[wid]} hrs sum").alignment = WRAP
    note_row += 1

set_widths(ws, [6, 12, 60, 10, 10, 14, 50])


# =========================================================================
# Tab 4 — Phase Matrix
# =========================================================================
ws = wb.create_sheet("Phase Matrix")
cols = ["Phase", "Weeks", "Workstream ID", "Workstream",
        "Median Hrs", "Description"]
ws.append(cols)
style_header(ws, 1, len(cols))

phase_matrix = [
    ("Phase 1", "Wk 1–4", "W1", "HubSpot Foundational Architecture", 18,
     "Objects, pipelines, lead status, lifecycle stages, properties, team/permissions"),
    ("Phase 1", "Wk 1–4", "W2", "Data QA + Pre-Production Audit", 18,
     "Sandbox review, data integrity, mapping validation, error handling + rollback plan, production-push memo"),
    ("Phase 1", "Wk 1–4", "W3", "Sales Stage Definitions + Gating", 10,
     "Workshop, stage build, required-field gating at 2 handoffs, stuck-deal report"),
    ("Phase 2", "Wk 5–9", "W4", "Quoting & CPQ Workflow (Commerce Hub)", 24,
     "Product library, CSV upload, quote templates, validation, approvals, native e-sig, quote-to-deal automation"),
    ("Phase 2", "Wk 5–9", "W5", "DocuSign Replacement Evaluation + Setup", 6,
     "Decision workshop; Commerce Hub native (default) or PandaDoc (alternate, +hrs)"),
    ("Phase 2", "Wk 5–9", "W6", "Automations & Workflows", 20,
     "Up to 12 workflows: deal-stage, lifecycle, customer-journey, ownership, alerts"),
    ("Phase 2", "Wk 5–9", "W7", "Reporting & Dashboards", 18,
     "3 dashboards (exec, sales, ops); custom views; report templates"),
    ("Phase 2", "Wk 5–9", "W8", "ClickUp Workflow Refinement", 12,
     "Workflow correctness, gating, status alignment (NOT integration build)"),
    ("Phase 3", "Wk 10–12", "W9", "Training & Enablement", 12,
     "Admin + end-user tracks, recorded; admin playbook + reference cards"),
    ("Phase 3", "Wk 10–12", "W10", "UAT & Go-Live", 10,
     "Test matrix, UAT facilitation, cutover, 30-day hypercare"),
    ("Cross-phase", "All 12 wks", "PM", "Project Management", 16,
     "Weekly standups, status reporting, phase reviews, escalation"),
    ("Optional", "TBD", "QP*", "QuotaPath Integration", 14,
     "Decision pending Jeremy 2026-05-08; HubSpot config + integration install + commission rules"),
]
for r, row in enumerate(phase_matrix, start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)
    if row[0] == "Optional":
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).fill = ACCENT_FILL
style_body(ws, 2, len(phase_matrix) + 1, len(cols), zebra=False)

set_widths(ws, [12, 11, 8, 40, 12, 70])


# =========================================================================
# Tab 5 — Phase Requirements
# =========================================================================
ws = wb.create_sheet("Phase Requirements")
cols = ["Phase", "WS", "Req ID", "Requirement", "Category", "Status"]
ws.append(cols)
style_header(ws, 1, len(cols))

reqs = [
    # Phase 1
    ("Phase 1", "W1", "ARC-01", "HubSpot tier validated; Commerce Hub Pro licensing path confirmed", "Architecture", "Pending"),
    ("Phase 1", "W1", "ARC-02", "Standard objects (Companies, Contacts, Deals) configured", "Architecture", "Pending"),
    ("Phase 1", "W1", "ARC-03", "Custom objects from Jeremy's sandbox reviewed and validated", "Architecture", "Pending"),
    ("Phase 1", "W1", "ARC-04", "Sales pipeline built with Sara's finalized stages", "Architecture", "Pending"),
    ("Phase 1", "W1", "ARC-05", "Lead status + lifecycle stage progression rules configured", "Architecture", "Pending"),
    ("Phase 1", "W1", "ARC-06", "Property strategy implemented (snake_case, group org)", "Architecture", "Pending"),
    ("Phase 1", "W1", "ARC-07", "Team + permissions configured (sales, ops, leadership)", "Architecture", "Pending"),
    ("Phase 1", "W1", "ARC-08", "Architecture decisions documented in admin playbook", "Documentation", "Pending"),
    ("Phase 1", "W2", "QA-01", "Sandbox access granted; initial inventory complete", "Data QA", "Pending"),
    ("Phase 1", "W2", "QA-02", "Data integrity audit complete (counts, completeness, dedupe, null rates)", "Data QA", "Pending"),
    ("Phase 1", "W2", "QA-03", "Object/property mapping validation complete", "Data QA", "Pending"),
    ("Phase 1", "W2", "QA-04", "Association validation complete (Contact↔Company↔Deal + custom)", "Data QA", "Pending"),
    ("Phase 1", "W2", "QA-05", "Custom-object structure review complete", "Data QA", "Pending"),
    ("Phase 1", "W2", "QA-06", "Error handling + rollback plan documented", "Data QA", "Pending"),
    ("Phase 1", "W2", "QA-07", "Production-push green-light memo delivered to Jeremy", "Data QA", "Pending"),
    ("Phase 1", "W3", "STG-01", "Sales-stage workshop with Sara + Christina complete", "Sales Process", "Pending"),
    ("Phase 1", "W3", "STG-02", "Stages built in HubSpot pipeline (probabilities + entry/exit criteria)", "Sales Process", "Pending"),
    ("Phase 1", "W3", "STG-03", "Required-field gating: sales → implementation (up to 10 fields)", "Governance", "Pending"),
    ("Phase 1", "W3", "STG-04", "Required-field gating: implementation → billing (up to 10 fields)", "Governance", "Pending"),
    ("Phase 1", "W3", "STG-05", "Stuck-deal HubSpot report live for leadership", "Reporting", "Pending"),
    # Phase 2
    ("Phase 2", "W4", "CPQ-01", "Commerce Hub Pro setup + Commerce Hub seat assignments", "CPQ", "Pending"),
    ("Phase 2", "W4", "CPQ-02", "Product library architecture configured (SKU, pricing, groups)", "CPQ", "Pending"),
    ("Phase 2", "W4", "CPQ-03", "CSV product upload pattern documented and validated", "CPQ", "Pending"),
    ("Phase 2", "W4", "CPQ-04", "Quote template configured with HelloSpoke branding", "CPQ", "Pending"),
    ("Phase 2", "W4", "CPQ-05", "Validation rules implemented for quotes above pricing thresholds", "CPQ", "Pending"),
    ("Phase 2", "W4", "CPQ-06", "Approval workflows configured", "CPQ", "Pending"),
    ("Phase 2", "W4", "CPQ-07", "Native e-signature configured on quotes (Dropbox Sign)", "CPQ", "Pending"),
    ("Phase 2", "W4", "CPQ-08", "Quote signature volume validated against Pro caps", "CPQ", "Pending"),
    ("Phase 2", "W4", "CPQ-09", "Quote-to-deal workflow automation configured", "CPQ", "Pending"),
    ("Phase 2", "W5", "SIG-01", "Decision-support workshop with Jeremy complete", "Signature", "Pending"),
    ("Phase 2", "W5", "SIG-02", "Replacement direction selected (Commerce Hub native OR PandaDoc)", "Signature", "Pending"),
    ("Phase 2", "W5", "SIG-03", "Selected platform configured (delegation, branded experience)", "Signature", "Pending"),
    ("Phase 2", "W5", "SIG-04", "Up to 5 templates migrated (PandaDoc path only)", "Signature", "Conditional"),
    ("Phase 2", "W6", "AUT-01", "Deal-stage notifications live", "Automation", "Pending"),
    ("Phase 2", "W6", "AUT-02", "Deal-stage task creation on advance", "Automation", "Pending"),
    ("Phase 2", "W6", "AUT-03", "Lifecycle progression workflows live", "Automation", "Pending"),
    ("Phase 2", "W6", "AUT-04", "Customer-journey automation live", "Automation", "Pending"),
    ("Phase 2", "W6", "AUT-05", "Sales ownership routing configured", "Automation", "Pending"),
    ("Phase 2", "W6", "AUT-06", "Stuck-deal alerts to managers", "Automation", "Pending"),
    ("Phase 2", "W6", "AUT-07..12", "Up to 12 workflows total (placeholder for additional)", "Automation", "Pending"),
    ("Phase 2", "W7", "DASH-01", "Executive dashboard live", "Reporting", "Pending"),
    ("Phase 2", "W7", "DASH-02", "Sales rep dashboard live", "Reporting", "Pending"),
    ("Phase 2", "W7", "DASH-03", "Ops dashboard live", "Reporting", "Pending"),
    ("Phase 2", "W7", "DASH-04..06", "Custom views + saved filters per role", "Reporting", "Pending"),
    ("Phase 2", "W7", "DASH-07", "Report templates configured", "Reporting", "Pending"),
    ("Phase 2", "W7", "DASH-08", "Scheduled-export configuration complete", "Reporting", "Pending"),
    ("Phase 2", "W8", "CLK-01", "Current state review of Jeremy-built integration complete", "ClickUp", "Pending"),
    ("Phase 2", "W8", "CLK-02", "Workflow correctness audit complete", "ClickUp", "Pending"),
    ("Phase 2", "W8", "CLK-03", "Status alignment confirmed (canonical statuses mirrored)", "ClickUp", "Pending"),
    ("Phase 2", "W8", "CLK-04", "Gating + handoff rules documented", "ClickUp", "Pending"),
    ("Phase 2", "W8", "CLK-05", "Who-does-what SOPs documented for implementation team", "ClickUp", "Pending"),
    # Phase 3
    ("Phase 3", "W9", "TRN-01", "Admin training delivered (2 sessions, recorded)", "Training", "Pending"),
    ("Phase 3", "W9", "TRN-02", "End-user sales training delivered (recorded + Q&A)", "Training", "Pending"),
    ("Phase 3", "W9", "TRN-03", "End-user ops/implementation training delivered", "Training", "Pending"),
    ("Phase 3", "W9", "TRN-04", "Admin playbook + quick-reference cards delivered", "Documentation", "Pending"),
    ("Phase 3", "W10", "UAT-01", "Test scenarios designed (top 10 sales, 5 ops, 5 reporting)", "UAT", "Pending"),
    ("Phase 3", "W10", "UAT-02", "Test matrix executed; bugs logged + triaged", "UAT", "Pending"),
    ("Phase 3", "W10", "UAT-03", "UAT facilitated with sales + ops users", "UAT", "Pending"),
    ("Phase 3", "W10", "UAT-04", "Go-live cutover complete; comms plan executed", "Go-Live", "Pending"),
    ("Phase 3", "W10", "UAT-05", "30-day hypercare initiated", "Hypercare", "Pending"),
    # Optional
    ("Optional", "QP*", "QP-01", "HubSpot deal/property configuration for QuotaPath data feed", "QuotaPath", "Pending decision"),
    ("Optional", "QP*", "QP-02", "Commission rule workshop + sales-engineer alignment", "QuotaPath", "Pending decision"),
    ("Optional", "QP*", "QP-03", "Validation testing across 1 commission cycle", "QuotaPath", "Pending decision"),
]
for r, row in enumerate(reqs, start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)
style_body(ws, 2, len(reqs) + 1, len(cols))
set_widths(ws, [10, 6, 12, 70, 16, 16])


# =========================================================================
# Tab 6 — Phase Success Criteria
# =========================================================================
ws = wb.create_sheet("Phase Success Criteria")
cols = ["Phase", "#", "Success Criterion"]
ws.append(cols)
style_header(ws, 1, len(cols))

criteria = [
    ("Phase 1 — Foundation + Audit (Wk 1–4)", 1, "HubSpot tier validated; Commerce Hub Pro licensing path confirmed"),
    ("Phase 1 — Foundation + Audit (Wk 1–4)", 2, "Foundational architecture configured (objects, pipelines, lead status, lifecycle, properties)"),
    ("Phase 1 — Foundation + Audit (Wk 1–4)", 3, "Sandbox audited; production-push green-light memo delivered to Jeremy"),
    ("Phase 1 — Foundation + Audit (Wk 1–4)", 4, "Sales pipeline live with stages + required-field gating + stuck-deal report"),
    ("Phase 1 — Foundation + Audit (Wk 1–4)", 5, "Architecture decisions documented in admin playbook"),
    ("Phase 2 — Build (Wk 5–9)", 1, "Commerce Hub CPQ live with quote templates, approvals, native e-signature on quotes"),
    ("Phase 2 — Build (Wk 5–9)", 2, "DocuSign replacement live (Commerce Hub native OR PandaDoc per Jeremy's decision)"),
    ("Phase 2 — Build (Wk 5–9)", 3, "12 automations live across deal-stage, lifecycle, and customer-journey workflows"),
    ("Phase 2 — Build (Wk 5–9)", 4, "3 dashboards live (executive, sales, ops); custom views per role"),
    ("Phase 2 — Build (Wk 5–9)", 5, "ClickUp workflow refinement complete; status sync verified end-to-end"),
    ("Phase 3 — Enable + Go-Live (Wk 10–12)", 1, "Admin + end-user training delivered and recorded"),
    ("Phase 3 — Enable + Go-Live (Wk 10–12)", 2, "UAT executed; bugs triaged and resolved"),
    ("Phase 3 — Enable + Go-Live (Wk 10–12)", 3, "Go-live cutover complete; communications plan executed"),
    ("Phase 3 — Enable + Go-Live (Wk 10–12)", 4, "30-day hypercare initiated"),
]
for r, row in enumerate(criteria, start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)
style_body(ws, 2, len(criteria) + 1, len(cols))
set_widths(ws, [38, 6, 90])


# =========================================================================
# Tab 7 — Risk Register
# =========================================================================
ws = wb.create_sheet("Risk Register")
cols = ["#", "Risk", "Severity", "Likelihood", "Impact",
        "Mitigation", "Owner", "Status"]
ws.append(cols)
style_header(ws, 1, len(cols))

risks = [
    (1, "HubSpot tier insufficient for Commerce Hub CPQ + native signature; upgrade to Commerce Hub Pro may be required",
     "High", "Medium",
     "W4 blocks; client-side licensing decision required mid-delivery",
     "Validate tier in W1 kickoff; surface upgrade decision before W4 starts",
     "Sayer", "Open"),
    (2, "Sandbox access delay — Jeremy must grant Sayer access to HubSpot sandbox before W2 begins",
     "High", "Medium",
     "W2 delayed; Phase 1 timeline slips",
     "Open item flagged 2026-05-04; access requested in v3 follow-up email",
     "HelloSpoke", "Open"),
    (3, "Pre-existing data quality issues in Jeremy-built migration surface during W2 audit",
     "Medium", "High",
     "Remediation may exceed 2 audit rounds; further rounds are change orders",
     "W2 limited to 2 audit rounds; surface remediation needs early; client decides remediation scope",
     "HelloSpoke", "Open"),
    (4, "Native signature volume exceeds Commerce Hub Pro caps (25 sigs/user/mo, pooled)",
     "Medium", "Low",
     "Forces Enterprise upgrade or PandaDoc fallback for non-quote docs",
     "Volume validation in W1; tier decision before W4; PandaDoc as fallback if needed",
     "HelloSpoke", "Open"),
    (5, "DocuSign replacement decision delayed — Jeremy hasn't confirmed Commerce Hub native vs PandaDoc",
     "Medium", "Medium",
     "W5 hours land at high end (PandaDoc) vs low end (Commerce Hub native)",
     "Decision deadline aligned with Phase 1 kickoff; v3 hours sized at median assuming Commerce Hub native",
     "HelloSpoke", "Open"),
    (6, "Sales-stage definitions delivered late — Sara is delivering this week post-validation; W3 build depends on it",
     "Medium", "Low",
     "W3 starts late, ripples to W6 dependencies",
     "Sara delivers in week 1 of delivery; if late, W3 slides to Phase 1 end",
     "HelloSpoke", "Open"),
    (7, "Phase 2 scope creep — Christina noted future ideas will surface 'once we live with the new system'",
     "Medium", "Medium",
     "New asks in mid-delivery erode v3 scope",
     "v3 explicitly defers Phase 2 ideas to a follow-on engagement; documented in scope",
     "Sayer", "Open"),
    (8, "QuotaPath decision changes scope mid-engagement",
     "Low", "Medium",
     "If Jeremy decides IN after v3 sends, +14 hrs change order",
     "Surface as explicit checkbox with EOW 2026-05-08 deadline; price impact transparent",
     "HelloSpoke", "Open"),
    (9, "Tier-priced products incompatible with bulk CSV import (HubSpot platform constraint)",
     "Low", "Low",
     "If HelloSpoke uses tier pricing, products created manually",
     "W4 inventory captures tier-pricing structure; if present, manual creation hours added",
     "Sayer", "Open"),
    (10, "Tight 12-week timeline with Phase 2 carrying 80 hrs",
     "Medium", "Medium",
     "Phase 2 risks running long if dependencies slip",
     "Weekly standups; phase-review checkpoints; early flag if Phase 1 deliverables run late",
     "Sayer", "Open"),
]
for r, row in enumerate(risks, start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)
style_body(ws, 2, len(risks) + 1, len(cols))
set_widths(ws, [4, 50, 10, 12, 40, 50, 12, 8])


# =========================================================================
# Tab 8 — Scope Assumptions
# =========================================================================
ws = wb.create_sheet("Scope Assumptions")
cols = ["#", "Assumption"]
ws.append(cols)
style_header(ws, 1, len(cols))

assumptions = [
    "HubSpot tier: Sales Hub Professional minimum; Commerce Hub Professional minimum for W4. Tier validation in W1; upgrade decisions are client-side licensing matters",
    "Sandbox access granted by Jeremy before W2 begins; sandbox contains current Salesforce + Rev IO data",
    "Jeremy completes Rev IO integration, basic ClickUp deal-won automation, Salesforce data migration, ALN integration. Sayer does NOT build these",
    "QuickBooks integration is OUT (Jeremy declined; financial data lives in Rev IO)",
    "Sara delivers finalized sales-stage definitions in week 1 of delivery",
    "Up to 2 rounds of W2 audit + remediation findings; further rounds are change orders",
    "Up to 10 required fields per gate, 2 gates total in W3",
    "Up to 12 workflows in W6; further or higher-complexity workflows are change orders",
    "Up to 3 dashboards (executive, sales, ops), 5–8 reports per dashboard in W7",
    "Up to 2 admin training sessions + 2 end-user sessions per team in W9",
    "30-day hypercare in W10 covers bugs and clarifications, not net-new feature requests",
    "No multi-currency quotes in W4",
    "No marketing email nurture flows (out of scope; covered in future engagement per 4/14)",
    "Phase 2 ideas deliberately deferred to a follow-on engagement (Christina's ask, 5/4)",
    "Rate: $175/hr; scope changes require written approval and may impact cost or timeline",
    "All work performed remotely unless otherwise agreed",
    "HelloSpoke maintains required Sales Hub + Commerce Hub licensing throughout engagement",
    "Salesforce contract expiration (October 2026) is the hard deadline for migration completion",
]
for r, a in enumerate(assumptions, start=2):
    ws.cell(row=r, column=1, value=r - 1)
    ws.cell(row=r, column=2, value=a)
style_body(ws, 2, len(assumptions) + 1, len(cols))
set_widths(ws, [4, 110])


# =========================================================================
# Tab 9 — Out of Scope
# =========================================================================
ws = wb.create_sheet("Out of Scope")
cols = ["Item", "Reason"]
ws.append(cols)
style_header(ws, 1, len(cols))

oos = [
    ("Rev IO integration build", "Jeremy completed (already syncing every 30 min)"),
    ("ClickUp integration build (deal-won automation, native integration setup, Make middleware)", "Jeremy completed"),
    ("Salesforce → HubSpot data migration build", "Jeremy completed; Sayer audits in W2 only"),
    ("QuickBooks integration", "Jeremy declined; financial data lives in Rev IO"),
    ("ALN integration", "Jeremy completing himself this week"),
    ("Make middleware", "Eliminated with stripped ClickUp integration build"),
    ("Legacy ClickUp ↔ Salesforce integration", "Explicitly out per 2026-04-23; Sayer not responsible for legacy plumbing"),
    ("Marketing email nurture flows", "Out per 2026-04-14"),
    ("Multi-currency quotes", "Out"),
    ("External BI tool integration for reporting", "Out"),
    ("Net-new HubSpot custom-object schema design", "Sayer validates Jeremy's existing schemas; doesn't design from scratch"),
    ("10DLC training flow", "Tracked separately per 2026-04-23"),
    ("Win-property separate survey flow", "Out per 2026-04-23 (80/20 standard flow handles)"),
    ("Phase 2 ideas surfaced post-go-live", "Deferred to follow-on engagement"),
    ("HubSpot Contracts (general contract signing)", "Spring 2026 beta; quote-tied; not relied on in v3"),
]
for r, (item, reason) in enumerate(oos, start=2):
    ws.cell(row=r, column=1, value=item)
    ws.cell(row=r, column=2, value=reason)
style_body(ws, 2, len(oos) + 1, len(cols))
set_widths(ws, [60, 80])


# =========================================================================
# Tab 10 — Open Items
# =========================================================================
ws = wb.create_sheet("Open Items")
cols = ["Owner", "Item", "Due / Status", "Type"]
ws.append(cols)
style_header(ws, 1, len(cols))

open_items = [
    ("Sayer (Cameron)", "Pricing alignment on $28,700 median before v3 ships to Jeremy", "Pre-send blocker", "Internal"),
    ("Sayer (Kyle)", "Manually apply v3 changes to authoritative SOW Google Sheet", "Pre-send blocker", "Internal"),
    ("Sayer (Kyle)", "Mirror authoritative changes into working copy SOW", "After authoritative apply", "Internal"),
    ("Sayer (Kyle)", "Generate new Gamma deck to supersede erumqig7pyx7rhm", "Pre-send blocker", "Internal"),
    ("Sayer (Kyle)", "Draft + send v3 follow-up email from kyle@gosayer.com", "Final step", "Internal"),
    ("Sayer (Kyle)", "Add v3 Project Baseline entry to calibration/calibration.md", "After v3 lands", "Internal"),
    ("Jeremy Wiley", "Answer on QuotaPath in/out", "EOW 2026-05-08", "Client decision"),
    ("Jeremy Wiley", "Confirm DocuSign replacement direction (Commerce Hub native vs PandaDoc)", "Before W5 begins", "Client decision"),
    ("Jeremy Wiley", "Grant Sayer access to HubSpot sandbox (Salesforce + Rev IO data)", "Before W2 begins", "Client deliverable"),
    ("Jeremy Wiley", "Confirm Commerce Hub Professional licensing (or upgrade decision)", "Before W4 begins", "Client decision"),
    ("Sara Hines", "Send finalized sales-stage definitions (post sales-team validation)", "Week 1 of delivery", "Client deliverable"),
    ("Christina Edwards", "Confirm sales-stage definitions align with current ClickUp setup; Phase 2 ideas deferred", "Week 1 of delivery", "Client confirmation"),
    ("Jeremy Wiley", "Finalize ALN + QuickBooks integrations himself this week (informational)", "Self-paced", "Client work (out of Sayer scope)"),
]
for r, row in enumerate(open_items, start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)
style_body(ws, 2, len(open_items) + 1, len(cols))
set_widths(ws, [22, 70, 22, 22])


# =========================================================================
# Save
# =========================================================================
out_path = "/Users/kyleharbuck.sayer/projects/project-scoping-tool/HelloSpoke/HelloSpoke_v3_SOW_Project_Sheet.xlsx"
wb.save(out_path)
print(f"Wrote {out_path}")

# Verification
print("\n=== VERIFICATION ===")
print(f"Tabs ({len(wb.sheetnames)}):")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {name}")

import os
print(f"\nFile size: {os.path.getsize(out_path):,} bytes")

# Hours sanity check from task table
total_task_hrs = sum(t[5] for t in tasks if t[0] != "QP*")
print(f"\nBase task-hour sum (W1..W10 + PM): {total_task_hrs} hrs")
print(f"Expected (median): 164 hrs")
print(f"Match: {total_task_hrs == 164}")

qp_hrs = sum(t[5] for t in tasks if t[0] == "QP*")
print(f"\nOptional QuotaPath task hours: {qp_hrs}")
print(f"Expected: 14")
print(f"Match: {qp_hrs == 14}")
