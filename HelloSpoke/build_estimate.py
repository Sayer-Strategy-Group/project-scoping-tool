#!/usr/bin/env python3
"""Generate HelloSpoke_Scoping_Estimate.xlsx with 5 sheets.

Sheets:
  1. Scoping Estimate — workstream hours with formula-driven rate cell
  2. Phase Detail — requirements per phase with milestones and success criteria
  3. Risk Register — 10 risks with severity color coding
  4. Assumptions & Exclusions — scope boundaries
  5. Approach Comparison — full vs reduced scope

Usage:
    python3 HelloSpoke/build_estimate.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
PHASE_FONT = Font(name="Arial", size=10, bold=True, color="1F4E79")
PHASE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
INPUT_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
ACTUALS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CURRENCY_FMT = "$#,##0"


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if (r - start_row) % 2 == 1:
                if not cell.fill or cell.fill.start_color.index == "00000000":
                    cell.fill = ALT_FILL


def auto_width(ws, max_col, padding=3):
    for col in range(1, max_col + 1):
        max_len = 0
        letter = get_column_letter(col)
        for cell in ws[letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + padding, 50)


# ---------------------------------------------------------------------------
# Sheet 1: Scoping Estimate (formula-driven)
# ---------------------------------------------------------------------------
def sheet_estimate(wb):
    ws = wb.active
    ws.title = "Scoping Estimate"

    # Rate input cell
    ws.cell(row=1, column=1, value="Hourly Rate:").font = BOLD_FONT
    rate_cell = ws.cell(row=1, column=2, value=150)
    rate_cell.number_format = CURRENCY_FMT
    rate_cell.fill = INPUT_FILL
    rate_cell.font = BOLD_FONT
    rate_cell.border = THIN_BORDER
    ws.cell(row=1, column=3, value="(adjust this cell to recalculate all costs)").font = Font(
        name="Arial", size=9, italic=True, color="666666"
    )

    # Headers at row 3
    headers = [
        "Phase", "Workstream", "Description",
        "Min Hours", "Max Hours", "Median Hours",
        "Min Cost", "Max Cost", "Median Cost",
        "Notes / Assumptions",
        "Actual Hours", "Actual Cost",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))

    # Workstream data: (phase, name, desc, min, max, med, notes)
    workstreams = [
        ("Phase 1", "CRM Architecture & Setup",
         "Audit existing instance, user roles, permissions, branding",
         8, 18, 12, "Existing HubSpot instance reconfiguration for 15-18 users"),
        ("Phase 1", "Pipeline & Deal Management",
         "Sales funnel redesign, stages, deal scoring, lead definitions",
         8, 18, 12, "Full redesign, not lift-and-shift from Salesforce"),
        ("Phase 1", "Contact & Company Management",
         "Custom props, multifamily objects, parent-child, dedup strategy",
         12, 24, 16, "Up to 25 custom properties, 2 custom objects"),
        ("Phase 2", "Salesforce Data Migration",
         "Phased migration: inactive -> active -> leads, audit + validate",
         16, 36, 24, "<50K records assumed; phased with validation between loads"),
        ("Phase 3", "Rev IO Integration",
         "REST API, custom object, billing->deal sync, webhooks",
         24, 44, 32, "First-time; telecom billing API; sandbox + production phases"),
        ("Phase 3", "QuickBooks Integration",
         "Revenue actuals sync, OAuth2 setup, reconciliation logic",
         12, 32, 20, "Native integration insufficient; custom build for revenue actuals"),
        ("Phase 3", "CPQ & Quoting (HubSpot CPQ native)",
         "Product library, dual dynamic templates, bulk import (up to 300 lines), parent-child auto-creation w/ dedup, native e-sig, close-won sync",
         28, 52, 38, "HubSpot CPQ native primary (AI CPQ per Billy's direction); PandaDoc fallback as change order; DocuSign sunset"),
        ("Phase 3", "QuotaPath Coordination",
         "HubSpot-side data flow: deal props, custom object feed, field mapping",
         6, 14, 10, "QuotaPath handles own implementation; Sayer handles HubSpot-side config"),
        ("Phase 4", "Automations & Workflows",
         "Lead routing, stage triggers, follow-up automation, notifications",
         10, 22, 16, "Up to 10 workflows"),
        ("Phase 4", "Reporting & Dashboards",
         "Executive dashboards, pipeline, forecasting, actuals vs targets",
         10, 22, 16, "Up to 2 dashboards with 10+ reports each"),
        ("Phase 4", "Training & Adoption",
         "3 sessions: admin, sales reps, ops/bookkeeping",
         10, 24, 16, "3 x 1-hour virtual sessions, recorded"),
        ("Cross", "Project Management",
         "Kickoff, weekly syncs, status tracking, change management",
         14, 40, 22, "~12% of total; weekly for 14 weeks"),
    ]

    for i, (phase, name, desc, mn, mx, med, notes) in enumerate(workstreams, 4):
        row = i
        ws.cell(row=row, column=1, value=phase)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=desc)
        ws.cell(row=row, column=4, value=mn)
        ws.cell(row=row, column=5, value=mx)
        ws.cell(row=row, column=6, value=med)
        # Cost formulas referencing rate cell $B$1
        ws.cell(row=row, column=7).value = f"=D{row}*$B$1"
        ws.cell(row=row, column=7).number_format = CURRENCY_FMT
        ws.cell(row=row, column=8).value = f"=E{row}*$B$1"
        ws.cell(row=row, column=8).number_format = CURRENCY_FMT
        ws.cell(row=row, column=9).value = f"=F{row}*$B$1"
        ws.cell(row=row, column=9).number_format = CURRENCY_FMT
        ws.cell(row=row, column=10, value=notes)
        # Actuals columns with green fill
        ws.cell(row=row, column=11).fill = ACTUALS_FILL
        ws.cell(row=row, column=12).fill = ACTUALS_FILL
        ws.cell(row=row, column=12).number_format = CURRENCY_FMT

    # Totals row
    total_row = len(workstreams) + 4
    ws.cell(row=total_row, column=1, value="").font = BOLD_FONT
    ws.cell(row=total_row, column=2, value="TOTAL").font = BOLD_FONT
    first_data = 4
    last_data = total_row - 1
    for col_letter, col_num in [("D", 4), ("E", 5), ("F", 6)]:
        ws.cell(row=total_row, column=col_num).value = (
            f"=SUM({col_letter}{first_data}:{col_letter}{last_data})"
        )
        ws.cell(row=total_row, column=col_num).font = BOLD_FONT
    for col_letter, col_num in [("G", 7), ("H", 8), ("I", 9)]:
        ws.cell(row=total_row, column=col_num).value = (
            f"=SUM({col_letter}{first_data}:{col_letter}{last_data})"
        )
        ws.cell(row=total_row, column=col_num).number_format = CURRENCY_FMT
        ws.cell(row=total_row, column=col_num).font = BOLD_FONT
    # Actuals total
    ws.cell(row=total_row, column=11).value = (
        f"=SUM(K{first_data}:K{last_data})"
    )
    ws.cell(row=total_row, column=11).font = BOLD_FONT
    ws.cell(row=total_row, column=11).fill = ACTUALS_FILL
    ws.cell(row=total_row, column=12).value = (
        f"=SUM(L{first_data}:L{last_data})"
    )
    ws.cell(row=total_row, column=12).font = BOLD_FONT
    ws.cell(row=total_row, column=12).number_format = CURRENCY_FMT
    ws.cell(row=total_row, column=12).fill = ACTUALS_FILL

    style_body(ws, 4, total_row, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "C4"


# ---------------------------------------------------------------------------
# Sheet 2: Phase Detail (new — requirements + milestones per phase)
# ---------------------------------------------------------------------------
def sheet_phase_detail(wb):
    ws = wb.create_sheet("Phase Detail")

    headers = ["Phase", "Weeks", "Req ID", "Requirement", "Category",
               "Success Criterion", "Status"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    phases = [
        {
            "phase": "Phase 1: CRM Foundation",
            "weeks": "Wk 1-4",
            "reqs": [
                ("CRM-01", "Existing HubSpot instance audited and cleaned of legacy configuration", "CRM Setup"),
                ("CRM-02", "User roles and permissions configured for 15-18 users", "CRM Setup"),
                ("CRM-03", "Team structure and ownership rules established", "CRM Setup"),
                ("CRM-04", "Sales pipeline redesigned with stages, entry/exit criteria, required fields", "Pipeline"),
                ("CRM-05", "Deal scoring model configured", "Pipeline"),
                ("CRM-06", "Lead vs. deal vs. account definitions standardized", "Pipeline"),
                ("CRM-07", "Contact and company properties configured for multifamily (up to 25)", "Data Model"),
                ("CRM-08", "Parent-child company associations built", "Data Model"),
                ("CRM-09", "Custom objects created for Rev IO billing and multifamily (up to 2)", "Data Model"),
                ("CRM-10", "Deduplication strategy defined and merge rules established", "Data Model"),
                ("PM-01", "Kickoff meeting completed with timeline and cadence aligned", "PM"),
                ("PM-02", "Weekly status syncs established", "PM"),
                ("PM-03", "Project tracker live with workstream status", "PM"),
            ],
            "criteria": [
                "All 15-18 users provisioned with correct roles",
                "Sales pipeline has defined stages with entry/exit criteria",
                "Custom objects configured with multifamily properties and parent-child",
                "Dedup strategy documented before any data imports",
                "Kickoff complete, weekly sync running, project tracker live",
            ],
        },
        {
            "phase": "Phase 2: Salesforce Migration",
            "weeks": "Wk 3-6",
            "reqs": [
                ("MIG-01", "Salesforce data audit completed — objects, counts, quality", "Audit"),
                ("MIG-02", "Field mapping document created (up to 50 mappings)", "Mapping"),
                ("MIG-03", "Phase A: inactive customers and closed deals imported + validated", "Migration"),
                ("MIG-04", "Phase B: active pipeline imported + validated", "Migration"),
                ("MIG-05", "Phase C: leads and prospects imported + validated", "Migration"),
                ("MIG-06", "Post-migration integrity verified — counts match, no data loss", "Validation"),
                ("MIG-07", "Outdated/low-quality records archived", "Cleanup"),
                ("PM-04", "Scope change process in place for discovered gaps", "PM"),
            ],
            "criteria": [
                "Field mapping document exists before any data moves",
                "Inactive data imported and verified — counts match SF",
                "Active pipeline imported with correct stages and owners",
                "Leads imported with quality flags resolved or archived",
                "Integrity check confirms no data loss",
            ],
        },
        {
            "phase": "Phase 3: Integration & Quoting Build",
            "weeks": "Wk 5-11",
            "reqs": [
                ("INT-01", "Rev IO REST API connection established with auth and error handling", "Rev IO"),
                ("INT-02", "HubSpot custom object for Rev IO billing created and associated with deals", "Rev IO"),
                ("INT-03", "Rev IO data syncing to HubSpot (accounts, invoices, payments, usage)", "Rev IO"),
                ("INT-04", "Webhook or batch sync operational for ongoing Rev IO flow", "Rev IO"),
                ("INT-05", "QuickBooks OAuth2 authentication configured with token refresh", "QuickBooks"),
                ("INT-06", "Revenue actuals from QB paid invoices syncing to HubSpot deals", "QuickBooks"),
                ("INT-07", "Revenue reconciliation logic defined — source-of-truth rules documented", "QuickBooks"),
                ("INT-08", "QuotaPath HubSpot-side data flow configured", "QuotaPath"),
                ("INT-09", "End-to-end chain tested: Rev IO -> HubSpot -> QuotaPath", "QuotaPath"),
                ("INT-10", "Sync monitoring and error alerting established for all integrations", "Monitoring"),
                ("CPQ-01", "HubSpot CPQ product library configured (up to 50 SKUs: VoIP + Notify + equipment)", "CPQ"),
                ("CPQ-02", "Dual dynamic quote templates (VoIP T&C, Notify/SaaS T&C) plus combined template", "CPQ"),
                ("CPQ-03", "Bulk line-item import for property management quotes (up to 300 lines w/ validation)", "CPQ"),
                ("CPQ-04", "Parent-child property auto-creation from quote lines with dedup logic", "CPQ"),
                ("CPQ-05", "Quote approval workflow with up to 2 tiers (rep self-serve + non-standard routing)", "CPQ"),
                ("CPQ-06", "HubSpot native e-signature live, DocuSign retired, templates migrated (MSA/NDA/SOW)", "CPQ"),
                ("CPQ-07", "Close-won automation: quoted amount syncs to deal amount; close date syncs to signature date", "CPQ"),
                ("CPQ-08", "Closed-won quote handoff to Rev IO billing record creation configured", "CPQ"),
                ("PM-05", "Cross-vendor coordination completed", "PM"),
            ],
            "criteria": [
                "Rev IO billing data syncing to HubSpot custom objects via API",
                "QB revenue actuals appearing on HubSpot deals with reconciliation rules",
                "Reps can build standard quotes end-to-end without admin assistance; bulk imports up to 300 lines succeed",
                "HubSpot native e-sig live, DocuSign retired, close-won auto-sync operational",
                "QuotaPath pulling commission data — full chain tested",
                "All integrations have error alerting and monitoring",
                "Cross-vendor coordination documented",
            ],
        },
        {
            "phase": "Phase 4: Enablement & Go-Live",
            "weeks": "Wk 9-14",
            "reqs": [
                ("ENB-01", "Up to 10 automated workflows deployed", "Automation"),
                ("ENB-02", "Executive dashboard built — pipeline, forecasting, revenue", "Reporting"),
                ("ENB-03", "Sales performance dashboard built — activity, conversion, velocity", "Reporting"),
                ("ENB-04", "Admin training delivered and recorded (Christina + ops)", "Training"),
                ("ENB-05", "Sales rep training delivered and recorded", "Training"),
                ("ENB-06", "Ops/bookkeeping training delivered and recorded", "Training"),
                ("ENB-07", "Workflow quick-reference guides provided per role", "Training"),
                ("PM-06", "Project handoff documentation delivered", "PM"),
            ],
            "criteria": [
                "Up to 10 workflows live and firing correctly",
                "Dashboards published with pipeline, forecasting, actuals vs targets",
                "All 3 user groups completed recorded training",
                "Christina can manage system without Sayer; role guides delivered",
                "Handoff docs delivered, engagement closed",
            ],
        },
    ]

    row = 2
    for phase_data in phases:
        # Phase header row
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = PHASE_FILL
            ws.cell(row=row, column=c).font = PHASE_FONT
        ws.cell(row=row, column=1, value=phase_data["phase"])
        ws.cell(row=row, column=2, value=phase_data["weeks"])
        row += 1

        # Requirements
        for req_id, req_desc, category in phase_data["reqs"]:
            ws.cell(row=row, column=1, value="")
            ws.cell(row=row, column=2, value="")
            ws.cell(row=row, column=3, value=req_id)
            ws.cell(row=row, column=4, value=req_desc)
            ws.cell(row=row, column=5, value=category)
            ws.cell(row=row, column=6, value="")
            ws.cell(row=row, column=7, value="Pending")
            row += 1

        # Success criteria (on separate rows after requirements)
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=3, value="").font = BOLD_FONT
        ws.cell(row=row, column=4, value="Success Criteria:").font = BOLD_FONT
        row += 1
        for i, criterion in enumerate(phase_data["criteria"], 1):
            ws.cell(row=row, column=4, value=f"{i}. {criterion}")
            ws.cell(row=row, column=6, value=criterion)
            row += 1

        row += 1  # blank row between phases

    style_body(ws, 2, row - 1, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "D2"


# ---------------------------------------------------------------------------
# Sheet 3: Risk Register
# ---------------------------------------------------------------------------
def sheet_risks(wb):
    ws = wb.create_sheet("Risk Register")

    headers = ["#", "Risk", "Severity", "Likelihood", "Impact",
               "Mitigation", "Owner", "Status"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    risks = [
        (1, "Salesforce data quality worse than expected", "High", "Medium",
         "Migration hours exceed estimate; duplicates propagate",
         "SF data audit in Phase 2 before migration; phased with validation",
         "Sayer + HelloSpoke", "Open"),
        (2, "Rev IO API limitations or poor documentation", "High", "Medium",
         "Integration scope increases; custom workarounds needed",
         "API investigation early in Phase 3; sandbox testing first",
         "Sayer", "Open"),
        (3, "Existing HubSpot instance has legacy config", "Medium", "Medium",
         "Cleanup adds hours to Phase 1",
         "Instance audit in Week 1; flag conflicts early",
         "Sayer", "Open"),
        (4, "HubSpot AI CPQ features unstable or missing functionality", "High", "Medium",
         "CPQ scope shifts to PandaDoc change order; Phase 3 delays",
         "Validate CPQ capabilities in Phase 3 Week 1 before build; PandaDoc fallback scoped",
         "Sayer + HelloSpoke", "Open"),
        (5, "QuotaPath deal falls through", "Medium", "Low",
         "Commission tracking deferred; HubSpot-side config repurposed for internal reporting",
         "HubSpot data flow architecture delivers value regardless",
         "HelloSpoke", "Open"),
        (6, "Stakeholder availability delays", "Medium", "Medium",
         "Pipeline design, quoting workshop, and training delayed",
         "Weekly cadence at kickoff; Christina as dedicated liaison",
         "HelloSpoke", "Open"),
        (7, "QuickBooks revenue reconciliation complexity", "Medium", "Medium",
         "Discrepancies between QB and HubSpot figures",
         "Define source-of-truth rules in Phase 3; reconciliation dashboard",
         "Sayer", "Open"),
        (8, "Bulk quote import edge cases (300+ line items)", "Medium", "High",
         "CPQ import fails for property management deals in production",
         "Early validation harness with HelloSpoke's actual CSVs; clear error UX reviewed with Ryan",
         "Sayer", "Open"),
        (9, "Salesforce contract timeline pressure", "Low", "Low",
         "October deadline creates urgency if project slips",
         "14-week timeline targets August; 2+ month buffer",
         "HelloSpoke", "Open"),
        (10, "Sales team adoption resistance", "Medium", "Medium",
         "Reps revert to old workflows or inconsistent CRM usage",
         "Role-tailored training in Phase 4; simplified data entry; deal scoring",
         "HelloSpoke + Sayer", "Open"),
        (11, "Scope creep from adjacent systems", "Medium", "High",
         "Elastic/ALN, Campfire, Zendesk get added mid-project",
         "Explicit exclusions documented; formal change order process",
         "Both", "Open"),
        (12, "Rev IO sandbox webhook limitation", "Low", "High",
         "Cannot test real-time billing events until production",
         "Plan production cutover in Phase 3; batch sync as interim",
         "Sayer", "Open"),
    ]

    severity_fills = {"High": RED_FILL, "Medium": YELLOW_FILL, "Low": GREEN_FILL}

    for i, (num, risk, sev, lik, impact, mit, owner, status) in enumerate(risks, 2):
        ws.cell(row=i, column=1, value=num)
        ws.cell(row=i, column=2, value=risk)
        ws.cell(row=i, column=3, value=sev)
        ws.cell(row=i, column=3).fill = severity_fills.get(sev, PatternFill())
        ws.cell(row=i, column=4, value=lik)
        ws.cell(row=i, column=5, value=impact)
        ws.cell(row=i, column=6, value=mit)
        ws.cell(row=i, column=7, value=owner)
        ws.cell(row=i, column=8, value=status)

    style_body(ws, 2, len(risks) + 1, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "C2"


# ---------------------------------------------------------------------------
# Sheet 4: Assumptions & Exclusions
# ---------------------------------------------------------------------------
def sheet_assumptions(wb):
    ws = wb.create_sheet("Assumptions & Exclusions")

    headers = ["Section", "Item"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 2)

    data = [
        ("Scope Assumptions", "HubSpot Sales Hub Professional (or Commerce Hub if required for AI CPQ) with 15-18 paid seats"),
        ("Scope Assumptions", "HubSpot AI CPQ features enabled and production-stable by start of Phase 3"),
        ("Scope Assumptions", "DocuSign sunset by HelloSpoke; HubSpot native e-sig replaces it (no parallel run)"),
        ("Scope Assumptions", "Salesforce data volume under 50,000 records total"),
        ("Scope Assumptions", "Rev IO REST API is accessible with documented endpoints"),
        ("Scope Assumptions", "QuotaPath proceeds with their own implementation in parallel"),
        ("Scope Assumptions", "QuickBooks Online admin access and API credentials provided"),
        ("Scope Assumptions", "Salesforce admin access provided for data audit and export"),
        ("Scope Assumptions", "Up to 25 custom contact/company properties"),
        ("Scope Assumptions", "Up to 2 custom objects (Rev IO billing + multifamily)"),
        ("Scope Assumptions", "Up to 50 field mappings for Salesforce migration"),
        ("Scope Assumptions", "Up to 30 field mappings for Rev IO integration"),
        ("Scope Assumptions", "Up to 50 products in HubSpot CPQ product library"),
        ("Scope Assumptions", "Up to 2 dynamic quote templates (VoIP, Notify SaaS) + 1 combined"),
        ("Scope Assumptions", "Up to 2 quote approval tiers (rep self-serve + Jeremy override)"),
        ("Scope Assumptions", "Bulk quote import supports up to 300 line items per quote"),
        ("Scope Assumptions", "Up to 10 automated workflows"),
        ("Scope Assumptions", "Up to 2 dashboards with 10+ reports each"),
        ("Scope Assumptions", "All work is remote unless otherwise agreed"),
        ("Out of Scope", "Elastic / ALN database integration (future phase)"),
        ("Out of Scope", "Campfire ERP evaluation or integration"),
        ("Out of Scope", "ClickUp-to-HubSpot sync"),
        ("Out of Scope", "Zendesk migration or integration"),
        ("Out of Scope", "Email template design or marketing automation"),
        ("Out of Scope", "Custom HubSpot UI extensions"),
        ("Out of Scope", "Ongoing data sync maintenance post-go-live"),
        ("Out of Scope", "HubSpot licensing costs (client responsibility)"),
        ("Out of Scope", "QuotaPath platform implementation (they handle their own)"),
        ("Client Responsibilities", "Assign project liaison (Christina Edwards)"),
        ("Client Responsibilities", "Provide access: HubSpot, Salesforce, Rev IO, QuickBooks"),
        ("Client Responsibilities", "Provide Rev IO and QuickBooks API credentials + sandbox"),
        ("Client Responsibilities", "Attend weekly syncs and milestone reviews"),
        ("Client Responsibilities", "Deliver dashboard requirements write-up (Jeremy Wiley)"),
        ("Client Responsibilities", "Provide Salesforce data exports and field mapping input"),
        ("Client Responsibilities", "Communicate scope changes through formal process"),
        ("Open Items", "Confirm HubSpot instance tier (Professional vs Enterprise)"),
        ("Open Items", "Salesforce data audit needed to confirm record counts and quality"),
        ("Open Items", "Rev IO API credentials and sandbox access needed"),
        ("Open Items", "Confirm multifamily property data needs custom object or standard associations"),
        ("Open Items", "QuotaPath implementation timeline alignment"),
    ]

    for i, (section, item) in enumerate(data, 2):
        ws.cell(row=i, column=1, value=section)
        ws.cell(row=i, column=2, value=item)

    style_body(ws, 2, len(data) + 1, 2)
    auto_width(ws, 2)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 5: Approach Comparison
# ---------------------------------------------------------------------------
def sheet_approach(wb):
    ws = wb.create_sheet("Approach Comparison")

    headers = ["Dimension", "Full Scope (Recommended)", "Reduced Scope (CRM Foundation)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    rows = [
        ("Phases Included", "1-4 (all)", "1-2 only (CRM + Migration)"),
        ("Total Investment", "$42,000", "$22,000"),
        ("Timeline", "14 weeks", "8 weeks"),
        ("Requirements Covered", "46 of 46", "21 of 46"),
        ("Integrations", "Rev IO + QuickBooks + QuotaPath assist", "None (Phase 3-4 scoped separately)"),
        ("CPQ / Quoting", "HubSpot CPQ native w/ e-sig, bulk import, dynamic templates", "Deferred to follow-on"),
        ("Salesforce Migration", "Included (phased)", "Included (phased)"),
        ("Sales Process Redesign", "Included", "Included"),
        ("Reporting", "Full dashboards with QB actuals", "Pipeline-only dashboards"),
        ("Training", "3 sessions (admin, sales, ops)", "3 sessions (admin, sales, ops)"),
        ("Commission Tracking", "End-to-end data flow configured", "Deferred to follow-on"),
        ("Risk Level", "Medium (integration unknowns)", "Low (proven CRM workstreams)"),
        ("SF Deadline Risk", "Low (Aug completion, 2+ mo buffer)", "Low (Jul completion, 3+ mo buffer)"),
        ("Recommendation",
         "RECOMMENDED -- delivers full value and completes integration chain before SF contract end",
         "Choose if budget constrained or API credentials unavailable at kickoff"),
    ]

    for i, (dim, full, reduced) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=dim)
        ws.cell(row=i, column=2, value=full)
        ws.cell(row=i, column=3, value=reduced)

    style_body(ws, 2, len(rows) + 1, len(headers))
    rec_row = len(rows) + 1
    ws.cell(row=rec_row, column=2).fill = GREEN_FILL
    ws.cell(row=rec_row, column=2).font = BOLD_FONT
    auto_width(ws, len(headers))
    ws.freeze_panes = "B2"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    sheet_estimate(wb)
    sheet_phase_detail(wb)
    sheet_risks(wb)
    sheet_assumptions(wb)
    sheet_approach(wb)

    out = Path(__file__).resolve().parent / "HelloSpoke_Scoping_Estimate.xlsx"
    wb.save(str(out))
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
