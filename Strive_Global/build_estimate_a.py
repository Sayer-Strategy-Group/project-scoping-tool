"""Generate Strive Global CRM Foundation Scoping Estimate (Option A)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
body_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', size=10, bold=True)
total_font = Font(name='Arial', size=10, bold=True, color='1F4E79')
rate_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
actuals_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
green_header_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
green_header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
currency_fmt = '$#,##0'


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def style_data_cell(ws, row, col, is_alt=False):
    cell = ws.cell(row=row, column=col)
    cell.font = body_font
    cell.border = thin_border
    cell.alignment = wrap_align
    if is_alt:
        cell.fill = alt_fill


def apply_data_styles(ws, data_start, data_end, max_col):
    for r in range(data_start, data_end + 1):
        is_alt = (r - data_start) % 2 == 1
        for c in range(1, max_col + 1):
            style_data_cell(ws, r, c, is_alt)


# ============================================================
# SHEET 1: APPROACH COMPARISON
# ============================================================
ws1 = wb.active
ws1.title = 'Approach Comparison'

ws1.merge_cells('A1:C1')
ws1.cell(row=1, column=1, value='STRIVE Global -- CRM Foundation: Approach Comparison').font = Font(name='Arial', size=14, bold=True, color='1F4E79')

headers = ['Dimension', 'Approach A: Foundation First', 'Approach B: Quick Wins First']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header_row(ws1, 3, 3)

comparisons = [
    ('Description',
     'Data cleanup and contact classification first (weeks 1-4), then pipeline restructuring and dashboards (weeks 5-8). Build on clean foundation.',
     'Pipeline restructuring and dashboards first (weeks 1-4), then data cleanup and enrichment (weeks 5-8). Deliver visible wins early.'),
    ('Total Hours (Median)', '119 hrs', '119 hrs'),
    ('Timeline', '6-8 weeks', '6-8 weeks'),
    ('Risk Level', 'LOW -- clean data enables accurate dashboards from day one',
     'MEDIUM -- dashboards built on dirty data may show misleading metrics initially'),
    ('Adoption Impact', 'Slower initial visibility but higher trust in data when dashboards launch',
     'Faster dashboard delivery but reps may distrust inaccurate early reports'),
    ('Change Management', 'Gradual -- team sees data improving before process changes',
     'Immediate -- team gets new pipeline and dashboards right away'),
    ('Dependency Risk', 'LOW -- data cleanup is independent of other workstreams',
     'MEDIUM -- dashboards depend on data quality for accuracy'),
    ('Client Resource Demand', 'Higher upfront (contact exports, classification rules, validation)',
     'Lower upfront but higher mid-project for data cleanup'),
    ('Board Reporting Value', 'Clean, trustworthy reports available by week 6',
     'Reports available by week 4 but accuracy improves through week 8'),
    ('Cost', '$17,850 (median at $150/hr)', '$17,850 (median at $150/hr)'),
]

for i, (dim, a, b) in enumerate(comparisons, 4):
    ws1.cell(row=i, column=1, value=dim)
    ws1.cell(row=i, column=2, value=a)
    ws1.cell(row=i, column=3, value=b)

apply_data_styles(ws1, 4, 13, 3)

# Recommendation
ws1.merge_cells('A15:C15')
ws1.cell(row=15, column=1, value='Recommendation').font = Font(name='Arial', size=11, bold=True, color='1F4E79')
ws1.merge_cells('A16:C16')
ws1.cell(row=16, column=1, value=(
    'Approach A: Foundation First is recommended. STRIVE has 10 years of uncoded contact data and '
    'unreliable pipeline metrics. Building dashboards on dirty data risks eroding trust in the CRM '
    'before it has a chance to prove value. Clean the data first, then deliver dashboards that the '
    'team and board can trust from day one. This also aligns with Jodi\'s emphasis on building a '
    'strong data infrastructure as the foundation for future AI capabilities.'
)).font = body_font
ws1.cell(row=16, column=1).alignment = wrap_align

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 50
ws1.column_dimensions['C'].width = 50
ws1.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws1.freeze_panes = 'A4'

# ============================================================
# SHEET 2: SCOPING ESTIMATE
# ============================================================
ws2 = wb.create_sheet('Scoping Estimate')

ws2.cell(row=1, column=1, value='Hourly Rate:').font = bold_font
ws2.cell(row=1, column=2, value=150).font = bold_font
ws2.cell(row=1, column=2).fill = rate_fill
ws2.cell(row=1, column=2).number_format = currency_fmt

headers2 = ['Workstream', 'Description', 'Min Hours', 'Max Hours', 'Median Hours',
            'Rate', 'Min Cost', 'Max Cost', 'Median Cost', 'Notes / Assumptions',
            'Actual Hours', 'Actual Cost', 'Variance']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, 13)

# Green headers for actuals columns
for col in [11, 12, 13]:
    cell = ws2.cell(row=3, column=col)
    cell.fill = green_header_fill
    cell.font = green_header_font

workstreams = [
    ('CRM Architecture & Config Review',
     'Portal audit, user roles/permissions, team structure reconfiguration for existing Enterprise instance',
     6, 12,
     'Assumes existing HubSpot Sales Hub Enterprise. Up to 23 seat users across sales, core, and service.'),
    ('Pipeline Restructuring',
     'Separate pre-pipeline from active pipeline. Partner-level vs. client-level deal categories. Stage probability mapping (5/20/40/50/60/80/90/100). Required fields per stage. Add Contracting stage.',
     10, 18,
     'Two deal categories (partner, client) in same pipeline. Pre-pipeline as separate pipeline or board. Jodi\'s probability framework as starting point.'),
    ('Contact & Company Classification + Cleanup',
     'Contact type taxonomy (client/prospect/broker/partner). AI enrichment using HubSpot Breeze credits. Normalization, deduplication, association fixes.',
     14, 24,
     '10 years of uncoded data. +30-50% data quality premium applied. Includes pilot enrichment on 500 records. Assumes up to 30 custom properties.'),
    ('Deal Health Scoring & Forecasting',
     'Deal score based on stage + time-in-stage + activity recency. Healthy/at-risk/stale classification. Standardized close won/lost reasons (replace free-form with categorized dropdowns).',
     6, 12,
     'Scoring formula and thresholds defined collaboratively with Andrea and Jodi.'),
    ('Reporting & Dashboards',
     'Partner dashboard (clonable with quick filters). Seller-specific KPI dashboard. CEO/board dashboard (sales revenue, new business, pipeline health). Up to 12 custom reports.',
     10, 18,
     'Includes: open deals, pipeline value, avg deals/week, closed-won revenue, loss reasons, avg deal size, avg term length, appointments, contacts added.'),
    ('Automations & Workflows',
     'Fix Outlook email logging issue. Lead source tracking automation. Form-to-CRM for microsite client capture. Deal stage-change notifications. Duplicate protection rules.',
     8, 16,
     'Up to 8 automated workflows. Outlook fix may require HubSpot support ticket if platform bug.'),
    ('Email & Communication Fix',
     'Outlook integration audit and fix. Email association rules to prevent logging to all related deals. Domain exclusions for internal communications.',
     4, 8,
     'Root cause diagnosis in week 1. If config fix: 2-4 hrs. If platform bug: HubSpot support escalation.'),
    ('Training: Admin',
     'Andrea + designated admin. Pipeline management, reporting, workflow administration, data hygiene practices, enrichment tool usage.',
     4, 8,
     '1-2 sessions. Includes recorded walkthrough for future reference.'),
    ('Training: End Users',
     'Sales team. Daily CRM usage, deal entry, activity logging, dashboard consumption, mobile app setup.',
     6, 10,
     '2-3 sessions. Tailored to current team size and workflow.'),
    ('Training: Executive',
     'Jodi. Board dashboard walkthrough, forecast consumption, report scheduling, export for board decks.',
     2, 4,
     '1 session. Focused on CEO dashboard and board reporting.'),
    ('UAT, Go-Live & Documentation',
     'Test scripts, 2 UAT sessions with key stakeholders, bug fix window, go-live cutover plan. Deliverables: 1 admin guide, 1 user guide. 2-week hypercare post-launch.',
     6, 10,
     'Hypercare includes issue resolution and adoption coaching.'),
    ('Project Management',
     'Kickoff, weekly syncs, status updates, change management, milestone reviews.',
     10, 16,
     '~12% of total estimated hours.'),
]

for i, (name, desc, min_h, max_h, notes) in enumerate(workstreams, 4):
    row = i
    ws2.cell(row=row, column=1, value=name)
    ws2.cell(row=row, column=2, value=desc)
    ws2.cell(row=row, column=3, value=min_h)
    ws2.cell(row=row, column=4, value=max_h)
    ws2.cell(row=row, column=5).value = f'=AVERAGE(C{row},D{row})'
    ws2.cell(row=row, column=6).value = f'=$B$1'
    ws2.cell(row=row, column=7).value = f'=C{row}*$B$1'
    ws2.cell(row=row, column=8).value = f'=D{row}*$B$1'
    ws2.cell(row=row, column=9).value = f'=E{row}*$B$1'
    ws2.cell(row=row, column=10, value=notes)
    # Actuals columns (green, empty for input)
    for col in [11, 12, 13]:
        ws2.cell(row=row, column=col).fill = actuals_fill

    ws2.cell(row=row, column=13).value = f'=IF(K{row}="","",K{row}-E{row})'

total_row = 4 + len(workstreams)
ws2.cell(row=total_row, column=1, value='TOTAL').font = total_font
for col in [3, 4, 5, 7, 8, 9, 11, 12]:
    cell = ws2.cell(row=total_row, column=col)
    col_letter = get_column_letter(col)
    cell.value = f'=SUM({col_letter}4:{col_letter}{total_row - 1})'
    cell.font = total_font
    cell.border = thin_border
ws2.cell(row=total_row, column=6).value = f'=$B$1'
ws2.cell(row=total_row, column=6).font = total_font

# Apply formatting
for col in [7, 8, 9, 12]:
    for row in range(4, total_row + 1):
        ws2.cell(row=row, column=col).number_format = currency_fmt
ws2.cell(row=1, column=2).number_format = currency_fmt

apply_data_styles(ws2, 4, total_row - 1, 10)
# Re-apply actuals fill
for row in range(4, total_row + 1):
    for col in [11, 12, 13]:
        ws2.cell(row=row, column=col).fill = actuals_fill
        ws2.cell(row=row, column=col).border = thin_border
        ws2.cell(row=row, column=col).font = body_font

# Column widths
widths = {1: 30, 2: 50, 3: 12, 4: 12, 5: 14, 6: 10, 7: 12, 8: 12, 9: 14, 10: 50, 11: 14, 12: 14, 13: 12}
for col, w in widths.items():
    ws2.column_dimensions[get_column_letter(col)].width = w

ws2.freeze_panes = 'B4'

# ============================================================
# SHEET 3: RISK REGISTER
# ============================================================
ws3 = wb.create_sheet('Risk Register')

headers3 = ['#', 'Risk', 'Severity', 'Likelihood', 'Impact', 'Mitigation', 'Owner', 'Status']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header_row(ws3, 1, 8)

sev_high_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
sev_med_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
sev_low_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

risks = [
    (1, '10 years of uncoded contact data -- AI enrichment scope may exceed estimates',
     'HIGH', 'HIGH',
     'Data cleanup takes 2-3x estimate; marketing segmentation delayed; project timeline extends',
     'Run AI enrichment pilot on 500 records first. Measure classification accuracy before committing to full database run. Budget 30% buffer.',
     'Consultant + Andrea', 'Open'),
    (2, 'Pipeline under-reporting -- reps using LinkedIn and manual quoting outside CRM',
     'HIGH', 'HIGH',
     'Forecasting unreliable during transition; board sees pipeline dip before improvement',
     'Jodi mandates CRM usage pre-launch. Set data entry expectations. Brief board that reported pipeline will increase as capture improves (not because business grew).',
     'Jodi (executive sponsor)', 'Open'),
    (3, 'Scope creep from close business relationship (shared office, referrals)',
     'HIGH', 'MEDIUM',
     'Informal requests bypass SOW; hours consumed without authorization; budget overrun',
     'SOW signed by budget authority. All change requests documented formally. Weekly hours tracking visible to client.',
     'Consultant + Jodi', 'Open'),
    (4, 'Board ROI expectations -- if narrative overpromises, credibility risk',
     'HIGH', 'MEDIUM',
     'Board approves but expects unrealistic timeline or savings; sets up failure',
     'ROI narrative uses conservative estimates with ranges, not point values. Tie ROI to measurable KPIs with 90-day check-in.',
     'Consultant + Jodi', 'Open'),
    (5, 'Seat/license confusion -- mixed versions across users, potential unused seats',
     'MEDIUM', 'HIGH',
     'Paying for unused seats; features expected but unavailable on current tier',
     'Complete license audit in week 1. Produce cost savings memo before any configuration work.',
     'Consultant + Jodi', 'Open'),
    (6, 'Adoption risk -- sales reps prefer LinkedIn Sales Navigator over CRM',
     'MEDIUM', 'HIGH',
     'New CRM configuration goes unused; ROI unrealized; board investment wasted',
     'Executive mandate from Jodi. Make CRM easier than alternatives (mobile, one-click logging). Rep-level dashboards show their own metrics.',
     'Jodi + Andrea', 'Open'),
    (7, 'Outlook email logging defect -- emails logging to all associated deals',
     'LOW', 'MEDIUM',
     'Fix may require HubSpot support escalation with unknown timeline; workaround needed in interim',
     'Diagnose root cause in week 1. If platform bug, open HubSpot support ticket immediately and track SLA.',
     'Consultant', 'Open'),
    (8, 'COO identity ambiguity -- Lauren vs. Zach Beegal referenced in different calls',
     'LOW', 'LOW',
     'Miscommunication on stakeholder authority; NDA or admin access granted to wrong person',
     'Clarify org chart and who controls HubSpot admin access during kickoff.',
     'Consultant', 'Open'),
]

for i, (num, risk, sev, lik, impact, mit, owner, status) in enumerate(risks, 2):
    ws3.cell(row=i, column=1, value=num)
    ws3.cell(row=i, column=2, value=risk)
    ws3.cell(row=i, column=3, value=sev)
    ws3.cell(row=i, column=4, value=lik)
    ws3.cell(row=i, column=5, value=impact)
    ws3.cell(row=i, column=6, value=mit)
    ws3.cell(row=i, column=7, value=owner)
    ws3.cell(row=i, column=8, value=status)

    # Color severity
    sev_cell = ws3.cell(row=i, column=3)
    if sev == 'HIGH':
        sev_cell.fill = sev_high_fill
    elif sev == 'MEDIUM':
        sev_cell.fill = sev_med_fill
    else:
        sev_cell.fill = sev_low_fill

apply_data_styles(ws3, 2, len(risks) + 1, 8)
# Re-apply severity colors
for i, (_, _, sev, *_rest) in enumerate(risks, 2):
    sev_cell = ws3.cell(row=i, column=3)
    if sev == 'HIGH':
        sev_cell.fill = sev_high_fill
    elif sev == 'MEDIUM':
        sev_cell.fill = sev_med_fill
    else:
        sev_cell.fill = sev_low_fill

widths3 = {1: 5, 2: 45, 3: 12, 4: 12, 5: 40, 6: 50, 7: 25, 8: 10}
for col, w in widths3.items():
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.freeze_panes = 'A2'

# ============================================================
# SHEET 4: ASSUMPTIONS & EXCLUSIONS
# ============================================================
ws4 = wb.create_sheet('Assumptions & Exclusions')

section_font = Font(name='Arial', size=11, bold=True, color='1F4E79')

# Section 1: Scope Assumptions
ws4.cell(row=1, column=1, value='Scope Assumptions').font = section_font
assumptions = [
    'HubSpot Sales Hub Enterprise (existing instance -- reconfiguration, not net-new setup)',
    'Up to 23 licensed users across 3 core, 10 sales, and 10 service seats',
    'Outlook as primary email platform (with known logging defect to be resolved)',
    'Up to 30 custom contact/company properties',
    '1 active sales pipeline with pre-pipeline separated (new pipeline or board view)',
    '2 deal categories: partner-level and client-level (same stages, different segmentation)',
    'Stage probability mapping using Jodi\'s framework as starting point (5/20/40/50/60/80/90/100)',
    'Contact type classification: client, prospect, broker, partner',
    'AI enrichment using HubSpot Breeze credits (5-10K credits available)',
    'Up to 8 automated workflows',
    '3-4 reporting dashboards with up to 12 custom reports',
    '6-8 week implementation timeline',
    'All work conducted remotely unless otherwise agreed',
    'Any scope changes managed through formal change request and may impact cost or timing',
]
for i, a in enumerate(assumptions, 2):
    ws4.cell(row=i, column=1, value=i - 1)
    ws4.cell(row=i, column=2, value=a)

# Section 2: Out of Scope
gap = len(assumptions) + 3
ws4.cell(row=gap, column=1, value='Out of Scope').font = section_font
exclusions = [
    'CPQ / quoting / product library build (see Option B: Full CRM Overhaul)',
    'Email sequencing or sales cadence automation',
    'Marketing automation (newsletters, campaigns, landing pages)',
    'NetSuite, Bill.com, or any ERP/financial system integration',
    'Client Success pipeline or client health scoring',
    'Commission tracking or calculation automation',
    'Contract management or e-signature workflows',
    'Custom API development or middleware (n8n/Zapier)',
    'Data migration from external systems (scope covers HubSpot internal cleanup only)',
    'Ongoing managed services beyond 2-week hypercare',
    'HubSpot license procurement or tier changes',
    'Hardware, IT infrastructure, or network changes',
]
for i, e in enumerate(exclusions, gap + 1):
    ws4.cell(row=i, column=1, value=i - gap)
    ws4.cell(row=i, column=2, value=e)

# Section 3: Client Responsibilities
gap2 = gap + len(exclusions) + 2
ws4.cell(row=gap2, column=1, value='Client Responsibilities').font = section_font
responsibilities = [
    'Designate primary point of contact for decisions and approvals (Andrea recommended)',
    'Ensure Jodi and Andrea availability for discovery, UAT, and training sessions',
    'Provide contact classification rules and validation for AI enrichment results',
    'Confirm deal stage definitions and probability framework before pipeline build',
    'Communicate standardized close won/lost reason categories',
    'Mandate CRM usage across sales team (executive sponsorship from Jodi)',
    'Complete UAT testing within agreed timeline',
    'Provide access to HubSpot admin portal',
    'Share 3-5 recent quotes/proposals for quoting workflow reference',
]
for i, r in enumerate(responsibilities, gap2 + 1):
    ws4.cell(row=i, column=1, value=i - gap2)
    ws4.cell(row=i, column=2, value=r)

# Section 4: Open Items
gap3 = gap2 + len(responsibilities) + 2
ws4.cell(row=gap3, column=1, value='Open Items').font = section_font
open_items = [
    'Exact contact/company/deal record counts in HubSpot (pending portal access)',
    'Service Hub usage details -- who uses it, preserve or deprecate?',
    'COO identity and HubSpot admin access ownership (Lauren vs. Zach Beegal)',
    'Outlook email logging defect root cause (config vs. platform bug)',
    'Existing workflow inventory beyond broker trial form',
    'Historical deal data reliability for probability calibration',
    'Microsite client intake form link from Andrea',
    'Jodi\'s Wagmo stage naming framework (if available)',
    'SOW signing authority (Jodi as CEO or board approval required)',
]
for i, o in enumerate(open_items, gap3 + 1):
    ws4.cell(row=i, column=1, value=i - gap3)
    ws4.cell(row=i, column=2, value=o)

# Apply styles
for row in ws4.iter_rows(min_row=2, max_col=3):
    for cell in row:
        cell.font = body_font
        cell.alignment = wrap_align

ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 80
ws4.column_dimensions['C'].width = 30

# ============================================================
# SAVE
# ============================================================
output_path = '/Users/harbuckconsulting/projects/Project Scoping Tool/Strive_Global/Strive_CRM_Foundation_Estimate.xlsx'
wb.save(output_path)
print(f'Option A saved to: {output_path}')
