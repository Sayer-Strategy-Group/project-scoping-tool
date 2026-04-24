"""Generate Strive Global Full CRM Overhaul Scoping Estimate V2 (Option B) - Wagmo-calibrated"""
import sys
import pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1] / 'scripts'))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from brand_styles import (
    HEADER_FONT, HEADER_FILL, BODY_FONT, BOLD_BODY_FONT, TOTAL_FONT,
    INPUT_FILL, ALT_ROW_FILL, SECONDARY_HEADER_FONT, SECONDARY_HEADER_FILL,
    THIN_BORDER, WRAP_ALIGN, TITLE_FONT, SECTION_FONT, CURRENCY_FMT,
    FONT_FAMILY, SAYER_BLACK,
    style_header_row, apply_data_styles_rows,
    get_column_letter,
)

wb = openpyxl.Workbook()

# Backwards-compat aliases — brand specs live in scripts/brand_styles.py.
header_font = HEADER_FONT
header_fill = HEADER_FILL
body_font = BODY_FONT
bold_font = BOLD_BODY_FONT
total_font = TOTAL_FONT
rate_fill = INPUT_FILL
alt_fill = ALT_ROW_FILL
actuals_fill = INPUT_FILL
green_header_fill = SECONDARY_HEADER_FILL
green_header_font = SECONDARY_HEADER_FONT
section_font = SECTION_FONT
thin_border = THIN_BORDER
wrap_align = WRAP_ALIGN
currency_fmt = CURRENCY_FMT

# This script's original apply_data_styles took a row list, not a start/end pair.
# Point the old name at the row-list variant so existing call sites keep working.
apply_data_styles = apply_data_styles_rows


# ============================================================
# SHEET 1: APPROACH COMPARISON
# ============================================================
ws1 = wb.active
ws1.title = 'Approach Comparison'

ws1.merge_cells('A1:C1')
ws1.cell(row=1, column=1, value='STRIVE Global -- Full CRM Overhaul V2: Approach Comparison').font = TITLE_FONT

headers = ['Dimension', 'Approach A: Phased (Recommended)', 'Approach B: Concurrent']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header_row(ws1, 3, 3)

comparisons = [
    ('Description',
     'Phase 1: CRM Foundation (weeks 1-8) -- pipeline, contacts, dashboards, automations. Phase 2: CPQ, enrichment, integrations, marketing (weeks 9-18). Each phase delivers standalone value.',
     'All workstreams in parallel across 12-14 weeks. Maximum resource intensity.'),
    ('Total Hours (Median)', '313 hrs across two phases', '313 hrs compressed'),
    ('Timeline', '16-18 weeks (Phase 1: 8-10 wks + Phase 2: 8-10 wks)',
     '12-14 weeks'),
    ('Risk Level', 'LOW -- Phase 1 de-risks Phase 2; clean data enables CPQ accuracy',
     'HIGH -- dirty data undermines CPQ and reporting; multiple moving parts'),
    ('Adoption Impact', 'Team absorbs changes gradually; builds confidence before CPQ launch',
     'All-at-once change; higher cognitive load'),
    ('Change Management', 'Two training cycles; Phase 1 habits established before Phase 2',
     'Single push covering all features'),
    ('Dependency Risk', 'LOW -- Phase 2 depends on Phase 1, but Phase 1 is self-contained',
     'HIGH -- CPQ depends on clean pipeline AND contact data simultaneously'),
    ('Client Resource Demand', 'Spread over 18 weeks; manageable alongside day-to-day',
     'Concentrated in 14 weeks; significant time commitment'),
    ('Board Reporting Value', 'Phase 1 dashboards by week 8; full CPQ and advanced reporting by week 18',
     'Full suite by week 14 if everything goes to plan'),
    ('Cost', '$46,950 (median at $150/hr)', '$46,950 (median at $150/hr)'),
    ('Budget Flexibility', 'Can pause after Phase 1 ($23,250) if budget constrained',
     'Full commitment required upfront'),
]

for i, (dim, a, b) in enumerate(comparisons, 4):
    ws1.cell(row=i, column=1, value=dim)
    ws1.cell(row=i, column=2, value=a)
    ws1.cell(row=i, column=3, value=b)

apply_data_styles(ws1, list(range(4, 15)), 3)

ws1.merge_cells('A16:C16')
ws1.cell(row=16, column=1, value='Recommendation').font = section_font
ws1.merge_cells('A17:C17')
ws1.cell(row=17, column=1, value=(
    'Approach A: Phased is strongly recommended. STRIVE\'s data quality issues make concurrent '
    'approach risky -- CPQ built on dirty data produces inaccurate quotes. Phase 1 delivers immediate '
    'operational value and can stand alone if Phase 2 is deferred. Estimates calibrated against prior '
    'comparable Salesforce engagement and HubSpot industry benchmarks (CPQ: 40-60 hrs industry standard).'
)).font = body_font
ws1.cell(row=17, column=1).alignment = wrap_align

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 55
ws1.column_dimensions['C'].width = 55
ws1.freeze_panes = 'A4'

# ============================================================
# SHEET 2: SCOPING ESTIMATE
# ============================================================
ws2 = wb.create_sheet('Scoping Estimate')

ws2.cell(row=1, column=1, value='Hourly Rate:').font = bold_font
ws2.cell(row=1, column=2, value=150).font = bold_font
ws2.cell(row=1, column=2).fill = rate_fill
ws2.cell(row=1, column=2).number_format = currency_fmt

headers2 = ['Workstream', 'Description', 'Min Hours', 'Max Hours', 'Median Hours',
            'Rate', 'Min Cost', 'Max Cost', 'Median Cost', 'Notes / Assumptions',
            'Actual Hours', 'Actual Cost', 'Variance']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, 13)
for col in [11, 12, 13]:
    ws2.cell(row=3, column=col).fill = green_header_fill
    ws2.cell(row=3, column=col).font = green_header_font

# PHASE 1
ws2.cell(row=4, column=1, value='PHASE 1: CRM FOUNDATION (Weeks 1-8)').font = section_font
ws2.merge_cells('A4:J4')

workstreams_p1 = [
    ('CRM Architecture & Config Review', 'Portal audit, user roles, permissions, team structure, 3-hub assessment, seat review.', 8, 16,
     'Wagmo comparable: 14 hrs. HubSpot simpler but Enterprise 3-hub complexity.'),
    ('Pipeline Restructuring', 'Pre-pipeline separation, partner/client categories, stage probabilities, required fields, Contracting stage, auto-transitions, deal age reset.', 16, 32,
     'Wagmo benchmark: 36 hrs SFDC pipeline. HubSpot -20%, Strive complexity +40%.'),
    ('Contact & Company Classification + Cleanup', 'Contact taxonomy, AI enrichment pilot + full run, normalization, dedup, association fixes, duplicate protection.', 16, 32,
     '10 yrs uncoded data. Industry: $5K-$15K for cleanup. Conservative with AI.'),
    ('Deal Health Scoring & Forecasting', 'Deal score formula, healthy/at-risk/stale, standardized close reasons, forecasting views.', 6, 14,
     'Scoring defined collaboratively.'),
    ('Reporting & Dashboards', 'Partner, seller, CEO/board, pipeline health dashboards. 15 custom reports. Scheduled exports.', 14, 26,
     'Wagmo: 20 hrs for simpler scope. Strive needs 4 dashboards + 15 reports.'),
    ('Automations & Workflows', 'Outlook fix, lead source, form-to-CRM, stage notifications, dup protection, pre-pipeline transitions.', 10, 18,
     'Up to 10 workflows.'),
    ('Email & Communication Fix', 'Outlook audit/fix, association rules, domain exclusions.', 4, 8, 'Root cause week 1.'),
    ('Training: Admin', 'Andrea + admin. Pipeline, reporting, workflows, data hygiene.', 4, 8, '1-2 sessions recorded.'),
    ('Training: End Users', 'Sales team. CRM usage, deals, activity, dashboards, mobile.', 6, 10, '2-3 sessions.'),
    ('Training: Executive', 'Jodi. Board dashboard, forecasting, scheduling.', 2, 4, '1 session.'),
    ('UAT, Go-Live & Documentation (Phase 1)', 'Test scripts, 2 UAT sessions, bug fixes, admin + user guide, 2-week hypercare.', 8, 12,
     'Wagmo testing: 7 hrs + documentation.'),
    ('Project Management (Phase 1)', 'Kickoff, weekly syncs, status updates, scope control.', 14, 22,
     '~12% of total. Wagmo lesson: formal PM prevents scope creep.'),
]

data_start = 5
all_p1_rows = []
for i, (name, desc, min_h, max_h, notes) in enumerate(workstreams_p1, data_start):
    row = i
    all_p1_rows.append(row)
    ws2.cell(row=row, column=1, value=name)
    ws2.cell(row=row, column=2, value=desc)
    ws2.cell(row=row, column=3, value=min_h)
    ws2.cell(row=row, column=4, value=max_h)
    ws2.cell(row=row, column=5).value = f'=AVERAGE(C{row},D{row})'
    ws2.cell(row=row, column=6).value = '=$B$1'
    ws2.cell(row=row, column=7).value = f'=C{row}*$B$1'
    ws2.cell(row=row, column=8).value = f'=D{row}*$B$1'
    ws2.cell(row=row, column=9).value = f'=E{row}*$B$1'
    ws2.cell(row=row, column=10, value=notes)
    for col in [11, 12, 13]:
        ws2.cell(row=row, column=col).fill = actuals_fill
    ws2.cell(row=row, column=13).value = f'=IF(K{row}="","",K{row}-E{row})'

p1_end = data_start + len(workstreams_p1) - 1
p1_total = p1_end + 1
ws2.cell(row=p1_total, column=1, value='Phase 1 Subtotal').font = total_font
for col in [3, 4, 5, 7, 8, 9, 11, 12]:
    cl = get_column_letter(col)
    ws2.cell(row=p1_total, column=col).value = f'=SUM({cl}{data_start}:{cl}{p1_end})'
    ws2.cell(row=p1_total, column=col).font = total_font
    ws2.cell(row=p1_total, column=col).border = thin_border
ws2.cell(row=p1_total, column=6).value = '=$B$1'

# PHASE 2
p2_label = p1_total + 2
ws2.cell(row=p2_label, column=1, value='PHASE 2: CPQ, ENRICHMENT & INTEGRATIONS (Weeks 9-18)').font = section_font
ws2.merge_cells(f'A{p2_label}:J{p2_label}')

workstreams_p2 = [
    ('CPQ Implementation', 'Product library (apps, microsites, AI, packs, Divi, bundles). Quote templates. E-signatures. Reminders. Multiple pricing models (license+rev share, flat, bundles/discounts).', 35, 55,
     'Industry benchmark: 40-60 hrs for CPQ rollout (Process Pro Consulting). Multiple pricing models add complexity. V1 was 18 hrs -- significantly under-scoped.'),
    ('Contract Management Workflows', 'Quote-to-contract automation. Status tracking. Renewal/expiry workflows. Replaces manual DocuSign through COO.', 8, 14,
     'Contract templates, status properties, auto-advance on signature.'),
    ('Data Enrichment at Scale', 'Full database Breeze AI enrichment. Auto-enrichment for inbound. QA and validation.', 8, 16,
     'Extends Phase 1 pilot to full database. Includes QA cycle.'),
    ('Data Normalization & Dedup (Extended)', 'Full normalization, advanced dedup, merge strategy, ongoing protection.', 10, 18,
     'Extends Phase 1 cleanup comprehensively.'),
    ('Parent-Child Structure (Broker Packs)', 'Association labels, per-broker forms, auto child contact creation, list generation.', 6, 10,
     'Enables microsite-to-app expansion. Depends on form link from Andrea.'),
    ('Email Sequencing', 'Broker prospecting, microsite-to-app expansion. 4 sequences + enrollment + reporting.', 6, 10,
     'HubSpot Enterprise sequences included in license.'),
    ('Marketing Automation Foundation', 'Newsletter template, 2-3 segmented campaigns, compliance.', 8, 14,
     'Leverages Marketing Hub (already licensed).'),
    ('Commission Visibility & Validation', 'CRM-based commission view, validation workflow for Jodi.', 4, 8,
     'Visibility only -- not full calculation engine.'),
    ('NetSuite Integration Prep', 'Data readiness, field mapping, gap analysis. PREP ONLY.', 8, 14,
     'Does NOT include API build. Separate SOW required.'),
    ('Bill.com Integration Prep', 'AR/AP field mapping, data readiness. PREP ONLY.', 4, 8,
     'Does NOT include build. Separate SOW.'),
    ('Seat/License Optimization', 'Audit 23 seats, identify waste, cost savings memo.', 2, 4,
     'Deliverable: cost savings recommendation.'),
    ('Advanced Reporting', 'Automated board reports, pipeline coverage, pace-to-goal, marketing ROI.', 8, 14,
     'Extends Phase 1 dashboards with Phase 2 data.'),
    ('Additional Project Management', 'Phase 2 PM, weekly syncs, milestone reviews.', 12, 20,
     '~12% of Phase 2 hours.'),
]

p2_start = p2_label + 1
all_p2_rows = []
for i, (name, desc, min_h, max_h, notes) in enumerate(workstreams_p2, p2_start):
    row = i
    all_p2_rows.append(row)
    ws2.cell(row=row, column=1, value=name)
    ws2.cell(row=row, column=2, value=desc)
    ws2.cell(row=row, column=3, value=min_h)
    ws2.cell(row=row, column=4, value=max_h)
    ws2.cell(row=row, column=5).value = f'=AVERAGE(C{row},D{row})'
    ws2.cell(row=row, column=6).value = '=$B$1'
    ws2.cell(row=row, column=7).value = f'=C{row}*$B$1'
    ws2.cell(row=row, column=8).value = f'=D{row}*$B$1'
    ws2.cell(row=row, column=9).value = f'=E{row}*$B$1'
    ws2.cell(row=row, column=10, value=notes)
    for col in [11, 12, 13]:
        ws2.cell(row=row, column=col).fill = actuals_fill
    ws2.cell(row=row, column=13).value = f'=IF(K{row}="","",K{row}-E{row})'

p2_end = p2_start + len(workstreams_p2) - 1
p2_total = p2_end + 1
ws2.cell(row=p2_total, column=1, value='Phase 2 Subtotal').font = total_font
for col in [3, 4, 5, 7, 8, 9, 11, 12]:
    cl = get_column_letter(col)
    ws2.cell(row=p2_total, column=col).value = f'=SUM({cl}{p2_start}:{cl}{p2_end})'
    ws2.cell(row=p2_total, column=col).font = total_font
    ws2.cell(row=p2_total, column=col).border = thin_border
ws2.cell(row=p2_total, column=6).value = '=$B$1'

# Optional add-on
addon_label = p2_total + 2
ws2.cell(row=addon_label, column=1, value='OPTIONAL ADD-ON').font = BOLD_BODY_FONT
ws2.merge_cells(f'A{addon_label}:J{addon_label}')

addon_row = addon_label + 1
ws2.cell(row=addon_row, column=1, value='Client Success Pipeline (Optional)')
ws2.cell(row=addon_row, column=2, value='Client health tracking, health score properties, renewal tracking. Requires CS team discovery.')
ws2.cell(row=addon_row, column=3, value=8)
ws2.cell(row=addon_row, column=4, value=14)
ws2.cell(row=addon_row, column=5).value = f'=AVERAGE(C{addon_row},D{addon_row})'
ws2.cell(row=addon_row, column=6).value = '=$B$1'
ws2.cell(row=addon_row, column=7).value = f'=C{addon_row}*$B$1'
ws2.cell(row=addon_row, column=8).value = f'=D{addon_row}*$B$1'
ws2.cell(row=addon_row, column=9).value = f'=E{addon_row}*$B$1'
ws2.cell(row=addon_row, column=10, value='Not in project total. Requires separate CS discovery.')
for col in [11, 12, 13]:
    ws2.cell(row=addon_row, column=col).fill = actuals_fill

# Grand total
grand_row = addon_row + 2
ws2.cell(row=grand_row, column=1, value='PROJECT TOTAL (Phase 1 + Phase 2)').font = SECTION_FONT
for col in [3, 4, 5, 7, 8, 9]:
    cl = get_column_letter(col)
    ws2.cell(row=grand_row, column=col).value = f'={cl}{p1_total}+{cl}{p2_total}'
    ws2.cell(row=grand_row, column=col).font = SECTION_FONT
    ws2.cell(row=grand_row, column=col).border = thin_border
ws2.cell(row=grand_row, column=6).value = '=$B$1'

# Formatting
all_data_rows = all_p1_rows + all_p2_rows + [addon_row]
apply_data_styles(ws2, all_data_rows, 10)
for row in all_data_rows:
    for col in [11, 12, 13]:
        ws2.cell(row=row, column=col).fill = actuals_fill
        ws2.cell(row=row, column=col).border = thin_border
        ws2.cell(row=row, column=col).font = body_font

for col in [7, 8, 9, 12]:
    for row in all_data_rows + [p1_total, p2_total, grand_row, addon_row]:
        ws2.cell(row=row, column=col).number_format = currency_fmt

widths = {1: 35, 2: 55, 3: 12, 4: 12, 5: 14, 6: 10, 7: 12, 8: 12, 9: 14, 10: 55, 11: 14, 12: 14, 13: 12}
for col, w in widths.items():
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.freeze_panes = 'B4'

# ============================================================
# SHEET 3: RISK REGISTER
# ============================================================
ws3 = wb.create_sheet('Risk Register')
headers3 = ['#', 'Risk', 'Severity', 'Likelihood', 'Impact', 'Mitigation', 'Owner', 'Status']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header_row(ws3, 1, 8)

from brand_styles import SEV_HIGH_FILL as sev_high, SEV_MED_FILL as sev_med, SEV_LOW_FILL as sev_low

risks = [
    (1, '10 years uncoded contact data -- enrichment may exceed estimates', 'HIGH', 'HIGH',
     'Cleanup 2-3x estimate; CPQ associations unreliable', 'Pilot 500 records in Phase 1. Budget 30% buffer.', 'Consultant + Andrea', 'Open'),
    (2, 'Pipeline under-reporting -- reps on LinkedIn', 'HIGH', 'HIGH',
     'Forecast unreliable; CPQ adoption depends on deal entry', 'Jodi mandates CRM usage. Board briefed on pipeline increase.', 'Jodi', 'Open'),
    (3, 'Scope creep from close relationship', 'HIGH', 'MEDIUM',
     'Informal requests; budget overrun on 18-week engagement', 'Formal SOW. Weekly hours tracking. Change request process.', 'Consultant + Jodi', 'Open'),
    (4, 'Board ROI expectations', 'HIGH', 'MEDIUM',
     'Unrealistic timeline expectations', 'Conservative ranges. 90-day check-in after Phase 1.', 'Consultant + Jodi', 'Open'),
    (5, 'CPQ product library complexity', 'MEDIUM', 'HIGH',
     'HubSpot CPQ may not handle all pricing models natively', 'Audit 5 recent quotes before build. Confirm Enterprise CPQ handles rev share.', 'Consultant', 'Open'),
    (6, 'Seat/license confusion', 'MEDIUM', 'HIGH',
     'Paying for unused seats; CPQ may need specific seats', 'License audit week 1. Cost savings memo.', 'Consultant + Jodi', 'Open'),
    (7, 'NetSuite/Bill.com scope bleed', 'MEDIUM', 'MEDIUM',
     'Prep becomes build without budget', 'Explicit boundary: prep only. Build = separate SOW.', 'Consultant', 'Open'),
    (8, 'Adoption risk', 'MEDIUM', 'HIGH',
     'CRM + CPQ unused; ROI unrealized', 'Executive mandate. Rep dashboards. Sequences reduce manual work.', 'Jodi + Andrea', 'Open'),
    (9, 'Outlook email logging defect', 'LOW', 'MEDIUM',
     'HubSpot support escalation timeline', 'Diagnose week 1. Open ticket if needed.', 'Consultant', 'Open'),
    (10, 'COO ambiguity', 'LOW', 'LOW', 'Admin access issues', 'Clarify at kickoff.', 'Consultant', 'Open'),
]

for i, (num, risk, sev, lik, impact, mit, owner, status) in enumerate(risks, 2):
    ws3.cell(row=i, column=1, value=num)
    ws3.cell(row=i, column=2, value=risk)
    ws3.cell(row=i, column=3, value=sev)
    ws3.cell(row=i, column=4, value=lik)
    ws3.cell(row=i, column=5, value=impact)
    ws3.cell(row=i, column=6, value=mit)
    ws3.cell(row=i, column=7, value=owner)
    ws3.cell(row=i, column=8, value=status)
    c = ws3.cell(row=i, column=3)
    if sev == 'HIGH': c.fill = sev_high
    elif sev == 'MEDIUM': c.fill = sev_med
    else: c.fill = sev_low

apply_data_styles(ws3, list(range(2, len(risks) + 2)), 8)
for i, (_, _, sev, *_) in enumerate(risks, 2):
    c = ws3.cell(row=i, column=3)
    if sev == 'HIGH': c.fill = sev_high
    elif sev == 'MEDIUM': c.fill = sev_med
    else: c.fill = sev_low

widths3 = {1: 5, 2: 50, 3: 12, 4: 12, 5: 45, 6: 55, 7: 25, 8: 10}
for col, w in widths3.items():
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.freeze_panes = 'A2'

# ============================================================
# SHEET 4: ASSUMPTIONS & EXCLUSIONS
# ============================================================
ws4 = wb.create_sheet('Assumptions & Exclusions')

ws4.cell(row=1, column=1, value='Scope Assumptions').font = section_font
assumptions = [
    'HubSpot Sales Hub Enterprise (existing instance -- reconfiguration)',
    'Up to 23 licensed users across core, sales, and service seats',
    'Outlook as primary email platform',
    'Up to 40 custom properties (expanded for CPQ and enrichment)',
    '1 active sales pipeline with pre-pipeline separated',
    '2 deal categories: partner-level and client-level',
    'Stage probability mapping using Jodi\'s framework',
    'AI enrichment using HubSpot Breeze credits (5-10K)',
    'Up to 12 automated workflows',
    '4-5 dashboards with up to 20 custom reports',
    'Complete product/pricing inventory provided before CPQ build',
    'CPQ hours based on industry benchmark of 40-60 hrs (Process Pro Consulting)',
    'Marketing Hub leveraged for campaigns (already licensed)',
    '16-18 week phased timeline',
    'Phase 1 can stand alone if Phase 2 deferred',
    'Estimates calibrated against prior Salesforce engagement and industry benchmarks',
    'All work conducted remotely',
    'Scope changes via formal change request',
]
for i, a in enumerate(assumptions, 2):
    ws4.cell(row=i, column=1, value=i - 1); ws4.cell(row=i, column=2, value=a)

gap = len(assumptions) + 3
ws4.cell(row=gap, column=1, value='Out of Scope').font = section_font
exclusions = [
    'NetSuite integration build (prep only)', 'Bill.com integration build (prep only)',
    'Client Success pipeline (optional add-on)', 'Full commission calculation engine',
    'Custom API development beyond HubSpot native', 'Other CRMs',
    'External data migration', 'Managed services beyond hypercare per phase',
    'License procurement', 'AI knowledge assistant', 'LinkedIn Navigator integration',
]
for i, e in enumerate(exclusions, gap + 1):
    ws4.cell(row=i, column=1, value=i - gap); ws4.cell(row=i, column=2, value=e)

gap2 = gap + len(exclusions) + 2
ws4.cell(row=gap2, column=1, value='Client Responsibilities').font = section_font
responsibilities = [
    'Designate primary POC (Andrea)', 'Jodi + Andrea availability for sessions',
    'Complete product/pricing inventory before Phase 2 CPQ',
    'Contact classification rules and enrichment validation',
    'Deal stage definitions and probability framework',
    'Close won/lost reason categories', 'Mandate CRM usage (Jodi exec sponsorship)',
    'Share 3-5 recent quotes across deal types', 'UAT testing per phase',
    'HubSpot admin access', 'NetSuite admin contact for integration prep',
    'Commission validation requirements',
]
for i, r in enumerate(responsibilities, gap2 + 1):
    ws4.cell(row=i, column=1, value=i - gap2); ws4.cell(row=i, column=2, value=r)

gap3 = gap2 + len(responsibilities) + 2
ws4.cell(row=gap3, column=1, value='Open Items').font = section_font
open_items = [
    'Contact/company/deal record counts (pending access)',
    'Product/pricing inventory for CPQ', 'Service Hub usage',
    'COO identity and admin access', 'Outlook defect root cause',
    'Existing workflow inventory', 'Historical deal data reliability',
    'Microsite form link', 'NetSuite entity structure',
    'SOW signing authority', 'Divi positioning',
]
for i, o in enumerate(open_items, gap3 + 1):
    ws4.cell(row=i, column=1, value=i - gap3); ws4.cell(row=i, column=2, value=o)

for row in ws4.iter_rows(min_row=2, max_col=3):
    for cell in row:
        cell.font = body_font; cell.alignment = wrap_align
ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 85

# SAVE
output_path = '/Users/harbuckconsulting/projects/Project Scoping Tool/Strive_Global/Strive_Full_Overhaul_Estimate_V2.xlsx'
wb.save(output_path)
print(f'Option B V2 saved to: {output_path}')
