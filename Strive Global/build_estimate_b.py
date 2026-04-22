"""Generate Strive Global Full CRM Overhaul Scoping Estimate (Option B)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
body_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', size=10, bold=True)
total_font = Font(name='Arial', size=10, bold=True, color='1F4E79')
rate_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
actuals_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
green_header_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
green_header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
section_font = Font(name='Arial', size=11, bold=True, color='1F4E79')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
currency_fmt = '$#,##0'


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def style_data_cell(ws, row, col, is_alt=False):
    cell = ws.cell(row=row, column=col)
    cell.font = body_font
    cell.border = thin_border
    cell.alignment = wrap_align
    if is_alt:
        cell.fill = alt_fill


def apply_data_styles(ws, data_start, data_end, max_col):
    for r in range(data_start, data_end + 1):
        is_alt = (r - data_start) % 2 == 1
        for c in range(1, max_col + 1):
            style_data_cell(ws, r, c, is_alt)


# ============================================================
# SHEET 1: APPROACH COMPARISON
# ============================================================
ws1 = wb.active
ws1.title = 'Approach Comparison'

ws1.merge_cells('A1:C1')
ws1.cell(row=1, column=1, value='STRIVE Global -- Full CRM Overhaul: Approach Comparison').font = Font(name='Arial', size=14, bold=True, color='1F4E79')

headers = ['Dimension', 'Approach A: Phased (Recommended)', 'Approach B: Concurrent']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header_row(ws1, 3, 3)

comparisons = [
    ('Description',
     'Phase 1: CRM Foundation (weeks 1-6) -- pipeline, contacts, dashboards, automations. Phase 2: CPQ, enrichment, integrations, marketing (weeks 7-14). Each phase delivers standalone value.',
     'All workstreams executed in parallel across 8-10 weeks. Single cutover. Maximum resource intensity.'),
    ('Total Hours (Median)', '216 hrs across two phases', '216 hrs compressed'),
    ('Timeline', '14 weeks (Phase 1: 6 wks + Phase 2: 8 wks)',
     '8-10 weeks'),
    ('Risk Level', 'LOW -- Phase 1 de-risks Phase 2; clean data enables CPQ accuracy',
     'HIGH -- dirty data undermines CPQ and reporting from day one; multiple moving parts'),
    ('Adoption Impact', 'Team absorbs changes gradually; builds confidence before CPQ launch',
     'All-at-once change; higher cognitive load; risk of feature overwhelm'),
    ('Change Management', 'Two training cycles; Phase 1 habits established before Phase 2 adds complexity',
     'Single training push covering all features; harder to absorb'),
    ('Dependency Risk', 'LOW -- Phase 2 depends on Phase 1, but Phase 1 is self-contained',
     'HIGH -- CPQ depends on product library AND clean pipeline AND contact data simultaneously'),
    ('Client Resource Demand', 'Spread over 14 weeks; manageable alongside day-to-day operations',
     'Concentrated in 8-10 weeks; significant time commitment from Andrea and Jodi'),
    ('Board Reporting Value', 'Phase 1 delivers board dashboards by week 6; Phase 2 adds CPQ and advanced reporting by week 14',
     'Full reporting suite available by week 10 if everything goes to plan'),
    ('Cost', '$32,400 (median at $150/hr)', '$32,400 (median at $150/hr)'),
    ('Budget Flexibility', 'Can pause after Phase 1 ($17,850) if budget constrained; Phase 2 becomes a separate engagement',
     'Full commitment required upfront; no natural pause point'),
]

for i, (dim, a, b) in enumerate(comparisons, 4):
    ws1.cell(row=i, column=1, value=dim)
    ws1.cell(row=i, column=2, value=a)
    ws1.cell(row=i, column=3, value=b)

apply_data_styles(ws1, 4, 14, 3)

# Recommendation
ws1.merge_cells('A16:C16')
ws1.cell(row=16, column=1, value='Recommendation').font = Font(name='Arial', size=11, bold=True, color='1F4E79')
ws1.merge_cells('A17:C17')
ws1.cell(row=17, column=1, value=(
    'Approach A: Phased is strongly recommended. STRIVE\'s data quality issues make a concurrent '
    'approach risky -- CPQ built on dirty contact and company data will produce inaccurate quotes and '
    'unreliable pipeline reporting. The phased approach also aligns with Jodi\'s stated goal of being '
    '"cost-effective this year while building foundation for future scaling." Phase 1 delivers immediate '
    'operational value (pipeline clarity, board dashboards, clean data), and Phase 2 builds on that '
    'foundation with CPQ, enrichment, and integration prep. If budget is constrained, Phase 1 can '
    'stand alone as a complete engagement.'
)).font = body_font
ws1.cell(row=17, column=1).alignment = wrap_align

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 55
ws1.column_dimensions['C'].width = 55
ws1.freeze_panes = 'A4'

# ============================================================
# SHEET 2: SCOPING ESTIMATE
# ============================================================
ws2 = wb.create_sheet('Scoping Estimate')

ws2.cell(row=1, column=1, value='Hourly Rate:').font = bold_font
ws2.cell(row=1, column=2, value=150).font = bold_font
ws2.cell(row=1, column=2).fill = rate_fill
ws2.cell(row=1, column=2).number_format = currency_fmt

headers2 = ['Workstream', 'Description', 'Min Hours', 'Max Hours', 'Median Hours',
            'Rate', 'Min Cost', 'Max Cost', 'Median Cost', 'Notes / Assumptions',
            'Actual Hours', 'Actual Cost', 'Variance']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, 13)

for col in [11, 12, 13]:
    cell = ws2.cell(row=3, column=col)
    cell.fill = green_header_fill
    cell.font = green_header_font

# Phase 1 label
ws2.cell(row=4, column=1, value='PHASE 1: CRM FOUNDATION (Weeks 1-6)').font = Font(name='Arial', size=10, bold=True, color='1F4E79')
ws2.merge_cells('A4:J4')

workstreams_p1 = [
    ('CRM Architecture & Config Review',
     'Portal audit, user roles/permissions, team structure reconfiguration for existing Enterprise instance',
     6, 12,
     'Assumes existing HubSpot Sales Hub Enterprise. Up to 23 seat users.'),
    ('Pipeline Restructuring',
     'Separate pre-pipeline from active pipeline. Partner-level vs. client-level deal categories. Stage probability mapping. Required fields per stage. Add Contracting stage.',
     10, 18,
     'Two deal categories in same pipeline. Jodi\'s probability framework as starting point.'),
    ('Contact & Company Classification + Cleanup',
     'Contact type taxonomy (client/prospect/broker/partner). AI enrichment using HubSpot Breeze credits. Normalization, deduplication, association fixes.',
     14, 24,
     '10 years of uncoded data. +30-50% data quality premium. Pilot enrichment on 500 records first.'),
    ('Deal Health Scoring & Forecasting',
     'Deal score: stage + time-in-stage + activity recency. Healthy/at-risk/stale classification. Standardized close won/lost reasons.',
     6, 12,
     'Scoring formula defined collaboratively with Andrea and Jodi.'),
    ('Reporting & Dashboards',
     'Partner dashboard (clonable, quick filters). Seller KPI dashboard. CEO/board dashboard. Up to 12 custom reports.',
     10, 18,
     'Covers all metrics Andrea identified: open deals, pipeline value, avg deals/week, etc.'),
    ('Automations & Workflows',
     'Fix Outlook email logging. Lead source tracking. Form-to-CRM for microsite capture. Stage-change notifications. Duplicate protection.',
     8, 16,
     'Up to 8 automated workflows.'),
    ('Email & Communication Fix',
     'Outlook integration audit/fix. Email association rules. Domain exclusions.',
     4, 8,
     'Root cause diagnosis week 1.'),
    ('Training: Admin',
     'Andrea + admin. Pipeline management, reporting, workflow admin, data hygiene.',
     4, 8,
     '1-2 sessions with recording.'),
    ('Training: End Users',
     'Sales team. Daily CRM usage, deal entry, activity logging, dashboards.',
     6, 10,
     '2-3 sessions.'),
    ('Training: Executive',
     'Jodi. Board dashboard, forecast consumption, scheduling.',
     2, 4,
     '1 session.'),
    ('UAT, Go-Live & Documentation (Phase 1)',
     'Test scripts, 2 UAT sessions, bug fixes, admin guide, user guide, 2-week hypercare.',
     6, 10,
     'Phase 1 go-live and hypercare.'),
    ('Project Management (Phase 1)',
     'Kickoff, weekly syncs, status updates.',
     10, 16,
     '~12% of Phase 1 hours.'),
]

data_start = 5
for i, (name, desc, min_h, max_h, notes) in enumerate(workstreams_p1, data_start):
    row = i
    ws2.cell(row=row, column=1, value=name)
    ws2.cell(row=row, column=2, value=desc)
    ws2.cell(row=row, column=3, value=min_h)
    ws2.cell(row=row, column=4, value=max_h)
    ws2.cell(row=row, column=5).value = f'=AVERAGE(C{row},D{row})'
    ws2.cell(row=row, column=6).value = f'=$B$1'
    ws2.cell(row=row, column=7).value = f'=C{row}*$B$1'
    ws2.cell(row=row, column=8).value = f'=D{row}*$B$1'
    ws2.cell(row=row, column=9).value = f'=E{row}*$B$1'
    ws2.cell(row=row, column=10, value=notes)
    for col in [11, 12, 13]:
        ws2.cell(row=row, column=col).fill = actuals_fill
    ws2.cell(row=row, column=13).value = f'=IF(K{row}="","",K{row}-E{row})'

p1_end = data_start + len(workstreams_p1) - 1

# Phase 1 subtotal
p1_total = p1_end + 1
ws2.cell(row=p1_total, column=1, value='Phase 1 Subtotal').font = total_font
for col in [3, 4, 5, 7, 8, 9, 11, 12]:
    cl = get_column_letter(col)
    ws2.cell(row=p1_total, column=col).value = f'=SUM({cl}{data_start}:{cl}{p1_end})'
    ws2.cell(row=p1_total, column=col).font = total_font
    ws2.cell(row=p1_total, column=col).border = thin_border
ws2.cell(row=p1_total, column=6).value = '=$B$1'

# Phase 2 label
p2_label = p1_total + 2
ws2.cell(row=p2_label, column=1, value='PHASE 2: CPQ, ENRICHMENT & INTEGRATIONS (Weeks 7-14)').font = Font(name='Arial', size=10, bold=True, color='1F4E79')
ws2.merge_cells(f'A{p2_label}:J{p2_label}')

workstreams_p2 = [
    ('CPQ Implementation',
     'Product library build (apps, microsites, AI assistant, broker packs, Divi, bundles). Quote template design. E-signature config. Automated quote reminders. Multiple pricing models.',
     12, 24,
     'Requires complete product/pricing inventory from Andrea. License+rev share, flat per-head, bundles/discounts.'),
    ('Contract Management Workflows',
     'Quote-to-contract automation. Contract status tracking. Renewal and expiry workflows.',
     6, 12,
     'Replaces manual DocuSign process through COO.'),
    ('Data Enrichment at Scale',
     'HubSpot Breeze AI credits (5-10K). Enrichment rules for auto-classification on inbound. QA and validation of results.',
     6, 12,
     'Extends Phase 1 pilot enrichment to full database.'),
    ('Data Normalization & Dedup (Extended)',
     'Full database normalization. Advanced deduplication rules. Merge strategy. Ongoing duplicate protection.',
     8, 16,
     'Extends Phase 1 basic cleanup to comprehensive database overhaul.'),
    ('Parent-Child Structure (Broker Packs)',
     'Association labels. Per-broker intake forms. Automated child contact creation from form submissions. List generation by parent broker.',
     4, 8,
     'Enables microsite-to-app expansion prospecting. Depends on form link from Andrea.'),
    ('Email Sequencing',
     'Broker prospecting sequences. Microsite-to-app expansion outreach. Up to 4 sequences.',
     4, 8,
     'Using HubSpot Sales Hub Enterprise sequences (included in current license).'),
    ('Marketing Automation Foundation',
     'Newsletter template. 2-3 segmented campaigns by contact type (client, broker, prospect). Unsubscribe and compliance setup.',
     6, 12,
     'Leverages Marketing Hub (already licensed). CAN-SPAM compliance.'),
    ('Commission Visibility & Validation',
     'CRM-based commission view tied to closed-won deals. Validation workflow for Jodi before release. Replaces manual spreadsheet reconciliation.',
     4, 8,
     'View and validation only -- not full commission calculation engine.'),
    ('NetSuite Integration Prep',
     'Data cleanup to enable future API connection. Field mapping documentation. Data readiness assessment.',
     6, 12,
     'PREP ONLY. Actual integration build requires separate SOW with NetSuite admin involvement.'),
    ('Bill.com Integration Prep',
     'AR/AP field mapping. Data readiness assessment for future integration.',
     4, 8,
     'PREP ONLY. Actual integration build requires separate SOW.'),
    ('Seat/License Optimization',
     'Audit 23 current seats (3 core, 10 sales, 10 service). Identify unused or misallocated seats. Cost savings memo.',
     2, 4,
     'Deliverable: cost savings recommendation with specific seat changes.'),
    ('Advanced Reporting',
     'Automated board reports (scheduled delivery). Pipeline coverage vs. quota. Pace-to-goal tracking. Marketing ROI reporting.',
     6, 12,
     'Extends Phase 1 dashboards with CPQ and marketing data.'),
    ('Additional Project Management',
     'Phase 2 PM: weekly syncs, milestone reviews, change management.',
     8, 14,
     '~12% of Phase 2 hours.'),
]

p2_start = p2_label + 1
for i, (name, desc, min_h, max_h, notes) in enumerate(workstreams_p2, p2_start):
    row = i
    ws2.cell(row=row, column=1, value=name)
    ws2.cell(row=row, column=2, value=desc)
    ws2.cell(row=row, column=3, value=min_h)
    ws2.cell(row=row, column=4, value=max_h)
    ws2.cell(row=row, column=5).value = f'=AVERAGE(C{row},D{row})'
    ws2.cell(row=row, column=6).value = f'=$B$1'
    ws2.cell(row=row, column=7).value = f'=C{row}*$B$1'
    ws2.cell(row=row, column=8).value = f'=D{row}*$B$1'
    ws2.cell(row=row, column=9).value = f'=E{row}*$B$1'
    ws2.cell(row=row, column=10, value=notes)
    for col in [11, 12, 13]:
        ws2.cell(row=row, column=col).fill = actuals_fill
    ws2.cell(row=row, column=13).value = f'=IF(K{row}="","",K{row}-E{row})'

p2_end = p2_start + len(workstreams_p2) - 1

# Phase 2 subtotal
p2_total = p2_end + 1
ws2.cell(row=p2_total, column=1, value='Phase 2 Subtotal').font = total_font
for col in [3, 4, 5, 7, 8, 9, 11, 12]:
    cl = get_column_letter(col)
    ws2.cell(row=p2_total, column=col).value = f'=SUM({cl}{p2_start}:{cl}{p2_end})'
    ws2.cell(row=p2_total, column=col).font = total_font
    ws2.cell(row=p2_total, column=col).border = thin_border
ws2.cell(row=p2_total, column=6).value = '=$B$1'

# Optional add-on
addon_label = p2_total + 2
ws2.cell(row=addon_label, column=1, value='OPTIONAL ADD-ON').font = Font(name='Arial', size=10, bold=True, color='548235')
ws2.merge_cells(f'A{addon_label}:J{addon_label}')

addon_row = addon_label + 1
ws2.cell(row=addon_row, column=1, value='Client Success Pipeline (Optional)')
ws2.cell(row=addon_row, column=2, value='Client health tracking (replace Excel). Health score properties. Renewal tracking. Requires separate discovery with CS team.')
ws2.cell(row=addon_row, column=3, value=6)
ws2.cell(row=addon_row, column=4, value=10)
ws2.cell(row=addon_row, column=5).value = f'=AVERAGE(C{addon_row},D{addon_row})'
ws2.cell(row=addon_row, column=6).value = '=$B$1'
ws2.cell(row=addon_row, column=7).value = f'=C{addon_row}*$B$1'
ws2.cell(row=addon_row, column=8).value = f'=D{addon_row}*$B$1'
ws2.cell(row=addon_row, column=9).value = f'=E{addon_row}*$B$1'
ws2.cell(row=addon_row, column=10, value='Requires separate discovery call with Client Success team. Not included in project total.')
for col in [11, 12, 13]:
    ws2.cell(row=addon_row, column=col).fill = actuals_fill

# Grand total
grand_row = addon_row + 2
ws2.cell(row=grand_row, column=1, value='PROJECT TOTAL (Phase 1 + Phase 2)').font = Font(name='Arial', size=11, bold=True, color='1F4E79')
for col in [3, 4, 5, 7, 8, 9]:
    cl = get_column_letter(col)
    ws2.cell(row=grand_row, column=col).value = f'={cl}{p1_total}+{cl}{p2_total}'
    ws2.cell(row=grand_row, column=col).font = Font(name='Arial', size=11, bold=True, color='1F4E79')
    ws2.cell(row=grand_row, column=col).border = thin_border
ws2.cell(row=grand_row, column=6).value = '=$B$1'

# Apply formatting
all_data_rows = list(range(data_start, p1_end + 1)) + list(range(p2_start, p2_end + 1)) + [addon_row]
for idx, row in enumerate(all_data_rows):
    is_alt = idx % 2 == 1
    for c in range(1, 11):
        style_data_cell(ws2, row, c, is_alt)
    for col in [11, 12, 13]:
        ws2.cell(row=row, column=col).fill = actuals_fill
        ws2.cell(row=row, column=col).border = thin_border
        ws2.cell(row=row, column=col).font = body_font

for col in [7, 8, 9, 12]:
    for row in all_data_rows + [p1_total, p2_total, grand_row, addon_row]:
        ws2.cell(row=row, column=col).number_format = currency_fmt

widths = {1: 35, 2: 55, 3: 12, 4: 12, 5: 14, 6: 10, 7: 12, 8: 12, 9: 14, 10: 55, 11: 14, 12: 14, 13: 12}
for col, w in widths.items():
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.freeze_panes = 'B4'

# ============================================================
# SHEET 3: RISK REGISTER
# ============================================================
ws3 = wb.create_sheet('Risk Register')

headers3 = ['#', 'Risk', 'Severity', 'Likelihood', 'Impact', 'Mitigation', 'Owner', 'Status']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header_row(ws3, 1, 8)

sev_high_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
sev_med_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
sev_low_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

risks = [
    (1, '10 years of uncoded contact data -- AI enrichment scope may exceed estimates',
     'HIGH', 'HIGH',
     'Data cleanup takes 2-3x estimate; CPQ product associations unreliable; marketing segmentation delayed',
     'Run AI enrichment pilot on 500 records in Phase 1. Measure accuracy before Phase 2 full run. Budget 30% buffer.',
     'Consultant + Andrea', 'Open'),
    (2, 'Pipeline under-reporting -- reps using LinkedIn and manual quoting outside CRM',
     'HIGH', 'HIGH',
     'Forecasting unreliable during transition; board sees pipeline dip before improvement; CPQ adoption depends on reps entering deals',
     'Jodi mandates CRM usage pre-launch. Set expectations. Brief board that pipeline will increase as capture improves.',
     'Jodi (executive sponsor)', 'Open'),
    (3, 'Scope creep from close business relationship',
     'HIGH', 'MEDIUM',
     'Informal requests bypass SOW; hours consumed without authorization; budget overrun on 14-week engagement',
     'SOW signed by budget authority. All changes documented. Weekly hours tracking visible to client.',
     'Consultant + Jodi', 'Open'),
    (4, 'Board ROI expectations overpromise',
     'HIGH', 'MEDIUM',
     'Board approves but expects unrealistic timeline or savings; Phase 2 pressure to deliver faster',
     'Conservative ROI estimates with ranges. 90-day KPI check-in after Phase 1. Phase 2 ROI measured separately.',
     'Consultant + Jodi', 'Open'),
    (5, 'CPQ product library complexity -- multiple pricing models (license+rev share, flat, bundles)',
     'MEDIUM', 'MEDIUM',
     'HubSpot CPQ may not handle all pricing models natively; partial manual process remains; quote accuracy issues',
     'Audit 5 recent quotes across deal types before building product library. Confirm Enterprise CPQ handles rev share structures. Identify edge cases early.',
     'Consultant', 'Open'),
    (6, 'Seat/license confusion -- mixed versions, potential unused seats',
     'MEDIUM', 'HIGH',
     'Paying for unused seats; features expected but unavailable; CPQ features may require specific seat types',
     'Complete license audit in week 1. Produce cost savings memo. Confirm CPQ capabilities on current Enterprise tier.',
     'Consultant + Jodi', 'Open'),
    (7, 'NetSuite/Bill.com integration scope bleed -- prep work reveals integration is harder than expected',
     'MEDIUM', 'MEDIUM',
     'Client pushes for actual integration build without additional budget; prep-only boundary violated',
     'Scope boundary explicitly states "prep only." Integration build requires separate SOW with NetSuite admin involvement. Document boundary in kickoff.',
     'Consultant', 'Open'),
    (8, 'Adoption risk -- sales reps prefer LinkedIn Sales Navigator over CRM',
     'MEDIUM', 'HIGH',
     'New CRM and CPQ configuration goes unused; ROI unrealized; board investment wasted',
     'Executive mandate from Jodi. Make CRM easier than alternatives. Rep-level dashboards. Email sequences reduce manual prospecting effort.',
     'Jodi + Andrea', 'Open'),
    (9, 'Outlook email logging defect -- emails logging to all associated deals',
     'LOW', 'MEDIUM',
     'Fix may require HubSpot support escalation; workaround needed; email sequences may compound the issue',
     'Diagnose root cause in week 1. If platform bug, open HubSpot support ticket. Email sequences designed to avoid multi-deal logging.',
     'Consultant', 'Open'),
    (10, 'COO identity ambiguity (Lauren vs. Zach Beegal)',
     'LOW', 'LOW',
     'Miscommunication on stakeholder authority; admin access issues',
     'Clarify org chart at kickoff.',
     'Consultant', 'Open'),
]

for i, (num, risk, sev, lik, impact, mit, owner, status) in enumerate(risks, 2):
    ws3.cell(row=i, column=1, value=num)
    ws3.cell(row=i, column=2, value=risk)
    ws3.cell(row=i, column=3, value=sev)
    ws3.cell(row=i, column=4, value=lik)
    ws3.cell(row=i, column=5, value=impact)
    ws3.cell(row=i, column=6, value=mit)
    ws3.cell(row=i, column=7, value=owner)
    ws3.cell(row=i, column=8, value=status)

    sev_cell = ws3.cell(row=i, column=3)
    if sev == 'HIGH':
        sev_cell.fill = sev_high_fill
    elif sev == 'MEDIUM':
        sev_cell.fill = sev_med_fill
    else:
        sev_cell.fill = sev_low_fill

apply_data_styles(ws3, 2, len(risks) + 1, 8)
for i, (_, _, sev, *_rest) in enumerate(risks, 2):
    sev_cell = ws3.cell(row=i, column=3)
    if sev == 'HIGH':
        sev_cell.fill = sev_high_fill
    elif sev == 'MEDIUM':
        sev_cell.fill = sev_med_fill
    else:
        sev_cell.fill = sev_low_fill

widths3 = {1: 5, 2: 50, 3: 12, 4: 12, 5: 45, 6: 55, 7: 25, 8: 10}
for col, w in widths3.items():
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.freeze_panes = 'A2'

# ============================================================
# SHEET 4: ASSUMPTIONS & EXCLUSIONS
# ============================================================
ws4 = wb.create_sheet('Assumptions & Exclusions')

ws4.cell(row=1, column=1, value='Scope Assumptions').font = section_font
assumptions = [
    'HubSpot Sales Hub Enterprise (existing instance -- reconfiguration, not net-new)',
    'Up to 23 licensed users across 3 core, 10 sales, and 10 service seats',
    'Outlook as primary email platform',
    'Up to 40 custom contact/company properties (expanded for CPQ and enrichment)',
    '1 active sales pipeline with pre-pipeline separated',
    '2 deal categories: partner-level and client-level',
    'Stage probability mapping using Jodi\'s framework',
    'Contact type classification: client, prospect, broker, partner',
    'AI enrichment using HubSpot Breeze credits (5-10K credits)',
    'Up to 12 automated workflows (expanded for CPQ and marketing)',
    '4-5 reporting dashboards with up to 20 custom reports',
    'Complete product/pricing inventory provided by Andrea before CPQ build',
    'HubSpot Marketing Hub leveraged for newsletter and campaign automation',
    '10-14 week implementation timeline (phased)',
    'Phase 1 can stand alone if Phase 2 is deferred',
    'All work conducted remotely unless otherwise agreed',
    'Any scope changes managed through formal change request',
]
for i, a in enumerate(assumptions, 2):
    ws4.cell(row=i, column=1, value=i - 1)
    ws4.cell(row=i, column=2, value=a)

gap = len(assumptions) + 3
ws4.cell(row=gap, column=1, value='Out of Scope').font = section_font
exclusions = [
    'NetSuite integration build (prep only -- actual build requires separate SOW)',
    'Bill.com integration build (prep only -- actual build requires separate SOW)',
    'Client Success pipeline (available as optional add-on; requires CS team discovery)',
    'Full commission calculation engine (visibility and validation only)',
    'Custom API development or middleware beyond HubSpot native integrations',
    'Salesforce or any CRM other than HubSpot',
    'Data migration from external systems (internal HubSpot cleanup only)',
    'Ongoing managed services beyond 2-week hypercare per phase',
    'HubSpot license procurement or tier changes (recommendations provided)',
    'Hardware, IT infrastructure, or network changes',
    'AI knowledge assistant for sales onboarding (future initiative)',
    'LinkedIn Sales Navigator integration (recommended for future)',
]
for i, e in enumerate(exclusions, gap + 1):
    ws4.cell(row=i, column=1, value=i - gap)
    ws4.cell(row=i, column=2, value=e)

gap2 = gap + len(exclusions) + 2
ws4.cell(row=gap2, column=1, value='Client Responsibilities').font = section_font
responsibilities = [
    'Designate primary point of contact (Andrea recommended)',
    'Ensure Jodi and Andrea availability for discovery, UAT, and training sessions',
    'Provide complete product/pricing inventory with all SKUs, bundles, and pricing models before Phase 2 CPQ build',
    'Provide contact classification rules and validate AI enrichment results',
    'Confirm deal stage definitions and probability framework',
    'Communicate standardized close won/lost reason categories',
    'Mandate CRM usage across sales team (executive sponsorship from Jodi)',
    'Share 3-5 recent quotes/proposals across deal types for CPQ reference',
    'Complete UAT testing within agreed timeline for each phase',
    'Provide access to HubSpot admin portal',
    'Identify NetSuite admin contact for integration prep coordination',
    'Provide commission validation requirements and current spreadsheet structure',
]
for i, r in enumerate(responsibilities, gap2 + 1):
    ws4.cell(row=i, column=1, value=i - gap2)
    ws4.cell(row=i, column=2, value=r)

gap3 = gap2 + len(responsibilities) + 2
ws4.cell(row=gap3, column=1, value='Open Items').font = section_font
open_items = [
    'Exact contact/company/deal record counts (pending portal access)',
    'Complete product/pricing inventory for CPQ build',
    'Service Hub usage details -- preserve or deprecate?',
    'COO identity and HubSpot admin access ownership',
    'Outlook email logging defect root cause',
    'Existing workflow inventory',
    'Historical deal data reliability for probability calibration',
    'Microsite client intake form link',
    'NetSuite entity structure and admin contact',
    'SOW signing authority (CEO or board)',
    'Divi product positioning (standalone vs. bundled)',
]
for i, o in enumerate(open_items, gap3 + 1):
    ws4.cell(row=i, column=1, value=i - gap3)
    ws4.cell(row=i, column=2, value=o)

for row in ws4.iter_rows(min_row=2, max_col=3):
    for cell in row:
        cell.font = body_font
        cell.alignment = wrap_align

ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 85
ws4.column_dimensions['C'].width = 30

# ============================================================
# SAVE
# ============================================================
output_path = '/Users/harbuckconsulting/projects/Project Scoping Tool/Strive_Global/Strive_Full_Overhaul_Estimate.xlsx'
wb.save(output_path)
print(f'Option B saved to: {output_path}')
