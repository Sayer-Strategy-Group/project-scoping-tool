"""Generate Strive Global V2 Side-by-Side Comparison -- Single Sheet"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Option A vs Option B'

# Styles
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
body_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', size=10, bold=True)
total_font = Font(name='Arial', size=11, bold=True, color='1F4E79')
section_font = Font(name='Arial', size=10, bold=True, color='1F4E79')
alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
b_only_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap = Alignment(wrap_text=True, vertical='top')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
currency_fmt = '$#,##0'

# Title
ws.merge_cells('A1:F1')
ws.cell(row=1, column=1, value='STRIVE Global -- V2 Scope Comparison: Option A vs Option B').font = Font(name='Arial', size=14, bold=True, color='1F4E79')

# Headers
headers = ['Workstream', 'Option A\nMedian Hrs', 'Option A\nMedian Cost', 'Option B\nMedian Hrs', 'Option B\nMedian Cost', 'Difference']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

# Data rows -- shared workstreams first, then B-only
rows_data = [
    # (workstream, a_hrs, b_hrs, is_section_label, is_b_only)
    ('SHARED WORKSTREAMS (in both options)', None, None, True, False),
    ('CRM Architecture & Config Review', 12, 12, False, False),
    ('Pipeline Restructuring', 24, 24, False, False),
    ('Contact & Company Classification + Cleanup', 24, 24, False, False),
    ('Deal Health Scoring & Forecasting', 10, 10, False, False),
    ('Reporting & Dashboards', 20, 20, False, False),
    ('Automations & Workflows', 14, 14, False, False),
    ('Email & Communication Fix', 6, 6, False, False),
    ('Training: Admin', 6, 6, False, False),
    ('Training: End Users', 8, 8, False, False),
    ('Training: Executive', 3, 3, False, False),
    ('UAT, Go-Live & Documentation', 10, 10, False, False),
    ('Project Management (Phase 1)', 18, 18, False, False),
    ('', None, None, True, False),
    ('OPTION B ONLY (Phase 2 additions)', None, None, True, False),
    ('CPQ Implementation', None, 45, False, True),
    ('Contract Management Workflows', None, 11, False, True),
    ('Data Enrichment at Scale', None, 12, False, True),
    ('Data Normalization & Dedup (Extended)', None, 14, False, True),
    ('Parent-Child Structure (Broker Packs)', None, 8, False, True),
    ('Email Sequencing', None, 8, False, True),
    ('Marketing Automation Foundation', None, 11, False, True),
    ('Commission Visibility & Validation', None, 6, False, True),
    ('NetSuite Integration Prep', None, 11, False, True),
    ('Bill.com Integration Prep', None, 6, False, True),
    ('Seat/License Optimization', None, 3, False, True),
    ('Advanced Reporting', None, 11, False, True),
    ('Additional Project Management (Phase 2)', None, 16, False, True),
]

rate = 150
r = 4
data_rows = []
for (name, a_hrs, b_hrs, is_section, is_b_only) in rows_data:
    if is_section:
        if name:
            ws.cell(row=r, column=1, value=name).font = section_font
            ws.merge_cells(f'A{r}:F{r}')
        r += 1
        continue

    ws.cell(row=r, column=1, value=name).font = body_font
    ws.cell(row=r, column=1).alignment = wrap
    ws.cell(row=r, column=1).border = thin_border

    if a_hrs is not None:
        ws.cell(row=r, column=2, value=a_hrs).font = body_font
        ws.cell(row=r, column=3, value=a_hrs * rate).font = body_font
        ws.cell(row=r, column=3).number_format = currency_fmt
    else:
        ws.cell(row=r, column=2, value='--').font = body_font
        ws.cell(row=r, column=3, value='--').font = body_font

    if b_hrs is not None:
        ws.cell(row=r, column=4, value=b_hrs).font = body_font
        ws.cell(row=r, column=5, value=b_hrs * rate).font = body_font
        ws.cell(row=r, column=5).number_format = currency_fmt
    else:
        ws.cell(row=r, column=4, value='--').font = body_font
        ws.cell(row=r, column=5, value='--').font = body_font

    # Difference
    if a_hrs is not None and b_hrs is not None:
        diff = b_hrs - a_hrs
        ws.cell(row=r, column=6, value=f'+{diff} hrs' if diff > 0 else ('--' if diff == 0 else f'{diff} hrs'))
    elif b_hrs is not None:
        ws.cell(row=r, column=6, value=f'+{b_hrs} hrs')
    else:
        ws.cell(row=r, column=6, value='--')
    ws.cell(row=r, column=6).font = body_font

    # Styling
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = thin_border
        ws.cell(row=r, column=c).alignment = Alignment(horizontal='center' if c > 1 else 'left', vertical='center', wrap_text=True)

    if is_b_only:
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = b_only_fill
    elif len(data_rows) % 2 == 1:
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = alt_fill

    data_rows.append(r)
    r += 1

# Totals
r += 1
ws.cell(row=r, column=1, value='TOTALS').font = total_font
ws.cell(row=r, column=1).border = thin_border

a_total_hrs = 155
b_total_hrs = 313
for col, val in [(2, a_total_hrs), (3, a_total_hrs * rate), (4, b_total_hrs), (5, b_total_hrs * rate)]:
    c = ws.cell(row=r, column=col, value=val)
    c.font = total_font
    c.border = thin_border
    c.alignment = center
    if col in [3, 5]:
        c.number_format = currency_fmt

ws.cell(row=r, column=6, value=f'+{b_total_hrs - a_total_hrs} hrs').font = total_font
ws.cell(row=r, column=6).border = thin_border
ws.cell(row=r, column=6).alignment = center

# Summary section below
r += 2
summary = [
    ('', 'Option A: CRM Foundation', 'Option B: Full Overhaul'),
    ('Median Hours', '155', '313'),
    ('Median Cost (@$150/hr)', '$23,250', '$46,950'),
    ('Timeline', '8-10 weeks', '16-18 weeks (phased)'),
    ('Payback Period', '5-7 months', '4-7 months'),
    ('Includes CPQ', 'No', 'Yes (45 hrs)'),
    ('Includes Marketing', 'No', 'Yes (11 hrs)'),
    ('Includes Integration Prep', 'No', 'Yes (NetSuite + Bill.com)'),
    ('Includes Commission Tracking', 'No', 'Yes (visibility + validation)'),
    ('Can Stand Alone', 'Yes', 'Yes (Phase 1 = Option A)'),
    ('Phase 2 Deferrable', 'N/A', 'Yes -- Phase 1 delivers value independently'),
]

ws.cell(row=r, column=1, value='SUMMARY COMPARISON').font = Font(name='Arial', size=12, bold=True, color='1F4E79')
ws.merge_cells(f'A{r}:C{r}')
r += 1

for i, (label, a_val, b_val) in enumerate(summary):
    ws.cell(row=r, column=1, value=label).font = bold_font if i == 0 else body_font
    ws.cell(row=r, column=2, value=a_val).font = bold_font if i == 0 else body_font
    ws.cell(row=r, column=3, value=b_val).font = bold_font if i == 0 else body_font
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = thin_border
        ws.cell(row=r, column=c).alignment = Alignment(vertical='center', wrap_text=True)
    if i == 0:
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = header_fill
            ws.cell(row=r, column=c).font = header_font
    elif i % 2 == 0:
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = alt_fill
    r += 1

# Column widths
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 14
ws.freeze_panes = 'A4'

# Legend
r += 1
ws.cell(row=r, column=1, value='Green rows = Option B only (not included in Option A)').font = Font(name='Arial', size=9, italic=True, color='548235')

output = '/Users/harbuckconsulting/projects/Project Scoping Tool/Strive_Global/Strive_V2_Comparison.xlsx'
wb.save(output)
print(f'Comparison saved to: {output}')
