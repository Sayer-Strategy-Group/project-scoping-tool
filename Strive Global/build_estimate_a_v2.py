"""Generate Strive Global CRM Foundation Scoping Estimate V2 (Option A) - Wagmo-calibrated"""
import sys
import pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1] / 'scripts'))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from brand_styles import (
    HEADER_FONT, HEADER_FILL, BODY_FONT, BOLD_BODY_FONT, TOTAL_FONT,
    INPUT_FILL, ALT_ROW_FILL, SECONDARY_HEADER_FONT, SECONDARY_HEADER_FILL,
    THIN_BORDER, WRAP_ALIGN, TITLE_FONT, SECTION_FONT, CURRENCY_FMT,
    FONT_FAMILY, SAYER_BLACK,
    style_header_row, apply_data_styles,
    get_column_letter,
)

wb = openpyxl.Workbook()

# Backwards-compat aliases — brand specs live in scripts/brand_styles.py.
header_font = HEADER_FONT
header_fill = HEADER_FILL
body_font = BODY_FONT
bold_font = BOLD_BODY_FONT
total_font = TOTAL_FONT
rate_fill = INPUT_FILL
alt_fill = ALT_ROW_FILL
actuals_fill = INPUT_FILL
green_header_fill = SECONDARY_HEADER_FILL
green_header_font = SECONDARY_HEADER_FONT
thin_border = THIN_BORDER
wrap_align = WRAP_ALIGN
currency_fmt = CURRENCY_FMT
section_font = SECTION_FONT


# ============================================================
# SHEET 1: APPROACH COMPARISON
# ============================================================
ws1 = wb.active
ws1.title = 'Approach Comparison'

ws1.merge_cells('A1:C1')
ws1.cell(row=1, column=1, value='STRIVE Global -- CRM Foundation V2: Approach Comparison').font = TITLE_FONT

headers = ['Dimension', 'Approach A: Foundation First (Recommended)', 'Approach B: Quick Wins First']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header_row(ws1, 3, 3)

comparisons = [
    ('Description',
     'Data cleanup and contact classification first (weeks 1-5), then pipeline restructuring and dashboards (weeks 6-10). Build on clean foundation.',
     'Pipeline restructuring and dashboards first (weeks 1-5), then data cleanup and enrichment (weeks 6-10). Deliver visible wins early.'),
    ('Total Hours (Median)', '155 hrs', '155 hrs'),
    ('Timeline', '8-10 weeks', '8-10 weeks'),
    ('Risk Level', 'LOW -- clean data enables accurate dashboards from day one',
     'MEDIUM -- dashboards built on dirty data show misleading metrics initially'),
    ('Adoption Impact', 'Slower initial visibility but higher trust in data when dashboards launch',
     'Faster dashboard delivery but reps may distrust inaccurate early reports'),
    ('Change Management', 'Gradual -- team sees data improving before process changes',
     'Immediate -- team gets new pipeline and dashboards right away'),
    ('Dependency Risk', 'LOW -- data cleanup is independent of other workstreams',
     'MEDIUM -- dashboards depend on data quality for accuracy'),
    ('Client Resource Demand', 'Higher upfront (contact exports, classification rules, validation)',
     'Lower upfront but higher mid-project for data cleanup'),
    ('Board Reporting Value', 'Clean, trustworthy reports available by week 8',
     'Reports available by week 5 but accuracy improves through week 10'),
    ('Cost', '$23,250 (median at $150/hr)', '$23,250 (median at $150/hr)'),
]

for i, (dim, a, b) in enumerate(comparisons, 4):
    ws1.cell(row=i, column=1, value=dim)
    ws1.cell(row=i, column=2, value=a)
    ws1.cell(row=i, column=3, value=b)

apply_data_styles(ws1, 4, 13, 3)

ws1.merge_cells('A15:C15')
ws1.cell(row=15, column=1, value='Recommendation').font = section_font
ws1.merge_cells('A16:C16')
ws1.cell(row=16, column=1, value=(
    'Approach A: Foundation First is recommended. STRIVE has 10 years of uncoded contact data and '
    'unreliable pipeline metrics. Building dashboards on dirty data risks eroding trust in the CRM. '
    'Clean the data first, then deliver dashboards the team and board can trust from day one. This '
    'also aligns with Jodi\'s emphasis on building a strong data infrastructure as the foundation '
    'for future AI capabilities.'
)).font = body_font
ws1.cell(row=16, column=1).alignment = wrap_align

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 50
ws1.column_dimensions['C'].width = 50
ws1.freeze_panes = 'A4'

# ============================================================
# SHEET 2: SCOPING ESTIMATE
# ============================================================
ws2 = wb.create_sheet('Scoping Estimate')

ws2.cell(row=1, column=1, value='Hourly Rate:').font = bold_font
ws2.cell(row=1, column=2, value=150).font = bold_font
ws2.cell(row=1, column=2).fill = rate_fill
ws2.cell(row=1, column=2).number_format = currency_fmt

headers2 = ['Workstream', 'Description', 'Min Hours', 'Max Hours', 'Median Hours',
            'Rate', 'Min Cost', 'Max Cost', 'Median Cost', 'Notes / Assumptions',
            'Actual Hours', 'Actual Cost', 'Variance']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, 13)

for col in [11, 12, 13]:
    cell = ws2.cell(row=3, column=col)
    cell.fill = green_header_fill
    cell.font = green_header_font

# V2 ESTIMATES -- Wagmo-calibrated
workstreams = [
    ('CRM Architecture & Config Review',
     'Portal audit, user roles/permissions, team structure reconfig. Seat/license review across 3 hubs (sales, marketing, service). Enterprise instance complexity assessment.',
     8, 16, 'Wagmo comparable: 14 hrs discovery. HubSpot simpler than SFDC but 3-hub Enterprise with 23 seats adds complexity.'),
    ('Pipeline Restructuring',
     'Separate pre-pipeline from active pipeline. Partner-level vs. client-level deal categories. Stage probability mapping (5/20/40/50/60/80/90/100). Required fields per stage. Add Contracting stage. Auto-transition workflows with deal age reset.',
     16, 32,
     'Wagmo benchmark: 36 hrs for simpler Salesforce pipeline. HubSpot offset: -20%. Strive complexity premium: +40% (2 categories, pre-pipeline, probabilities, auto-transitions). Net: ~24 hrs median.'),
    ('Contact & Company Classification + Cleanup',
     'Contact type taxonomy (client/prospect/broker/partner). AI enrichment pilot (500 records) then full run using HubSpot Breeze credits. Normalization, deduplication, association fixes. Duplicate protection rules.',
     16, 32,
     '10 years uncoded data. Industry benchmark: $5K-$15K for cleanup = 33-100 hrs at $150. Conservative estimate with AI acceleration. +30-50% data quality premium.'),
    ('Deal Health Scoring & Forecasting',
     'Deal score: stage + time-in-stage + activity recency. Healthy/at-risk/stale classification. Standardized close won/lost reasons (categorized dropdowns replacing free-form). Forecasting views.',
     6, 14,
     'Scoring formula and thresholds defined collaboratively. Close reason categories defined before implementation.'),
    ('Reporting & Dashboards',
     'Partner dashboard (clonable, quick filters). Seller KPI dashboard. CEO/board dashboard (bookings, pipeline coverage, win rate, deal age). Pipeline health dashboard. Up to 15 custom reports.',
     14, 26,
     'Wagmo benchmark: 20 hrs for 2 dashboards + 6 reports. Strive needs 4 dashboards + 15 reports. AI accelerates report logic but iteration cycles take time.'),
    ('Automations & Workflows',
     'Fix Outlook email logging issue. Lead source tracking automation. Form-to-CRM for microsite client capture. Deal stage-change notifications. Duplicate protection rules. Pre-pipeline auto-transition.',
     10, 18,
     'Up to 10 automated workflows (increased from 8). Outlook fix may require HubSpot support ticket. Includes pre-pipeline transition automation.'),
    ('Email & Communication Fix',
     'Outlook integration audit and fix. Email association rules to prevent logging to all related deals. Domain exclusions for internal communications.',
     4, 8,
     'Root cause diagnosis week 1. If config fix: 2-4 hrs. If platform bug: HubSpot support escalation.'),
    ('Training: Admin',
     'Andrea + designated admin. Pipeline management, reporting, workflow admin, data hygiene, enrichment tools.',
     4, 8,
     'Wagmo training was minimal (2.5 hrs total). Strive needs proper formal sessions. 1-2 sessions with recording.'),
    ('Training: End Users',
     'Sales team. Daily CRM usage, deal entry, activity logging, dashboard consumption, mobile app.',
     6, 10,
     '2-3 sessions. Emphasis on showing reps how CRM data benefits them.'),
    ('Training: Executive',
     'Jodi. Board dashboard walkthrough, forecast consumption, report scheduling, export for board decks.',
     2, 4,
     '1 session focused on board reporting.'),
    ('UAT, Go-Live & Documentation',
     'Test scripts, 2 UAT sessions with key stakeholders, bug fix window, go-live cutover plan. Admin guide + user guide. 2-week hypercare.',
     8, 12,
     'Wagmo testing: 7 hrs. Add documentation deliverables and hypercare setup. Formal test scripts not done at Wagmo.'),
    ('Project Management',
     'Kickoff, weekly syncs, status updates, change management, milestone reviews. Scope creep prevention.',
     14, 22,
     '~12% of total. Wagmo lesson: zero formal PM led to scope creep. This prevents that.'),
]

for i, (name, desc, min_h, max_h, notes) in enumerate(workstreams, 4):
    row = i
    ws2.cell(row=row, column=1, value=name)
    ws2.cell(row=row, column=2, value=desc)
    ws2.cell(row=row, column=3, value=min_h)
    ws2.cell(row=row, column=4, value=max_h)
    ws2.cell(row=row, column=5).value = f'=AVERAGE(C{row},D{row})'
    ws2.cell(row=row, column=6).value = f'=$B$1'
    ws2.cell(row=row, column=7).value = f'=C{row}*$B$1'
    ws2.cell(row=row, column=8).value = f'=D{row}*$B$1'
    ws2.cell(row=row, column=9).value = f'=E{row}*$B$1'
    ws2.cell(row=row, column=10, value=notes)
    for col in [11, 12, 13]:
        ws2.cell(row=row, column=col).fill = actuals_fill
    ws2.cell(row=row, column=13).value = f'=IF(K{row}="","",K{row}-E{row})'

total_row = 4 + len(workstreams)
ws2.cell(row=total_row, column=1, value='TOTAL').font = total_font
for col in [3, 4, 5, 7, 8, 9, 11, 12]:
    cl = get_column_letter(col)
    ws2.cell(row=total_row, column=col).value = f'=SUM({cl}4:{cl}{total_row - 1})'
    ws2.cell(row=total_row, column=col).font = total_font
    ws2.cell(row=total_row, column=col).border = thin_border
ws2.cell(row=total_row, column=6).value = '=$B$1'
ws2.cell(row=total_row, column=6).font = total_font

for col in [7, 8, 9, 12]:
    for row in range(4, total_row + 1):
        ws2.cell(row=row, column=col).number_format = currency_fmt

apply_data_styles(ws2, 4, total_row - 1, 10)
for row in range(4, total_row + 1):
    for col in [11, 12, 13]:
        ws2.cell(row=row, column=col).fill = actuals_fill
        ws2.cell(row=row, column=col).border = thin_border
        ws2.cell(row=row, column=col).font = body_font

widths = {1: 30, 2: 55, 3: 12, 4: 12, 5: 14, 6: 10, 7: 12, 8: 12, 9: 14, 10: 55, 11: 14, 12: 14, 13: 12}
for col, w in widths.items():
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.freeze_panes = 'B4'

# ============================================================
# SHEET 3: RISK REGISTER (same as V1 -- risks unchanged)
# ============================================================
ws3 = wb.create_sheet('Risk Register')

headers3 = ['#', 'Risk', 'Severity', 'Likelihood', 'Impact', 'Mitigation', 'Owner', 'Status']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header_row(ws3, 1, 8)

from brand_styles import SEV_HIGH_FILL, SEV_MED_FILL, SEV_LOW_FILL
sev_high_fill = SEV_HIGH_FILL
sev_med_fill = SEV_MED_FILL
sev_low_fill = SEV_LOW_FILL

risks = [
    (1, '10 years of uncoded contact data -- AI enrichment scope may exceed estimates',
     'HIGH', 'HIGH',
     'Data cleanup takes 2-3x estimate; marketing segmentation delayed; project timeline extends',
     'Run AI enrichment pilot on 500 records first. Measure classification accuracy before full run. Budget 30% buffer.',
     'Consultant + Andrea', 'Open'),
    (2, 'Pipeline under-reporting -- reps using LinkedIn and manual quoting outside CRM',
     'HIGH', 'HIGH',
     'Forecasting unreliable during transition; board sees pipeline dip before improvement',
     'Jodi mandates CRM usage pre-launch. Set data entry expectations. Brief board that reported pipeline will increase as capture improves.',
     'Jodi (executive sponsor)', 'Open'),
    (3, 'Scope creep from close business relationship',
     'HIGH', 'MEDIUM',
     'Informal requests bypass SOW; hours consumed without authorization; budget overrun',
     'SOW signed by budget authority. All change requests documented formally. Weekly hours tracking visible to client. Lesson learned from prior engagement.',
     'Consultant + Jodi', 'Open'),
    (4, 'Board ROI expectations overpromise',
     'HIGH', 'MEDIUM',
     'Board approves but expects unrealistic timeline or savings; sets up failure',
     'ROI narrative uses conservative estimates with ranges. 90-day KPI check-in.',
     'Consultant + Jodi', 'Open'),
    (5, 'Seat/license confusion -- mixed versions, unused seats',
     'MEDIUM', 'HIGH',
     'Paying for unused seats; features expected but unavailable',
     'License audit in week 1. Cost savings memo before configuration work.',
     'Consultant + Jodi', 'Open'),
    (6, 'Adoption risk -- reps prefer LinkedIn',
     'MEDIUM', 'HIGH',
     'New CRM configuration goes unused; ROI unrealized',
     'Executive mandate. Make CRM easier than alternatives. Rep-level dashboards.',
     'Jodi + Andrea', 'Open'),
    (7, 'Outlook email logging defect',
     'LOW', 'MEDIUM',
     'Fix may require HubSpot support escalation with unknown timeline',
     'Diagnose week 1. Open HubSpot ticket if platform bug.',
     'Consultant', 'Open'),
    (8, 'COO identity ambiguity (Lauren vs. Zach Beegal)',
     'LOW', 'LOW',
     'Miscommunication on stakeholder authority',
     'Clarify org chart at kickoff.',
     'Consultant', 'Open'),
]

for i, (num, risk, sev, lik, impact, mit, owner, status) in enumerate(risks, 2):
    ws3.cell(row=i, column=1, value=num)
    ws3.cell(row=i, column=2, value=risk)
    ws3.cell(row=i, column=3, value=sev)
    ws3.cell(row=i, column=4, value=lik)
    ws3.cell(row=i, column=5, value=impact)
    ws3.cell(row=i, column=6, value=mit)
    ws3.cell(row=i, column=7, value=owner)
    ws3.cell(row=i, column=8, value=status)
    sev_cell = ws3.cell(row=i, column=3)
    if sev == 'HIGH': sev_cell.fill = sev_high_fill
    elif sev == 'MEDIUM': sev_cell.fill = sev_med_fill
    else: sev_cell.fill = sev_low_fill

apply_data_styles(ws3, 2, len(risks) + 1, 8)
for i, (_, _, sev, *_) in enumerate(risks, 2):
    c = ws3.cell(row=i, column=3)
    if sev == 'HIGH': c.fill = sev_high_fill
    elif sev == 'MEDIUM': c.fill = sev_med_fill
    else: c.fill = sev_low_fill

widths3 = {1: 5, 2: 45, 3: 12, 4: 12, 5: 40, 6: 50, 7: 25, 8: 10}
for col, w in widths3.items():
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.freeze_panes = 'A2'

# ============================================================
# SHEET 4: ASSUMPTIONS & EXCLUSIONS
# ============================================================
ws4 = wb.create_sheet('Assumptions & Exclusions')

ws4.cell(row=1, column=1, value='Scope Assumptions').font = section_font
assumptions = [
    'HubSpot Sales Hub Enterprise (existing instance -- reconfiguration, not net-new)',
    'Up to 23 licensed users across 3 core, 10 sales, and 10 service seats',
    'Outlook as primary email platform (with known logging defect to be resolved)',
    'Up to 30 custom contact/company properties',
    '1 active sales pipeline with pre-pipeline separated',
    '2 deal categories: partner-level and client-level (same stages, different segmentation)',
    'Stage probability mapping using Jodi\'s framework (5/20/40/50/60/80/90/100)',
    'Contact type classification: client, prospect, broker, partner',
    'AI enrichment using HubSpot Breeze credits (5-10K credits available)',
    'Up to 10 automated workflows',
    '4 reporting dashboards with up to 15 custom reports',
    '8-10 week implementation timeline',
    'All work conducted remotely unless otherwise agreed',
    'Estimates calibrated against prior comparable engagement (Wagmo/Salesforce) and industry benchmarks',
    'Any scope changes managed through formal change request and may impact cost or timing',
]
for i, a in enumerate(assumptions, 2):
    ws4.cell(row=i, column=1, value=i - 1)
    ws4.cell(row=i, column=2, value=a)

gap = len(assumptions) + 3
ws4.cell(row=gap, column=1, value='Out of Scope').font = section_font
exclusions = [
    'CPQ / quoting / product library build (see Option B: Full CRM Overhaul)',
    'Email sequencing or sales cadence automation',
    'Marketing automation (newsletters, campaigns, landing pages)',
    'NetSuite, Bill.com, or any ERP/financial system integration',
    'Client Success pipeline or client health scoring',
    'Commission tracking or calculation automation',
    'Contract management or e-signature workflows',
    'Custom API development or middleware (n8n/Zapier)',
    'Data migration from external systems (HubSpot internal cleanup only)',
    'Ongoing managed services beyond 2-week hypercare',
    'HubSpot license procurement or tier changes',
]
for i, e in enumerate(exclusions, gap + 1):
    ws4.cell(row=i, column=1, value=i - gap)
    ws4.cell(row=i, column=2, value=e)

gap2 = gap + len(exclusions) + 2
ws4.cell(row=gap2, column=1, value='Client Responsibilities').font = section_font
responsibilities = [
    'Designate primary point of contact (Andrea recommended)',
    'Ensure Jodi and Andrea availability for discovery, UAT, and training sessions',
    'Provide contact classification rules and validate AI enrichment results',
    'Confirm deal stage definitions and probability framework before pipeline build',
    'Communicate standardized close won/lost reason categories',
    'Mandate CRM usage across sales team (executive sponsorship from Jodi)',
    'Complete UAT testing within agreed timeline',
    'Provide access to HubSpot admin portal',
    'Share 3-5 recent quotes/proposals for reference',
]
for i, r in enumerate(responsibilities, gap2 + 1):
    ws4.cell(row=i, column=1, value=i - gap2)
    ws4.cell(row=i, column=2, value=r)

gap3 = gap2 + len(responsibilities) + 2
ws4.cell(row=gap3, column=1, value='Open Items').font = section_font
open_items = [
    'Exact contact/company/deal record counts (pending portal access)',
    'Service Hub usage details -- preserve or deprecate?',
    'COO identity and HubSpot admin access ownership',
    'Outlook email logging defect root cause',
    'Existing workflow inventory',
    'Historical deal data reliability for probability calibration',
    'Microsite client intake form link from Andrea',
    'SOW signing authority (CEO or board)',
]
for i, o in enumerate(open_items, gap3 + 1):
    ws4.cell(row=i, column=1, value=i - gap3)
    ws4.cell(row=i, column=2, value=o)

for row in ws4.iter_rows(min_row=2, max_col=3):
    for cell in row:
        cell.font = body_font
        cell.alignment = wrap_align

ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 80

# SAVE
output_path = '/Users/harbuckconsulting/projects/Project Scoping Tool/Strive_Global/Strive_CRM_Foundation_Estimate_V2.xlsx'
wb.save(output_path)
print(f'Option A V2 saved to: {output_path}')
