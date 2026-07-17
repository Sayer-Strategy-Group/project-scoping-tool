#!/usr/bin/env python3
"""Shared scoping-workbook generator.

Reads a structured ``scope.json`` (conforming to ``templates/scoping-schema.json``)
and emits the branded 5-sheet Sayer scoping workbook (+ optional Approach
Comparison sheet, + optional Schedule sheet when ``engagement.startDate`` is set —
see ``compute_schedule``). This REPLACES the per-client, hand-written
``build_estimate.py`` scripts that previously lived in each client folder: one
generator, one source of truth, driven entirely by data.

Usage:
    python3 scripts/build_estimate.py <path/to/scope.json> [--out <file.xlsx>]
    python3 scripts/build_estimate.py <scope.json> --validate-only

If ``--out`` is omitted the workbook is written next to the scope.json as
``{ClientName}_Scoping_Estimate.xlsx``.

Brand styling comes exclusively from ``scripts/brand_styles.py`` — no inline hex.
Python 3.9 compatible (no ``X | None`` / ``tuple[X, Y]`` syntax).
"""
import argparse
import json
import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Dict, List, Optional

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "scripts"))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment  # noqa: E402
from openpyxl.worksheet.worksheet import Worksheet  # noqa: E402

from brand_styles import (  # noqa: E402
    BODY_FONT,
    BOLD_BODY_FONT,
    CAPTION_FONT,
    CURRENCY_FMT,
    DATE_FMT,
    INPUT_FILL,
    SECTION_FONT,
    THIN_BORDER,
    TITLE_FONT,
    apply_data_styles,
    apply_data_styles_rows,
    apply_input_fill_cells,
    get_column_letter,
    severity_fill,
    style_data_cell,
    style_header_row,
    style_secondary_header_cells,
)

SCHEMA_PATH = REPO_ROOT / "templates" / "scoping-schema.json"


# --------------------------------------------------------------------------- #
# Loading + validation
# --------------------------------------------------------------------------- #
def load_scope(path: Path) -> dict:
    with path.open() as fh:
        return json.load(fh)


def validate_scope(scope: dict) -> List[str]:
    """Validate against the JSON Schema if ``jsonschema`` is available.

    Returns a list of human-readable error strings (empty == valid). Schema
    validation is best-effort: if ``jsonschema`` is not installed we fall back
    to a minimal structural check so the generator still runs.
    """
    try:
        from jsonschema import Draft202012Validator
    except ImportError:
        required = ["client", "engagement", "workstreams", "budgetSummary", "risks"]
        return ["missing required key: %s" % k for k in required if k not in scope]

    with SCHEMA_PATH.open() as fh:
        schema = json.load(fh)
    validator = Draft202012Validator(schema)
    return [
        "%s: %s" % (list(err.path) or "<root>", err.message)
        for err in sorted(validator.iter_errors(scope), key=lambda e: list(e.path))
    ]


# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #
def _title(ws: Worksheet, text: str, last_col: int) -> None:
    ws.cell(row=1, column=1, value=text).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)


def _set_widths(ws: Worksheet, widths: Dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _finalize(ws: Worksheet, last_col: int, last_row: int) -> None:
    ws.freeze_panes = "B%d" % 1  # placeholder; per-sheet panes set by caller
    ws.print_area = "A1:%s%d" % (get_column_letter(last_col), last_row)


# --------------------------------------------------------------------------- #
# Sheet 1 — Scoping Estimate
# --------------------------------------------------------------------------- #
def build_scoping_estimate(ws: Worksheet, scope: dict) -> Dict[str, int]:
    """Build Sheet 1. Returns a map of workstream id -> its data row (for Sheet-2 cross-refs)."""
    rate = scope["engagement"]["defaultRate"]
    workstreams = scope["workstreams"]

    _title(ws, "%s — Scoping Estimate" % scope["client"]["name"], last_col=13)

    # Single configurable rate cell ($B$2) that all cost columns reference.
    ws.cell(row=2, column=1, value="Rate (USD/hr)").font = BOLD_BODY_FONT
    ws.cell(row=2, column=2, value=rate)
    apply_input_fill_cells(ws, row=2, cols=[2])
    ws.cell(row=2, column=2).number_format = CURRENCY_FMT

    header_row = 4
    headers = [
        "Workstream", "Description", "Min Hours", "Max Hours", "Median Hours",
        "Rate", "Min Cost", "Max Cost", "Median Cost", "Notes/Assumptions",
        "Actual Hours", "Actual Cost", "Variance",
    ]
    for col, label in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=label)
    style_header_row(ws, row=header_row, max_col=10)            # forecast columns A-J
    style_secondary_header_cells(ws, row=header_row, cols=[11, 12, 13])  # actuals K-M

    ws_row_map: Dict[str, int] = {}
    first_data = header_row + 1
    r = first_data
    for w in workstreams:
        ws_row_map[w["id"]] = r
        ws.cell(row=r, column=1, value="%s  %s" % (w["id"], w["name"]))
        ws.cell(row=r, column=2, value=w.get("description", ""))
        ws.cell(row=r, column=3, value=w["hours"]["min"])
        ws.cell(row=r, column=4, value=w["hours"]["max"])
        ws.cell(row=r, column=5, value=w["hours"]["median"])
        ws.cell(row=r, column=6, value="=$B$2")
        ws.cell(row=r, column=7, value="=C%d*F%d" % (r, r))
        ws.cell(row=r, column=8, value="=D%d*F%d" % (r, r))
        ws.cell(row=r, column=9, value="=E%d*F%d" % (r, r))
        ws.cell(row=r, column=10, value=w.get("notes", ""))
        # Actuals: K is the only input; L derives from it; M is the variance.
        ws.cell(row=r, column=12, value="=K%d*$B$2" % r)
        ws.cell(row=r, column=13, value="=L%d-I%d" % (r, r))
        r += 1
    last_data = r - 1

    # Totals row
    total_row = r
    ws.cell(row=total_row, column=1, value="TOTAL").font = BOLD_BODY_FONT
    for col_letter, col_idx in (("C", 3), ("D", 4), ("E", 5), ("G", 7), ("H", 8),
                                ("I", 9), ("K", 11), ("L", 12)):
        ws.cell(row=total_row, column=col_idx,
                value="=SUM(%s%d:%s%d)" % (col_letter, first_data, col_letter, last_data))
    ws.cell(row=total_row, column=13, value="=L%d-I%d" % (total_row, total_row))

    # Styling: data rows (striped), then currency + input affordances.
    apply_data_styles(ws, data_start=first_data, data_end=last_data, max_col=13)
    for rr in range(first_data, last_data + 1):
        for col_idx in (6, 7, 8, 9, 12, 13):
            ws.cell(row=rr, column=col_idx).number_format = CURRENCY_FMT
        apply_input_fill_cells(ws, row=rr, cols=[11])  # Actual Hours input
    # Totals row styling
    for col_idx in range(1, 14):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = BOLD_BODY_FONT
        cell.border = THIN_BORDER
        if col_idx in (7, 8, 9, 12, 13):
            cell.number_format = CURRENCY_FMT

    _set_widths(ws, {"A": 30, "B": 42, "C": 11, "D": 11, "E": 13, "F": 10,
                     "G": 12, "H": 12, "I": 13, "J": 40, "K": 12, "L": 12, "M": 12})
    ws.freeze_panes = "B%d" % first_data
    ws.print_area = "A1:M%d" % total_row
    return ws_row_map


# --------------------------------------------------------------------------- #
# Sheet 2 — Task Breakdown (with cross-sheet integrity check)
# --------------------------------------------------------------------------- #
def build_task_breakdown(ws: Worksheet, scope: dict, ws_row_map: Dict[str, int]) -> None:
    _title(ws, "%s — Task Breakdown" % scope["client"]["name"], last_col=6)
    header_row = 2
    headers = ["#", "Phase", "Workstream", "Task", "Hours", "Check"]
    for col, label in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=label)
    style_header_row(ws, row=header_row, max_col=6)

    phase_names = {p["id"]: p["name"] for p in scope.get("phases", [])}
    data_rows: List[int] = []
    subtotal_rows: List[int] = []
    r = header_row + 1
    seq = 1
    for w in scope["workstreams"]:
        tasks = w.get("tasks") or []
        task_cells: List[str] = []
        phase_label = phase_names.get(w.get("phase"), "Cross-phase" if not w.get("phase") else w.get("phase"))
        for t in tasks:
            ws.cell(row=r, column=1, value=seq)
            ws.cell(row=r, column=2, value=phase_label)
            ws.cell(row=r, column=3, value="%s  %s" % (w["id"], w["name"]))
            ws.cell(row=r, column=4, value=t["name"])
            ws.cell(row=r, column=5, value=t["hours"])
            task_cells.append("E%d" % r)
            data_rows.append(r)
            seq += 1
            r += 1
        # Workstream subtotal row with integrity check against Sheet 1 median.
        sub_row = r
        ws.cell(row=sub_row, column=3, value="%s — subtotal" % w["id"]).font = BOLD_BODY_FONT
        if task_cells:
            ws.cell(row=sub_row, column=5, value="=SUM(%s)" % ",".join(task_cells))
            ws.cell(row=sub_row, column=6,
                    value="=IF(E%d='Scoping Estimate'!E%d,\"OK\",\"MISMATCH\")"
                          % (sub_row, ws_row_map[w["id"]]))
        else:
            # No task detail — surface that the breakdown is incomplete.
            ws.cell(row=sub_row, column=5, value=0)
            ws.cell(row=sub_row, column=6, value="NO TASKS")
        subtotal_rows.append(sub_row)
        r += 1

    last_row = r - 1
    apply_data_styles_rows(ws, rows=data_rows, max_col=6)
    for sr in subtotal_rows:
        for col_idx in range(1, 7):
            cell = ws.cell(row=sr, column=col_idx)
            cell.font = BOLD_BODY_FONT
            cell.border = THIN_BORDER
    _set_widths(ws, {"A": 5, "B": 34, "C": 30, "D": 44, "E": 9, "F": 12})
    ws.freeze_panes = "A%d" % (header_row + 1)
    ws.print_area = "A1:F%d" % last_row


# --------------------------------------------------------------------------- #
# Sheet 3 — Deliverables & Acceptance
# --------------------------------------------------------------------------- #
def build_deliverables(ws: Worksheet, scope: dict) -> None:
    _title(ws, "%s — Deliverables & Acceptance" % scope["client"]["name"], last_col=5)
    header_row = 2
    headers = ["#", "Deliverable", "Description", "Workstream", "Acceptance Criteria"]
    for col, label in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=label)
    style_header_row(ws, row=header_row, max_col=5)

    items = scope.get("deliverables", [])
    r = header_row + 1
    for d in items:
        ws.cell(row=r, column=1, value=d.get("id", r - header_row))
        ws.cell(row=r, column=2, value=d.get("name", ""))
        ws.cell(row=r, column=3, value=d.get("description", ""))
        ws.cell(row=r, column=4, value=d.get("workstream", ""))
        ws.cell(row=r, column=5, value=d.get("acceptanceCriteria", ""))
        r += 1
    last_row = max(r - 1, header_row)
    if items:
        apply_data_styles(ws, data_start=header_row + 1, data_end=last_row, max_col=5)
    _set_widths(ws, {"A": 5, "B": 30, "C": 44, "D": 16, "E": 46})
    ws.freeze_panes = "A%d" % (header_row + 1)
    ws.print_area = "A1:E%d" % last_row


# --------------------------------------------------------------------------- #
# Sheet 4 — Risk Register
# --------------------------------------------------------------------------- #
def build_risk_register(ws: Worksheet, scope: dict) -> None:
    _title(ws, "%s — Risk Register" % scope["client"]["name"], last_col=8)
    header_row = 2
    headers = ["#", "Risk", "Severity", "Likelihood", "Impact", "Mitigation", "Owner", "Status"]
    for col, label in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=label)
    style_header_row(ws, row=header_row, max_col=8)

    risks = scope.get("risks", [])
    r = header_row + 1
    for i, risk in enumerate(risks, start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=risk.get("name", ""))
        sev = (risk.get("severity") or "").upper()
        ws.cell(row=r, column=3, value=sev)
        ws.cell(row=r, column=4, value=(risk.get("likelihood") or "").upper())
        ws.cell(row=r, column=5, value=risk.get("impact", ""))
        ws.cell(row=r, column=6, value=risk.get("mitigation", ""))
        ws.cell(row=r, column=7, value=risk.get("owner", ""))
        ws.cell(row=r, column=8, value=risk.get("status", "open"))
        r += 1
    last_row = max(r - 1, header_row)
    if risks:
        apply_data_styles(ws, data_start=header_row + 1, data_end=last_row, max_col=8)
        # Severity color overlay (after striping so the semantic fill wins).
        for i, risk in enumerate(risks):
            cell = ws.cell(row=header_row + 1 + i, column=3)
            cell.fill = severity_fill(risk.get("severity", "low"))
    ws.auto_filter.ref = "A%d:H%d" % (header_row, last_row)
    _set_widths(ws, {"A": 5, "B": 40, "C": 11, "D": 12, "E": 38, "F": 40, "G": 10, "H": 9})
    ws.freeze_panes = "A%d" % (header_row + 1)
    ws.print_area = "A1:H%d" % last_row


# --------------------------------------------------------------------------- #
# Sheet 5 — Assumptions
# --------------------------------------------------------------------------- #
def build_assumptions(ws: Worksheet, scope: dict) -> None:
    _title(ws, "%s — Assumptions & Exclusions" % scope["client"]["name"], last_col=2)
    sections = [
        ("Scope Assumptions", scope.get("assumptions", [])),
        ("Out of Scope", scope.get("exclusions", [])),
        ("Client Responsibilities", scope.get("clientResponsibilities", [])),
        ("Open Items", scope.get("openItems", [])),
    ]
    r = 3
    data_rows: List[int] = []
    for title, items in sections:
        ws.cell(row=r, column=1, value=title).font = SECTION_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        style_header_row(ws, row=r, max_col=2)
        r += 1
        for item in (items or ["(none)"]):
            ws.cell(row=r, column=1, value="•")
            ws.cell(row=r, column=2, value=item)
            data_rows.append(r)
            r += 1
        r += 1  # blank spacer between sections
    last_row = max(r - 1, 3)
    for dr in data_rows:
        style_data_cell(ws, dr, 1)
        cell = ws.cell(row=dr, column=2)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    _set_widths(ws, {"A": 4, "B": 90})
    ws.print_area = "A1:B%d" % last_row


# --------------------------------------------------------------------------- #
# Sheet 6 — Approach Comparison (optional)
# --------------------------------------------------------------------------- #
def build_approach_comparison(ws: Worksheet, scope: dict) -> None:
    approaches = scope.get("approaches", [])
    last_col = 1 + len(approaches)
    _title(ws, "%s — Approach Comparison" % scope["client"]["name"], last_col=max(last_col, 2))

    header_row = 2
    ws.cell(row=header_row, column=1, value="Dimension")
    for j, a in enumerate(approaches, start=2):
        label = a["name"] + (" (recommended)" if a.get("recommended") else "")
        ws.cell(row=header_row, column=j, value=label)
    style_header_row(ws, row=header_row, max_col=last_col)

    def row_vals(label, fn):
        return [label] + [fn(a) for a in approaches]

    rows = [
        row_vals("Total Hours (median)", lambda a: a.get("totalHours", {}).get("median", "")),
        row_vals("Total Cost (median)", lambda a: a.get("totalCost", {}).get("median", "")),
        row_vals("Timeline (weeks)", lambda a: a.get("timelineWeeks", "")),
        row_vals("Risk Level", lambda a: (a.get("riskLevel") or "").upper()),
        row_vals("Dependencies", lambda a: "; ".join(a.get("dependencies", []))),
        row_vals("Recommended", lambda a: "YES" if a.get("recommended") else "—"),
    ]
    r = header_row + 1
    for vals in rows:
        for j, v in enumerate(vals, start=1):
            ws.cell(row=r, column=j, value=v)
        r += 1
    data_end = r - 1
    apply_data_styles(ws, data_start=header_row + 1, data_end=data_end, max_col=last_col)
    # Cost row currency
    for j in range(2, last_col + 1):
        ws.cell(row=header_row + 2, column=j).number_format = CURRENCY_FMT

    # Recommendation callout
    rec = scope.get("recommendation")
    if rec:
        rr = data_end + 2
        ws.cell(row=rr, column=1, value="Recommendation").font = SECTION_FONT
        ws.cell(row=rr + 1, column=1, value=rec).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=rr + 1, start_column=1, end_row=rr + 1, end_column=last_col)
        data_end = rr + 1

    widths = {"A": 24}
    for j in range(2, last_col + 1):
        widths[get_column_letter(j)] = 34
    _set_widths(ws, widths)
    ws.print_area = "A1:%s%d" % (get_column_letter(last_col), data_end)


# --------------------------------------------------------------------------- #
# Committed / T&M mode — Sheet 1 (phase rollup) + Sheet 2 (reconcile to phase)
# --------------------------------------------------------------------------- #
def _phase_label(phase: dict) -> str:
    if phase.get("kind") == "contingency":
        return "CONT"
    return "P%d" % phase.get("order", 0)


def build_scoping_estimate_committed(ws: Worksheet, scope: dict) -> Dict[str, int]:
    """Committed (T&M) Sheet 1: metadata block + phase rollup + tech fee + grand total.

    Returns phase id -> data row (Sheet-2 reconciles task subtotals to the 'Est. Hours'
    cell at that row). Mirrors the validated committed/T&M phase-rollup layout.
    """
    eng = scope["engagement"]
    blended = eng.get("blendedRate") or eng["defaultRate"]
    tech_pct = eng.get("techFeePct", 0)
    phases = sorted(scope.get("phases", []), key=lambda p: p.get("order", 0))

    _title(ws, "%s — %s" % (scope["client"]["name"], eng.get("name", "Scoping Estimate")), last_col=8)
    sub = ws.cell(row=2, column=1, value=eng.get("feeStructure", ""))
    sub.font = CAPTION_FONT
    sub.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # Engagement metadata block.
    r = 4

    def meta(label: str, value: str) -> None:
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = BOLD_BODY_FONT
        cell = ws.cell(row=r, column=3, value=value)
        cell.font = BODY_FONT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        r += 1

    meta("Engagement", eng.get("name", ""))
    if eng.get("lead"):
        meta("Sayer engagement lead", eng["lead"])
    for s in eng.get("staffMix", []):
        meta(s.get("role", "Staff"),
             "%s — $%s/hr at %s%% allocation" % (s.get("name", ""), s.get("rate"), s.get("allocationPct")))
    meta("Blended rate", "$%s/hr" % blended)
    if eng.get("feeStructure"):
        meta("Fee structure", eng["feeStructure"])

    header_row = r + 1
    headers = ["ID", "Phase / Workstream", "Est. Hours", "Est. Amount",
               "Actual Hours", "Actual Amount", "Variance Hrs", "Notes"]
    for col, label in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=label)
    style_header_row(ws, row=header_row, max_col=4)              # forecast A-D
    style_secondary_header_cells(ws, row=header_row, cols=[5, 6, 7])  # actuals E-G
    style_header_row(ws, row=header_row, max_col=8, start_col=8)  # Notes H

    phase_row_map: Dict[str, int] = {}
    first = header_row + 1
    rr = first
    for p in phases:
        hrs = p.get("committedHours", 0)
        amt = p.get("committedAmount", hrs * blended)
        ws.cell(row=rr, column=1, value=_phase_label(p))
        ws.cell(row=rr, column=2, value=p.get("description") or p.get("name", ""))
        ws.cell(row=rr, column=3, value=hrs)
        ws.cell(row=rr, column=4, value=amt)
        ws.cell(row=rr, column=6, value="=E%d*%s" % (rr, blended))   # actual amount
        ws.cell(row=rr, column=7, value="=E%d-C%d" % (rr, rr))       # variance hrs
        phase_row_map[p["id"]] = rr
        rr += 1
    last_phase = rr - 1

    sub_row, fee_row, gt_row = rr, rr + 1, rr + 2
    ws.cell(row=sub_row, column=2, value="Consulting fees subtotal")
    ws.cell(row=sub_row, column=3, value="=SUM(C%d:C%d)" % (first, last_phase))
    ws.cell(row=sub_row, column=4, value="=SUM(D%d:D%d)" % (first, last_phase))
    ws.cell(row=fee_row, column=2, value="Technology & Administrative Fee (%g%%)" % tech_pct)
    ws.cell(row=fee_row, column=4, value="=ROUND(D%d*%s,0)" % (sub_row, tech_pct / 100.0))
    ws.cell(row=gt_row, column=2, value="Grand total estimate")
    ws.cell(row=gt_row, column=4, value="=D%d+D%d" % (sub_row, fee_row))

    apply_data_styles(ws, data_start=first, data_end=last_phase, max_col=8)
    for r2 in range(first, last_phase + 1):
        ws.cell(row=r2, column=4).number_format = CURRENCY_FMT
        ws.cell(row=r2, column=6).number_format = CURRENCY_FMT
        apply_input_fill_cells(ws, row=r2, cols=[5])  # Actual Hours input
    for trow in (sub_row, fee_row, gt_row):
        for col in range(1, 9):
            cell = ws.cell(row=trow, column=col)
            cell.font = BOLD_BODY_FONT
            cell.border = THIN_BORDER
        ws.cell(row=trow, column=4).number_format = CURRENCY_FMT

    # Optional payment / invoice schedule (committed SOW). Sum should reconcile to grand total.
    last_row = gt_row
    schedule = eng.get("paymentSchedule") or []
    if schedule:
        sched_header = gt_row + 2
        hcell = ws.cell(row=sched_header, column=1, value="Payment Schedule")
        hcell.font = BOLD_BODY_FONT
        ws.merge_cells(start_row=sched_header, start_column=1, end_row=sched_header, end_column=3)
        ws.cell(row=sched_header, column=4, value="Amount").font = BOLD_BODY_FONT
        sched_first = sched_header + 1
        rr2 = sched_first
        for inv in schedule:
            ws.cell(row=rr2, column=2, value=inv.get("label", ""))
            amt_cell = ws.cell(row=rr2, column=4, value=inv.get("amount", 0))
            amt_cell.number_format = CURRENCY_FMT
            if inv.get("notes"):
                ws.cell(row=rr2, column=8, value=inv["notes"])
            rr2 += 1
        sched_last = rr2 - 1
        total_row = rr2
        ws.cell(row=total_row, column=2, value="Total").font = BOLD_BODY_FONT
        tcell = ws.cell(row=total_row, column=4, value="=SUM(D%d:D%d)" % (sched_first, sched_last))
        tcell.number_format = CURRENCY_FMT
        tcell.font = BOLD_BODY_FONT
        for rrow in range(sched_header, total_row + 1):
            for col in range(1, 9):
                ws.cell(row=rrow, column=col).border = THIN_BORDER
        last_row = total_row

    _set_widths(ws, {"A": 6, "B": 40, "C": 12, "D": 14, "E": 13, "F": 14, "G": 12, "H": 28})
    ws.freeze_panes = "A%d" % first
    ws.print_area = "A1:H%d" % last_row
    return phase_row_map


def build_task_breakdown_committed(ws: Worksheet, scope: dict, phase_row_map: Dict[str, int]) -> None:
    _title(ws, "%s — Task Breakdown" % scope["client"]["name"], last_col=5)
    header_row = 2
    for col, label in enumerate(["#", "Phase", "Task", "Hours", "Check"], start=1):
        ws.cell(row=header_row, column=col, value=label)
    style_header_row(ws, row=header_row, max_col=5)

    phases = sorted(scope.get("phases", []), key=lambda p: p.get("order", 0))
    data_rows: List[int] = []
    subtotal_rows: List[int] = []
    r = header_row + 1
    seq = 1
    for p in phases:
        tasks = p.get("tasks") or []
        disp = _phase_label(p)
        cells: List[str] = []
        for t in tasks:
            ws.cell(row=r, column=1, value=seq)
            ws.cell(row=r, column=2, value="%s  %s" % (disp, p.get("name", "")))
            ws.cell(row=r, column=3, value=t["name"])
            ws.cell(row=r, column=4, value=t["hours"])
            cells.append("D%d" % r)
            data_rows.append(r)
            seq += 1
            r += 1
        sub_row = r
        ws.cell(row=sub_row, column=3, value="%s — subtotal" % disp).font = BOLD_BODY_FONT
        if cells:
            ws.cell(row=sub_row, column=4, value="=SUM(%s)" % ",".join(cells))
            ws.cell(row=sub_row, column=5,
                    value="=IF(D%d='Scoping Estimate'!C%d,\"OK\",\"MISMATCH\")"
                          % (sub_row, phase_row_map[p["id"]]))
        else:
            ws.cell(row=sub_row, column=4, value=0)
            ws.cell(row=sub_row, column=5, value="NO TASKS")
        subtotal_rows.append(sub_row)
        r += 1

    last_row = r - 1
    apply_data_styles_rows(ws, rows=data_rows, max_col=5)
    for sr in subtotal_rows:
        for col in range(1, 6):
            cell = ws.cell(row=sr, column=col)
            cell.font = BOLD_BODY_FONT
            cell.border = THIN_BORDER
    _set_widths(ws, {"A": 5, "B": 30, "C": 50, "D": 9, "E": 12})
    ws.freeze_panes = "A%d" % (header_row + 1)
    ws.print_area = "A1:E%d" % last_row


# --------------------------------------------------------------------------- #
# Schedule — derive start/end dates from engagement.startDate (ranges mode only)
# --------------------------------------------------------------------------- #
def compute_schedule(scope: dict) -> Optional[dict]:
    """Derive start/end dates per workstream + task from ``engagement.startDate``.

    Simplified single-track model: SEQUENTIAL workstreams (the default) consume
    capacity in ``workstreams[]`` array order — no explicit dependency graph or
    resource contention. CONCURRENT workstreams (``scheduling: "concurrent"``)
    span the full computed sequential timeline instead of blocking it — use for
    work that genuinely runs alongside the build rather than gating it (PM /
    governance, an externally-gated feasibility spike).

    Returns ``None`` (no Schedule sheet) if ``engagement.startDate`` is absent —
    fully backward compatible with scope.json records written before scheduling
    existed — or if the scope is in committed/phase mode (v1 covers ranges mode
    only; committed-mode scheduling is a documented limitation, not silently
    wrong).
    """
    eng = scope.get("engagement", {})
    start_str = eng.get("startDate")
    if not start_str or "workstreams" not in scope:
        return None
    start_date = date.fromisoformat(start_str)

    capacity = eng.get("capacityHoursPerWeek")
    if not capacity:
        weeks = eng.get("estimatedTimelineWeeks")
        median_hours = (eng.get("totalHours") or {}).get("median")
        capacity = (median_hours / weeks) if (weeks and median_hours) else 20.0
    capacity = capacity or 20.0  # guard an explicit 0/None from a bad scope.json

    def _add_weeks(base: date, weeks: float) -> date:
        return base + timedelta(days=round(weeks * 7))

    workstreams = scope["workstreams"]
    sequential = [w for w in workstreams if w.get("scheduling", "sequential") == "sequential"]
    concurrent = [w for w in workstreams if w.get("scheduling") == "concurrent"]

    ws_schedule: Dict[str, dict] = {}
    cumulative_hours = 0.0
    for w in sequential:
        tasks = w.get("tasks") or [{"name": w["name"], "hours": w["hours"]["median"]}]
        synthetic = not w.get("tasks")
        ws_start_hours = cumulative_hours
        task_rows = []
        for t in tasks:
            t_start, t_end = cumulative_hours, cumulative_hours + t["hours"]
            task_rows.append({
                "name": t["name"],
                "hours": t["hours"],
                "start": _add_weeks(start_date, t_start / capacity),
                "end": _add_weeks(start_date, t_end / capacity),
            })
            cumulative_hours = t_end
        ws_schedule[w["id"]] = {
            "start": _add_weeks(start_date, ws_start_hours / capacity),
            "end": _add_weeks(start_date, cumulative_hours / capacity),
            "tasks": task_rows,
            "synthetic_tasks": synthetic,
        }

    overall_end = _add_weeks(start_date, cumulative_hours / capacity) if sequential else start_date
    span_days = (overall_end - start_date).days

    for w in concurrent:
        tasks = w.get("tasks") or [{"name": w["name"], "hours": w["hours"]["median"]}]
        synthetic = not w.get("tasks")
        local_total = sum(t["hours"] for t in tasks) or 1
        local_cum = 0.0
        task_rows = []
        for t in tasks:
            frac_start, frac_end = local_cum / local_total, (local_cum + t["hours"]) / local_total
            task_rows.append({
                "name": t["name"],
                "hours": t["hours"],
                "start": start_date + timedelta(days=round(span_days * frac_start)),
                "end": start_date + timedelta(days=round(span_days * frac_end)),
            })
            local_cum += t["hours"]
        ws_schedule[w["id"]] = {
            "start": start_date,
            "end": overall_end,
            "tasks": task_rows,
            "synthetic_tasks": synthetic,
        }

    return {"start_date": start_date, "end_date": overall_end, "capacity": capacity, "workstreams": ws_schedule}


def build_schedule_sheet(ws: Worksheet, scope: dict, schedule: dict) -> None:
    """Render the Schedule sheet with LIVE Excel formulas, not static dates.

    Project Start Date ($C$2) and Assumed Capacity ($C$3) are input cells —
    editing either in Excel recalculates every task/workstream/project date
    automatically, no regeneration required. A hidden helper column ("Track")
    tags each row SEQ/CONC so a bounded SUMIFS can compute "cumulative
    sequential hours before this row" while correctly ignoring interleaved
    concurrent rows, matching ``compute_schedule()``'s Python model exactly.
    ``schedule`` (from ``compute_schedule``) supplies structure (task order,
    hours, synthetic-task detection) and seeds the two input cells; its
    computed date VALUES are not written to the sheet — the formulas derive
    dates independently once seeded, which is the point.
    """
    _title(ws, "%s — Schedule" % scope["client"]["name"], last_col=7)

    meta_row = 2
    ws.cell(row=meta_row, column=1, value="Project Start Date").font = BOLD_BODY_FONT
    start_cell = ws.cell(row=meta_row, column=3, value=schedule["start_date"])
    start_cell.number_format = DATE_FMT
    apply_input_fill_cells(ws, row=meta_row, cols=[3])

    cap_row = meta_row + 1
    ws.cell(row=cap_row, column=1, value="Assumed Capacity (hrs/week)").font = BOLD_BODY_FONT
    ws.cell(row=cap_row, column=3, value=round(schedule["capacity"], 1))
    apply_input_fill_cells(ws, row=cap_row, cols=[3])

    end_row = cap_row + 1
    ws.cell(row=end_row, column=1, value="Computed Project End Date").font = BOLD_BODY_FONT
    end_cell = ws.cell(row=end_row, column=3,
                        value='=$C$2+(SUMIF($H:$H,"SEQ",$D:$D)/$C$3)*7')
    end_cell.number_format = DATE_FMT

    caption_row = end_row + 1
    note = ws.cell(
        row=caption_row, column=1,
        value=("Sequential single-track model -- no dependency graph or resource contention. "
               "Concurrent workstreams (PM, ongoing feasibility) span the full computed project "
               "window rather than blocking it. Formulas recalculate automatically in Excel when "
               "you change the Project Start Date or Assumed Capacity cells above."))
    note.font = CAPTION_FONT
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=caption_row, start_column=1, end_row=caption_row, end_column=7)

    header_row = caption_row + 2
    headers = ["#", "Workstream", "Task", "Hours", "Start Date", "End Date", "Duration (wks)"]
    for col, label in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=label)
    style_header_row(ws, row=header_row, max_col=7)
    ws.cell(row=header_row, column=8, value="Track (internal)")  # hidden SUMIFS helper

    data_rows: List[int] = []
    subtotal_rows: List[int] = []
    first_data_row = header_row + 1
    r = first_data_row
    seq = 1
    for w in scope["workstreams"]:
        wsched = schedule["workstreams"].get(w["id"])
        if not wsched:
            continue  # defensive: shouldn't happen, but never crash workbook generation over it
        is_concurrent = w.get("scheduling", "sequential") == "concurrent"
        track = "CONC" if is_concurrent else "SEQ"
        ws_first_row = None
        for t_sched in wsched["tasks"]:
            if ws_first_row is None:
                ws_first_row = r
            ws.cell(row=r, column=1, value=seq)
            ws.cell(row=r, column=2, value="%s  %s" % (w["id"], w["name"]))
            ws.cell(row=r, column=3, value=t_sched["name"])
            ws.cell(row=r, column=4, value=t_sched["hours"])
            ws.cell(row=r, column=8, value=track)

            if is_concurrent:
                start_formula, end_formula = "=$C$2", "=$C$4"
            else:
                cum_before = ("0" if r == first_data_row else
                              'SUMIFS($D$%d:D%d,$H$%d:H%d,"SEQ")' % (first_data_row, r - 1, first_data_row, r - 1))
                start_formula = "=$C$2+(%s/$C$3)*7" % cum_before
                end_formula = "=$C$2+((%s+D%d)/$C$3)*7" % (cum_before, r)

            sc = ws.cell(row=r, column=5, value=start_formula); sc.number_format = DATE_FMT
            ec = ws.cell(row=r, column=6, value=end_formula); ec.number_format = DATE_FMT
            ws.cell(row=r, column=7, value="=(F%d-E%d)/7" % (r, r))
            data_rows.append(r)
            seq += 1
            r += 1
        ws_last_row = r - 1

        sub_row = r
        ws.cell(row=sub_row, column=2, value="%s — window" % w["id"]).font = BOLD_BODY_FONT
        if is_concurrent:
            sc = ws.cell(row=sub_row, column=5, value="=$C$2")
            ec = ws.cell(row=sub_row, column=6, value="=$C$4")
        else:
            sc = ws.cell(row=sub_row, column=5, value="=E%d" % ws_first_row)
            ec = ws.cell(row=sub_row, column=6, value="=F%d" % ws_last_row)
        sc.number_format = DATE_FMT
        ec.number_format = DATE_FMT
        ws.cell(row=sub_row, column=7, value="=(F%d-E%d)/7" % (sub_row, sub_row))
        subtotal_rows.append(sub_row)
        r += 1

    proj_row = r
    ws.cell(row=proj_row, column=2, value="PROJECT").font = BOLD_BODY_FONT
    sc = ws.cell(row=proj_row, column=5, value="=$C$2"); sc.number_format = DATE_FMT
    ec = ws.cell(row=proj_row, column=6, value="=$C$4"); ec.number_format = DATE_FMT
    ws.cell(row=proj_row, column=7, value="=(F%d-E%d)/7" % (proj_row, proj_row))

    last_row = proj_row
    apply_data_styles_rows(ws, rows=data_rows, max_col=7)
    for sr in subtotal_rows + [proj_row]:
        for col_idx in range(1, 8):
            cell = ws.cell(row=sr, column=col_idx)
            cell.font = BOLD_BODY_FONT
            cell.border = THIN_BORDER

    _set_widths(ws, {"A": 5, "B": 30, "C": 44, "D": 9, "E": 13, "F": 13, "G": 15})
    ws.column_dimensions[get_column_letter(8)].hidden = True
    ws.freeze_panes = "A%d" % first_data_row
    ws.print_area = "A1:G%d" % last_row


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #
def build_workbook(scope: dict) -> Workbook:
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = "Scoping Estimate"
    mode = scope["engagement"].get("estimateMode", "ranges")
    if mode == "committed":
        phase_map = build_scoping_estimate_committed(sheet1, scope)
        build_task_breakdown_committed(wb.create_sheet("Task Breakdown"), scope, phase_map)
    else:
        ws_row_map = build_scoping_estimate(sheet1, scope)
        build_task_breakdown(wb.create_sheet("Task Breakdown"), scope, ws_row_map)
        schedule = compute_schedule(scope)
        if schedule:
            build_schedule_sheet(wb.create_sheet("Schedule", index=2), scope, schedule)
    build_deliverables(wb.create_sheet("Deliverables & Acceptance"), scope)
    build_risk_register(wb.create_sheet("Risk Register"), scope)
    build_assumptions(wb.create_sheet("Assumptions"), scope)
    if scope.get("approaches"):
        build_approach_comparison(wb.create_sheet("Approach Comparison"), scope)
    return wb


def default_out_path(scope_path: Path, scope: dict) -> Path:
    safe = "".join(c for c in scope["client"]["name"] if c.isalnum() or c in (" ", "-", "_")).strip()
    safe = safe.replace(" ", "_")
    return scope_path.parent / ("%s_Scoping_Estimate.xlsx" % safe)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Generate the branded scoping workbook from scope.json")
    parser.add_argument("scope_json", help="Path to scope.json (conforms to templates/scoping-schema.json)")
    parser.add_argument("--out", help="Output .xlsx path (default: {ClientName}_Scoping_Estimate.xlsx next to scope.json)")
    parser.add_argument("--validate-only", action="store_true", help="Validate scope.json against the schema and exit")
    args = parser.parse_args(argv)

    scope_path = Path(args.scope_json).resolve()
    if not scope_path.exists():
        print("error: scope.json not found: %s" % scope_path, file=sys.stderr)
        return 2
    scope = load_scope(scope_path)

    errors = validate_scope(scope)
    if errors:
        print("scope.json FAILED schema validation:", file=sys.stderr)
        for e in errors:
            print("  - %s" % e, file=sys.stderr)
        return 1
    print("scope.json validates against schema: OK")
    if args.validate_only:
        return 0

    out_path = Path(args.out).resolve() if args.out else default_out_path(scope_path, scope)
    wb = build_workbook(scope)
    wb.save(str(out_path))
    print("wrote workbook: %s (%d sheets)" % (out_path, len(wb.sheetnames)))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
