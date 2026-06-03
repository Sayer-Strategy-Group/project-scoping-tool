"""Sayer brand styles for openpyxl workbooks.

Single source of truth for every Excel generator in this repo.
Authoritative brand spec:
  /Users/harbuckconsulting/projects/AIVA/skills/sayer-brand-guidelines/SKILL.md

Design notes:
  - Primary header = Sayer Yellow + Black (forecast/scoping columns).
  - Secondary header = Grey 700 + White (actuals / post-project columns).
    Preserves the two-tier "forecast vs reality" visual split used in every
    scoping workbook, but on-brand.
  - Input cells (rate, actuals entry) = Cool Grey fill. One input affordance
    color instead of the prior mix of light blue + light green.
  - Risk severity fills retain traffic-light semantics. Medium uses Sayer
    Yellow (only brand-native color that already means "attention").
  - Calibri, not Rethink Sans: the brand font does not survive xlsx
    round-tripping. Calibri is the sanctioned fallback.

Python 3.9 compatible. Do NOT introduce 3.10+ syntax (X | None, tuple[X, Y]).
"""
from typing import Iterable

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter  # re-exported for caller convenience

# --- Colors (hex, no leading '#') ---
SAYER_YELLOW = "FEC700"
SAYER_BLACK = "000000"
SAYER_WHITE = "FFFFFF"
GREY_700 = "2E2E2E"
GREY_600 = "7E7E7E"
GREY_500 = "BCBCBC"
GREY_300 = "E3E3E3"
COOL_GREY = "D6D6D6"

# Traffic-light severity fills (semantic, not strict brand).
SEV_HIGH = "FFC7CE"        # light red
SEV_MEDIUM = "FEC700"      # Sayer Yellow (replaces generic FFEB9C)
SEV_LOW = "C6EFCE"         # light green

# --- Typography ---
FONT_FAMILY = "Calibri"
TITLE_SIZE = 14
HEADER_SIZE = 11
BODY_SIZE = 11
CAPTION_SIZE = 10

CURRENCY_FMT = "$#,##0"


def _solid(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


# --- Font singletons ---
TITLE_FONT = Font(name=FONT_FAMILY, size=TITLE_SIZE, bold=True, color=SAYER_BLACK)
SECTION_FONT = Font(name=FONT_FAMILY, size=HEADER_SIZE, bold=True, color=SAYER_BLACK)
HEADER_FONT = Font(name=FONT_FAMILY, size=HEADER_SIZE, bold=True, color=SAYER_BLACK)
SECONDARY_HEADER_FONT = Font(name=FONT_FAMILY, size=HEADER_SIZE, bold=True, color=SAYER_WHITE)
BODY_FONT = Font(name=FONT_FAMILY, size=BODY_SIZE, color=SAYER_BLACK)
BOLD_BODY_FONT = Font(name=FONT_FAMILY, size=BODY_SIZE, bold=True, color=SAYER_BLACK)
TOTAL_FONT = Font(name=FONT_FAMILY, size=BODY_SIZE, bold=True, color=SAYER_BLACK)
CAPTION_FONT = Font(name=FONT_FAMILY, size=CAPTION_SIZE, color=GREY_600)

# --- Fill singletons ---
HEADER_FILL = _solid(SAYER_YELLOW)
SECONDARY_HEADER_FILL = _solid(GREY_700)
ALT_ROW_FILL = _solid(GREY_300)
INPUT_FILL = _solid(COOL_GREY)

SEV_HIGH_FILL = _solid(SEV_HIGH)
SEV_MED_FILL = _solid(SEV_MEDIUM)
SEV_LOW_FILL = _solid(SEV_LOW)

# --- Border singleton ---
_thin_side = Side(style="thin", color=GREY_500)
THIN_BORDER = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)

# --- Alignment singletons ---
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_WRAP_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def severity_fill(level: str) -> PatternFill:
    """Return the severity fill for HIGH / MEDIUM / LOW (case-insensitive)."""
    lvl = (level or "").strip().upper()
    if lvl == "HIGH":
        return SEV_HIGH_FILL
    if lvl == "MEDIUM":
        return SEV_MED_FILL
    return SEV_LOW_FILL


def style_header_row(ws, row: int, max_col: int, start_col: int = 1) -> None:
    """Apply primary (Sayer Yellow + Black) header styling to a row range."""
    for col in range(start_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_WRAP_ALIGN
        cell.border = THIN_BORDER


def style_secondary_header_cells(ws, row: int, cols: Iterable[int]) -> None:
    """Apply secondary (Grey 700 + White) header styling to specific columns.

    Use for 'Actuals' columns and other quieter header variants.
    """
    for col in cols:
        cell = ws.cell(row=row, column=col)
        cell.font = SECONDARY_HEADER_FONT
        cell.fill = SECONDARY_HEADER_FILL
        cell.alignment = CENTER_WRAP_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(ws, row: int, col: int, is_alt: bool = False) -> None:
    cell = ws.cell(row=row, column=col)
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = WRAP_ALIGN
    if is_alt:
        cell.fill = ALT_ROW_FILL


def apply_data_styles(ws, data_start: int, data_end: int, max_col: int) -> None:
    """Style a contiguous range of data rows with alternating fill."""
    for r in range(data_start, data_end + 1):
        is_alt = (r - data_start) % 2 == 1
        for c in range(1, max_col + 1):
            style_data_cell(ws, r, c, is_alt)


def apply_data_styles_rows(ws, rows: Iterable[int], max_col: int) -> None:
    """Style a non-contiguous list of data rows (alternating by enumeration index).

    Use this when phase subtotals / section labels interleave with data rows,
    so that striping still looks correct across the visible data.
    """
    for idx, r in enumerate(rows):
        is_alt = idx % 2 == 1
        for c in range(1, max_col + 1):
            style_data_cell(ws, r, c, is_alt)


def style_data_row(ws, row: int, max_col: int, fill=None, font=None) -> None:
    """Style a single row with optional fill / font overrides.

    Defaults to BODY_FONT + THIN_BORDER + WRAP_ALIGN. Pass an explicit fill
    (e.g. an accent/callout fill) or font to override for highlighted / callout rows.
    """
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font if font is not None else BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill is not None:
            cell.fill = fill


def apply_input_fill_cells(ws, row: int, cols: Iterable[int]) -> None:
    """Mark cells as user-input (Cool Grey fill, thin border, body font)."""
    for col in cols:
        cell = ws.cell(row=row, column=col)
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        cell.font = BODY_FONT


__all__ = [
    # Colors
    "SAYER_YELLOW", "SAYER_BLACK", "SAYER_WHITE",
    "GREY_700", "GREY_600", "GREY_500", "GREY_300", "COOL_GREY",
    "SEV_HIGH", "SEV_MEDIUM", "SEV_LOW",
    # Typography
    "FONT_FAMILY", "TITLE_SIZE", "HEADER_SIZE", "BODY_SIZE", "CAPTION_SIZE",
    "CURRENCY_FMT",
    # Fonts
    "TITLE_FONT", "SECTION_FONT", "HEADER_FONT", "SECONDARY_HEADER_FONT",
    "BODY_FONT", "BOLD_BODY_FONT", "TOTAL_FONT", "CAPTION_FONT",
    # Fills
    "HEADER_FILL", "SECONDARY_HEADER_FILL", "ALT_ROW_FILL", "INPUT_FILL",
    "SEV_HIGH_FILL", "SEV_MED_FILL", "SEV_LOW_FILL",
    # Border + alignment
    "THIN_BORDER", "WRAP_ALIGN", "CENTER_WRAP_ALIGN",
    # Helpers
    "severity_fill", "style_header_row", "style_secondary_header_cells",
    "style_data_cell", "apply_data_styles", "apply_data_styles_rows",
    "style_data_row", "apply_input_fill_cells",
    # Re-exports
    "get_column_letter",
]
