"""
Build American Bedding CPQ project plan as Excel workbook for Google Sheets import.
Three sheets: Project Plan (task-level), Workstream Summary, Milestones.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional
import os

# --- Style constants ---
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
PHASE1_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
PHASE2_FILL = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
MILESTONE_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D5D8DC")
)


def style_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    cols: int,
    fill: Optional[PatternFill] = None,
    font: Optional[Font] = None,
) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font or NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill


def build_project_plan(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Sheet 1: Task-level project plan with week assignments."""
    headers = ["Phase", "Stage", "Week(s)", "Task", "Owner", "Status", "Notes"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    tasks = [
        # Phase 1 - Foundation (Weeks 1-4)
        ("Phase 1", "Foundation", "1-2", "Project kickoff meeting", "Sayer + AB Team", "Not Started", ""),
        ("Phase 1", "Foundation", "1-2", "NetSuite data audit (SKU quality, duplicates)", "Sayer", "Not Started", "SuiteQL queries"),
        ("Phase 1", "Foundation", "1-2", "Product catalog extraction from NetSuite", "Sayer", "Not Started", "~700 SKUs"),
        ("Phase 1", "Foundation", "1-2", "HubSpot CPQ module configuration", "Sayer", "Not Started", "Commerce Hub setup"),
        ("Phase 1", "Foundation", "1-2", "NetSuite OAuth 2.0 M2M credential setup", "Sayer + NetSuite Team", "Not Started", "API access required"),
        ("Phase 1", "Foundation", "1-2", "Kuebix API credential validation", "Sayer + Sarah-Beth", "Not Started", "Basic Auth"),
        ("Phase 1", "Foundation", "3-4", "Product catalog loaded into HubSpot", "Sayer", "Not Started", "Custom properties, bulk load"),
        ("Phase 1", "Foundation", "3-4", "NetSuite integration active (customer data sync)", "Sayer", "Not Started", "Credit status, payment terms"),
        ("Phase 1", "Foundation", "3-4", "Quote approval workflows configured", "Sayer", "Not Started", ""),
        ("Phase 1", "Foundation", "4", "Milestone review with Mike and Patrick", "Sayer + Leadership", "Not Started", "Foundation checkpoint"),
        # Phase 1 - Integration (Weeks 5-8)
        ("Phase 1", "Integration", "5-6", "Weight engine: Dorm Mattress module", "Sayer", "Not Started", "HIGH complexity -- 180+ paths"),
        ("Phase 1", "Integration", "5-6", "Weight engine: Camp Mattress module", "Sayer", "Not Started", "MEDIUM complexity -- 48 paths"),
        ("Phase 1", "Integration", "5-6", "Weight engine: Dura-Last module", "Sayer", "Not Started", "LOW-MEDIUM -- 12 paths"),
        ("Phase 1", "Integration", "5-6", "Kuebix freight integration operational", "Sayer", "Not Started", "Multi-carrier LTL rate shopping"),
        ("Phase 1", "Integration", "5-6", "Origin warehouse routing configured", "Sayer", "Not Started", "Map all shipping origins"),
        ("Phase 1", "Integration", "5-6", "Quote template design (draft)", "Sayer + Sarah-Beth", "Not Started", "Branding and layout"),
        ("Phase 1", "Integration", "7-8", "Weight engine: SoFlux + Vinyl modules", "Sayer", "Not Started", "LOW complexity"),
        ("Phase 1", "Integration", "7-8", "Quote template finalized", "Sayer + Sarah-Beth", "Not Started", "Matching NetSuite format"),
        ("Phase 1", "Integration", "7-8", "Volume discount + tax automation", "Sayer", "Not Started", "10% volume, state-based tax"),
        ("Phase 1", "Integration", "7-8", "Payment terms logic (prepay/net 10/30/60)", "Sayer", "Not Started", "Credit hold integration"),
        ("Phase 1", "Integration", "7-8", "Quote-to-order bridge tested", "Sayer", "Not Started", "HubSpot quote to NetSuite Estimate"),
        ("Phase 1", "Integration", "8", "Milestone review with leadership", "Sayer + Leadership", "Not Started", "Integration checkpoint"),
        # Phase 1 - Launch (Weeks 9-12)
        ("Phase 1", "Launch", "9-10", "End-to-end integration testing", "Sayer", "Not Started", "All three systems"),
        ("Phase 1", "Launch", "9-10", "UAT sessions with Caleb and Don", "Sayer + Sales Team", "Not Started", "2 sessions planned"),
        ("Phase 1", "Launch", "9-10", "Bug fix window", "Sayer", "Not Started", ""),
        ("Phase 1", "Launch", "9-10", "Training: Sales team (Caleb, Don)", "Sayer", "Not Started", "Recorded sessions"),
        ("Phase 1", "Launch", "9-10", "Training: Admin team (Sarah-Beth, Patrick)", "Sayer", "Not Started", "Recorded sessions"),
        ("Phase 1", "Launch", "11-12", "Go-live cutover", "Sayer + AB Team", "Not Started", ""),
        ("Phase 1", "Launch", "11-12", "Hypercare monitoring (2 weeks)", "Sayer", "Not Started", "Integration flows, freight, templates"),
        ("Phase 1", "Launch", "12", "Admin guide + user reference card delivered", "Sayer", "Not Started", ""),
        ("Phase 1", "Launch", "12", "Phase 1 close-out", "Sayer + Leadership", "Not Started", "Go/no-go for Phase 2"),
        # Phase 2 - Hardening (Weeks 13-16)
        ("Phase 2", "Hardening", "13-14", "Production API setup (Railway)", "Sayer", "Not Started", "CI/CD, monitoring, logging"),
        ("Phase 2", "Hardening", "13-14", "Weight engine migration to production API", "Sayer", "Not Started", "All 5 modules + shared utilities"),
        ("Phase 2", "Hardening", "13-14", "Unit test coverage for weight calculations", "Sayer", "Not Started", "Cell-by-cell validation"),
        ("Phase 2", "Hardening", "13-14", "Regression testing against Phase 1 baseline", "Sayer", "Not Started", ""),
        ("Phase 2", "Hardening", "15-16", "Kuebix freight migration to production API", "Sayer", "Not Started", "Retry logic for carrier failures"),
        ("Phase 2", "Hardening", "15-16", "NetSuite integration migration", "Sayer", "Not Started", "OAuth token auto-refresh"),
        ("Phase 2", "Hardening", "15-16", "HubSpot rewired to call production API", "Sayer", "Not Started", "Thin trigger layer"),
        ("Phase 2", "Hardening", "15-16", "Load testing at production volume", "Sayer", "Not Started", "65+ quotes/week"),
        # Phase 2 - Finalization (Weeks 17-18)
        ("Phase 2", "Finalization", "17-18", "Final regression testing", "Sayer", "Not Started", "Full E2E validation"),
        ("Phase 2", "Finalization", "17-18", "Documentation update", "Sayer", "Not Started", "Architecture and admin guide"),
        ("Phase 2", "Finalization", "18", "Phase 2 close-out", "Sayer + Leadership", "Not Started", "Project complete"),
    ]

    for i, task in enumerate(tasks):
        row = i + 2
        for col, val in enumerate(task, 1):
            ws.cell(row=row, column=col, value=val)

        fill = PHASE1_FILL if task[0] == "Phase 1" else PHASE2_FILL
        if "milestone" in task[3].lower() or "close-out" in task[3].lower() or "go-live" in task[3].lower():
            fill = MILESTONE_FILL
        style_data_row(ws, row, len(headers), fill=fill)

    # Column widths
    widths = [10, 14, 10, 50, 24, 14, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A2"


def build_workstream_summary(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Sheet 2: Workstream effort summary by phase."""
    headers = ["Phase", "Workstream", "Weeks", "Description"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    workstreams = [
        ("Phase 1", "CPQ Architecture & Configuration", "1-4", "Commerce Hub setup, custom properties, approval workflows, payment terms"),
        ("Phase 1", "Product Catalog Sync", "1-4", "700 SKUs extracted from NetSuite via SuiteQL, loaded into HubSpot with custom properties"),
        ("Phase 1", "NetSuite Quote-to-Order Integration", "1-4, 7-8", "OAuth 2.0 M2M authentication, quote-to-estimate-to-sales-order automation, credit hold checks"),
        ("Phase 1", "Kuebix Freight Integration", "5-6", "Multi-carrier LTL rate shopping via API, origin warehouse routing, manual override"),
        ("Phase 1", "Dynamic Weight Calculation Engine", "5-8", "Replaces 5 Excel spreadsheets with modular engine (Dorm, Camp, Dura-Last, SoFlux, Vinyl)"),
        ("Phase 1", "Quote Template Design", "5-8", "Branded PDF matching NetSuite format -- line items, discounts, tax, freight, terms"),
        ("Phase 1", "Discount, Tax & Payment Terms", "7-8", "10% volume discount, state-based tax, prepay/net 10/30/60, credit hold integration"),
        ("Phase 1", "Training & Documentation", "9-10", "Sales team + admin training sessions (recorded), admin guide, user reference card"),
        ("Phase 1", "Testing & QA", "9-12", "E2E integration testing, weight validation vs Excel, freight validation, UAT, 2-week hypercare"),
        ("Phase 1", "Project Management", "1-12", "Kickoff, weekly syncs, milestone reviews, scope management"),
        ("Phase 2", "Production API Setup (Railway)", "13-14", "Cloud hosting, CI/CD pipeline, monitoring, logging, health checks"),
        ("Phase 2", "Weight Engine Migration", "13-14", "All 5 product modules + shared utilities migrated with full test coverage"),
        ("Phase 2", "Kuebix Freight Migration", "15-16", "Freight orchestration moved to production API with retry logic"),
        ("Phase 2", "NetSuite Integration Migration", "15-16", "OAuth token management, catalog sync, quote-to-order with proper error handling"),
        ("Phase 2", "HubSpot Rewiring", "15-16", "Custom code actions updated to call production API instead of running logic locally"),
        ("Phase 2", "Integration Testing", "13-18", "E2E testing, load testing at 65+ quotes/week, regression against Phase 1 baseline"),
        ("Phase 2", "Project Management", "13-18", "Milestone reviews, documentation updates"),
    ]

    for i, ws_data in enumerate(workstreams):
        row = i + 2
        for col, val in enumerate(ws_data, 1):
            ws.cell(row=row, column=col, value=val)

        fill = PHASE1_FILL if ws_data[0] == "Phase 1" else PHASE2_FILL
        style_data_row(ws, row, len(headers), fill=fill)

    widths = [10, 38, 12, 80]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def build_milestones(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Sheet 3: Key milestones and checkpoints."""
    headers = ["Week", "Milestone", "Attendees", "Decision Point"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    milestones = [
        ("1", "Project Kickoff", "Sayer + Full AB Team", "Confirm scope, timeline, and access requirements"),
        ("4", "Foundation Milestone Review", "Sayer + Mike, Patrick", "Product catalog and NetSuite integration validated"),
        ("8", "Integration Milestone Review", "Sayer + Leadership", "Weight engine, freight, and quote-to-order tested"),
        ("9-10", "UAT Sessions", "Sayer + Caleb, Don", "Sales team validates quoting workflow"),
        ("11-12", "Phase 1 Go-Live", "Sayer + Full AB Team", "Reps begin quoting in HubSpot"),
        ("12", "Phase 1 Close-Out", "Sayer + Mike, Patrick", "Go/no-go decision for Phase 2"),
        ("14", "Phase 2 Midpoint Check", "Sayer + Mike, Patrick", "Weight engine and API migration validated"),
        ("18", "Phase 2 Close-Out / Project Complete", "Sayer + Leadership", "Production platform live, documentation delivered"),
    ]

    for i, ms in enumerate(milestones):
        row = i + 2
        for col, val in enumerate(ms, 1):
            ws.cell(row=row, column=col, value=val)
        style_data_row(ws, row, len(headers), fill=MILESTONE_FILL)

    widths = [10, 38, 28, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def main() -> None:
    wb = openpyxl.Workbook()

    # Sheet 1: Project Plan
    ws_plan = wb.active
    ws_plan.title = "Project Plan"
    build_project_plan(ws_plan)

    # Sheet 2: Workstream Summary
    ws_workstreams = wb.create_sheet("Workstream Summary")
    build_workstream_summary(ws_workstreams)

    # Sheet 3: Milestones
    ws_milestones = wb.create_sheet("Milestones")
    build_milestones(ws_milestones)

    output_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(output_dir)
    output_path = os.path.join(parent_dir, "American_Bedding_CPQ_Project_Plan.xlsx")
    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
