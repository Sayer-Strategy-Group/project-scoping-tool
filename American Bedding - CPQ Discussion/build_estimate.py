"""Generate American Bedding CPQ Scoping Estimate -- Two-Phase Hybrid Architecture

Phase 1 (V1): HubSpot-Native CPQ (~$30,000)
Phase 2 (V2): Railway Production Platform (~$11,500)

Architecture: HubSpot custom code + Railway API (not n8n).
Weight engine: composition pattern -- shared utilities + per-product modules.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
body_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', size=10, bold=True)
total_font = Font(name='Arial', size=10, bold=True, color='1F4E79')
rate_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
actuals_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
green_header_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
green_header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
phase1_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
phase2_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
grand_total_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
currency_fmt = '$#,##0'
section_font = Font(name='Arial', size=11, bold=True, color='1F4E79')
phase_label_font = Font(name='Arial', size=12, bold=True, color='1F4E79')

sev_high_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
sev_med_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
sev_low_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
resolved_font = Font(name='Arial', size=10, color='808080')


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def apply_data_styles(ws, data_start, data_end, max_col):
    for r in range(data_start, data_end + 1):
        is_alt = (r - data_start) % 2 == 1
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = wrap_align
            if is_alt:
                cell.fill = alt_fill


# ============================================================
# SHEET 1: APPROACH COMPARISON
# ============================================================
ws1 = wb.active
ws1.title = 'Approach Comparison'

ws1.merge_cells('A1:C1')
ws1.cell(row=1, column=1, value='American Bedding -- CPQ: Approach Comparison').font = Font(
    name='Arial', size=14, bold=True, color='1F4E79'
)

headers = ['Dimension', 'Recommended: HubSpot CPQ + Hybrid Architecture', 'Alternative: NetSuite Quoting + HubSpot Visibility']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header_row(ws1, 3, 3)

comparisons = [
    ('Description',
     'HubSpot becomes the quoting interface. HubSpot custom code actions handle lightweight triggers. A dedicated cloud API (Railway) orchestrates product data from NetSuite, freight rates from Kuebix, and dynamic weight calculations. Reps create and manage quotes without leaving HubSpot.',
     'Keep quoting in NetSuite. Build enhanced sync to push quote PDFs, line items, and status into HubSpot for visibility. Reps still create quotes in NetSuite but leadership sees everything in HubSpot.'),
    ('Investment',
     'Phase 1 (V1): ~$30,000 | Phase 2 (V2): ~$11,500 | Total: ~$41,500',
     '~$18,000-$22,000 (reduced scope)'),
    ('Rep Experience',
     'Single interface -- no system jumping. Quote creation, freight, pricing, and approval all in HubSpot.',
     'Still requires NetSuite for quote creation. HubSpot is read-only for quote data.'),
    ('Freight Automation',
     'Full -- Kuebix API with dynamic weight calculation. Multi-carrier LTL rate shopping automated. Phase 1 via HubSpot custom code, Phase 2 migrates to production API for unlimited execution time.',
     'None or partial -- manual Kuebix entry continues, or NetSuite-side integration (TPI scope).'),
    ('Weight Calculation',
     'Modular engine with composition architecture: shared utilities + per-product modules. Each product has unique physics (Dorm: innerspring/foam/batting, Camp: foam series + fire barrier, Dura-Last: poly fiber, SoFlux/Vinyl: covers only). Validated: no VBA, no macros, no external refs.',
     'Manual -- sales reps continue using Excel spreadsheets for weight/dimension lookups.'),
    ('NetSuite Integration',
     'Bi-directional: product catalog sync, quote-to-estimate creation, estimate-to-sales-order conversion, credit hold check.',
     'Enhanced one-way: quote visibility in HubSpot. No quote creation from HubSpot.'),
    ('Shipping Origin Routing',
     'Automated -- system determines correct warehouse origin per product/order for Kuebix API calls.',
     'Not applicable -- freight handled manually.'),
    ('Risk Level',
     'MEDIUM -- 3-system integration. Kuebix API fully reviewed. Excel calculators validated. HubSpot 20s timeout is Phase 1 limitation, eliminated in Phase 2.',
     'LOW -- simpler scope but does not solve the core pain point (4-system juggling).'),
    ('Value Delivery',
     'HIGH -- eliminates system-jumping, automates freight, reduces quote time from 10-60 min to under 5 min.',
     'LOW -- improves visibility for leadership but does not reduce rep workload or automate anything.'),
    ('Payback Period',
     'Phase 1: 6-8 months on labor savings alone',
     '12-18 months (visibility value harder to quantify)'),
    ('Production Readiness',
     'Phase 2 delivers production-grade codebase: version-controlled, unit-tested, CI/CD deployed. Maintainable by any developer.',
     'Minimal -- limited integration means less production code to maintain.'),
    ('Future Phase Readiness',
     'Strong foundation for government RFP automation, customer configurator, and external storefront.',
     'Minimal -- future phase would still need to move quoting to HubSpot.'),
]

for i, (dim, a, b) in enumerate(comparisons, 4):
    ws1.cell(row=i, column=1, value=dim)
    ws1.cell(row=i, column=2, value=a)
    ws1.cell(row=i, column=3, value=b)

apply_data_styles(ws1, 4, 4 + len(comparisons) - 1, 3)

rec_row = 4 + len(comparisons) + 1
ws1.merge_cells(f'A{rec_row}:C{rec_row}')
ws1.cell(row=rec_row, column=1, value='Recommendation').font = section_font
ws1.merge_cells(f'A{rec_row + 1}:C{rec_row + 1}')
ws1.cell(row=rec_row + 1, column=1, value=(
    'HubSpot CPQ with hybrid architecture is recommended. The core pain point is 4-system juggling '
    'that costs Caleb 10-60 minutes per quote across 65 quotes/week. The alternative (enhanced '
    'visibility) does not reduce that workload. The recommended approach eliminates system-jumping '
    'entirely, automates freight calculation, and builds the foundation for future phases. '
    'Two-phase delivery gets reps quoting in HubSpot fast (Phase 1) then hardens the platform '
    'for long-term reliability (Phase 2). The Kuebix API and Excel calculators have both been '
    'fully reviewed and validated.'
)).font = body_font
ws1.cell(row=rec_row + 1, column=1).alignment = wrap_align

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 55
ws1.column_dimensions['C'].width = 55
ws1.freeze_panes = 'A4'

# ============================================================
# SHEET 2: SCOPING ESTIMATE
# ============================================================
ws2 = wb.create_sheet('Scoping Estimate')

ws2.cell(row=1, column=1, value='Hourly Rate:').font = bold_font
ws2.cell(row=1, column=2, value=150).font = bold_font
ws2.cell(row=1, column=2).fill = rate_fill
ws2.cell(row=1, column=2).number_format = currency_fmt

headers2 = ['Workstream', 'Description', 'Min Hours', 'Max Hours', 'Median Hours',
            'Rate', 'Min Cost', 'Max Cost', 'Median Cost', 'Notes / Assumptions',
            'Actual Hours', 'Actual Cost', 'Variance']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, 13)

for col in [11, 12, 13]:
    cell = ws2.cell(row=3, column=col)
    cell.fill = green_header_fill
    cell.font = green_header_font

# -- Phase 1 label --
row = 4
ws2.merge_cells(f'A{row}:J{row}')
ws2.cell(row=row, column=1, value='PHASE 1 (V1): HUBSPOT-NATIVE CPQ').font = phase_label_font
for c in range(1, 14):
    ws2.cell(row=row, column=c).fill = phase1_fill
    ws2.cell(row=row, column=c).border = thin_border
row += 1

p1_workstreams = [
    ('HubSpot CPQ Architecture & Configuration',
     'Configure Commerce Hub CPQ module. Custom deal/quote/line item properties. Quote approval workflows. Quote numbering and validity settings (30-day window). Payment terms configuration (prepay, net 10/30/60).',
     14, 22,
     'Commerce Hub Professional confirmed. Standard CPQ config with custom properties for manufacturing context. Approval workflows for discount overrides.'),

    ('Product Catalog Sync (NetSuite to HubSpot)',
     'SuiteQL extraction of 700 SKUs from NetSuite. Map to HubSpot products with custom properties: weight components, freight class, dimensions (L x W x H), cover type, foam series, construction type. HubSpot custom code sync trigger for ongoing price and product updates.',
     20, 32,
     'Flat catalog -- one product per SKU. SuiteQL batch queries (1-2 API calls). HubSpot custom code trigger replaces n8n scheduled sync. OAuth 2.0 M2M auth.'),

    ('NetSuite Quote-to-Order Integration',
     'OAuth 2.0 M2M setup. Quote acceptance triggers NetSuite Estimate creation via HubSpot custom code -> REST API. Estimate-to-Sales Order conversion. Customer credit hold check. Payment terms sync. Price level sync.',
     22, 36,
     'Bi-directional: HubSpot quote -> NetSuite estimate -> sales order -> work order. Credit hold reads AR data. 1.25x ERP adjacency premium (existing integration). HubSpot custom code handles orchestration.'),

    ('Kuebix Freight Integration + Origin Routing',
     'HubSpot custom code: collect line item weights/dims -> determine origin warehouse -> POST /action/quickRate to Kuebix -> parse multi-carrier LTL response -> write cheapest rate as freight line item. Manual override option. Error handling with fallback.',
     22, 34,
     'Kuebix API fully reviewed (OpenAPI 3.0.2). Basic Auth. quickRate returns multi-carrier quotes. Full street address required. Origin routing for multiple warehouses. persist=false prevents phantom shipments. 20s HubSpot timeout risk for multi-line quotes.'),

    ('Dynamic Weight Calculation Engine',
     'Composition architecture: shared utilities (fabricCalc, cubeVolume, coverWeights, packagingCalc) + one module per product line. Dorm Mattress (innerspring + foam + batting + cover), Camp Mattress (foam core + fire barrier), Dura-Last (poly fiber), SoFlux OX/Vinyl (covers only). Each validated cell-by-cell against its Excel file. Cube volume for LTL/TL determination.',
     28, 44,
     'HIGHEST-COMPLEXITY WORKSTREAM. Composition architecture -- no single template. Each product has unique physics, seam allowances, and construction constants that cannot be derived by analogy. '
     'Dorm (HIGH -- 180+ paths, 9-level nesting), Camp (MED -- 48 paths, foam series + fire barrier), Dura-Last (LOW-MED -- 12 paths, poly fiber), SoFlux OX/Vinyl (LOW -- ~95% shared base). '
     'Validated 2026-04-04: no VBA, no external refs, no circular formulas. Open: "Special Case Check with Delbert" in Dorm Mattress.'),

    ('Quote Template Design',
     'Custom HubL/HTML/CSS quote template matching current NetSuite format. Line item table (Qty, Item #, Description, Unit Price, Amount). Subtotal, volume discount (10%), state-based tax, shipping line, total. Damaged freight policy terms. Customer acceptance flow. PDF optimization.',
     12, 20,
     'Reference: 5 NetSuite PDF examples. Must match header/footer, line item format, terms. HubSpot Flow theme supports custom properties. @media print CSS for PDF.'),

    ('Discount, Tax & Payment Terms Logic',
     'Volume discount automation (10% threshold). State-based tax calculation (9.75% TN, exempt OOS). Payment terms workflow (prepay/net 10/30/60). Credit hold integration from NetSuite AR. All via HubSpot custom code.',
     10, 18,
     'Tax may need external service for multi-state. Discount approval thresholds TBD at kickoff. Credit hold reads NetSuite -- flags but does not hard-block. Fits within HubSpot custom code 20s limit.'),

    ('Training & Documentation',
     'Sales training (Caleb, Don): quoting workflow, product selection, freight review. Admin training (Sarah-Beth, Patrick): system management, catalog updates, sync monitoring. All virtual, recorded. Admin guide + user quick reference card.',
     10, 16,
     '2-3 sessions. Emphasis on time savings. Admin guide covers HubSpot custom code monitoring and product sync. Record for onboarding.'),

    ('Testing & QA',
     'E2E integration testing across all 3 systems. Quote accuracy vs. NetSuite examples. Weight calc vs. Excel baselines (cell-by-cell per product -- no cross-product shortcuts). Freight vs. manual Kuebix lookups. UAT with Caleb and Don. 2-week hypercare.',
     14, 24,
     'Test matrix: product sync accuracy, weight calc per product line, freight quote accuracy, template rendering, NetSuite estimate creation, SO conversion, credit hold. 2 UAT sessions. Dorm Mattress sampling strategy for 180+ paths.'),

    ('Project Management',
     'Kickoff, weekly syncs, milestone reviews. Coordination with American Bedding team. Documentation and change request management. Scope creep prevention. Go-live planning.',
     16, 24,
     '~10% of total. Weekly syncs with Sarah-Beth. Milestone reviews at weeks 4, 8, and go-live. All change requests documented.'),
]

p1_start = row
for name, desc, min_h, max_h, notes in p1_workstreams:
    ws2.cell(row=row, column=1, value=name)
    ws2.cell(row=row, column=2, value=desc)
    ws2.cell(row=row, column=3, value=min_h)
    ws2.cell(row=row, column=4, value=max_h)
    ws2.cell(row=row, column=5).value = f'=AVERAGE(C{row},D{row})'
    ws2.cell(row=row, column=6).value = '=$B$1'
    ws2.cell(row=row, column=7).value = f'=C{row}*$B$1'
    ws2.cell(row=row, column=8).value = f'=D{row}*$B$1'
    ws2.cell(row=row, column=9).value = f'=E{row}*$B$1'
    ws2.cell(row=row, column=10, value=notes)
    ws2.cell(row=row, column=13).value = f'=IF(K{row}="","",K{row}-E{row})'
    for c in range(1, 14):
        ws2.cell(row=row, column=c).font = body_font
        ws2.cell(row=row, column=c).border = thin_border
        ws2.cell(row=row, column=c).alignment = wrap_align
    for c in [11, 12, 13]:
        ws2.cell(row=row, column=c).fill = actuals_fill
    for c in [7, 8, 9]:
        ws2.cell(row=row, column=c).number_format = currency_fmt
    row += 1

# Phase 1 subtotal
p1_total = row
ws2.cell(row=row, column=1, value='PHASE 1 SUBTOTAL').font = total_font
for c in range(1, 14):
    ws2.cell(row=row, column=c).fill = total_fill
    ws2.cell(row=row, column=c).border = thin_border
for col in [3, 4, 5]:
    cl = get_column_letter(col)
    ws2.cell(row=row, column=col).value = f'=SUM({cl}{p1_start}:{cl}{row - 1})'
    ws2.cell(row=row, column=col).font = total_font
ws2.cell(row=row, column=6).value = '=$B$1'
ws2.cell(row=row, column=6).font = total_font
for col in [7, 8, 9]:
    cl = get_column_letter(col)
    ws2.cell(row=row, column=col).value = f'=SUM({cl}{p1_start}:{cl}{row - 1})'
    ws2.cell(row=row, column=col).font = total_font
    ws2.cell(row=row, column=col).number_format = currency_fmt
for col in [11, 12]:
    cl = get_column_letter(col)
    ws2.cell(row=row, column=col).value = f'=SUM({cl}{p1_start}:{cl}{row - 1})'
row += 2

# -- Phase 2 label --
ws2.merge_cells(f'A{row}:J{row}')
ws2.cell(row=row, column=1, value='PHASE 2 (V2): RAILWAY PRODUCTION PLATFORM').font = phase_label_font
for c in range(1, 14):
    ws2.cell(row=row, column=c).fill = phase2_fill
    ws2.cell(row=row, column=c).border = thin_border
row += 1

p2_workstreams = [
    ('Railway Infrastructure Setup',
     'Production API project setup on Railway. CI/CD pipeline (git push deploy). Environment config (staging + production). Health monitoring endpoints. Error logging and alerting. Security hardening.',
     8, 12,
     'Railway provides managed hosting with auto-deploy from git. Node.js or Python runtime. Environment variables for secrets. Health check endpoint for uptime monitoring.'),

    ('Weight Engine Migration',
     'Refactor 5 product modules + shared utilities into Railway API. Composition architecture: shared layer (fabricCalc, cubeVolume, coverWeights, packagingCalc) + per-product modules (dormMattress, campMattress, duraLast, sofluxCover, vinylCover). Each product built and tested independently.',
     10, 16,
     'Code exists from Phase 1 -- migration is refactor + test coverage. Each module validated cell-by-cell against its own Excel file -- no cross-product shortcuts. '
     'Dorm needs ~20 representative test cases for 180+ paths. SoFlux/Vinyl share ~95% code. Target: 100% coverage for critical paths.'),

    ('Kuebix Freight Migration',
     'Migrate Kuebix quickRate orchestration to Railway API. Multi-line, multi-carrier quote assembly without timeout risk. Retry logic for carrier API failures. Response caching for repeat shipping lanes.',
     8, 14,
     'Eliminates HubSpot 20s timeout constraint. Retry logic handles intermittent Kuebix API failures. Cache reduces API calls for repeat lanes.'),

    ('NetSuite Integration Migration',
     'Migrate NetSuite OAuth token management, catalog sync, and quote-to-order flow to Railway. Automated catalog sync via Railway cron. Proper error handling and retry for REST API calls.',
     10, 16,
     'OAuth 2.0 M2M token refresh automated. SuiteQL catalog extraction on schedule. Estimate/SO creation with retry. Full audit logging.'),

    ('HubSpot Rewiring',
     'Update HubSpot custom code actions to call Railway API endpoints instead of running logic locally. HubSpot becomes a thin trigger layer. Webhook configuration.',
     6, 10,
     'Simple HTTP calls from HubSpot custom code to Railway API. Minimal logic in HubSpot -- just trigger + response handling.'),

    ('Integration & Regression Testing',
     'Full regression suite: every Phase 1 workflow operates identically on Railway. Load testing at 65+ quotes/week. Performance benchmarking. Documentation update.',
     8, 12,
     'Regression against Phase 1 baseline ensures zero behavior change. Load test confirms production volume capacity. Updated admin guide and architecture docs.'),

    ('Project Management',
     'Milestone reviews, documentation updates, Phase 2 close-out.',
     6, 10,
     'Lighter PM load than Phase 1. ~10% of Phase 2 hours.'),
]

p2_start = row
for name, desc, min_h, max_h, notes in p2_workstreams:
    ws2.cell(row=row, column=1, value=name)
    ws2.cell(row=row, column=2, value=desc)
    ws2.cell(row=row, column=3, value=min_h)
    ws2.cell(row=row, column=4, value=max_h)
    ws2.cell(row=row, column=5).value = f'=AVERAGE(C{row},D{row})'
    ws2.cell(row=row, column=6).value = '=$B$1'
    ws2.cell(row=row, column=7).value = f'=C{row}*$B$1'
    ws2.cell(row=row, column=8).value = f'=D{row}*$B$1'
    ws2.cell(row=row, column=9).value = f'=E{row}*$B$1'
    ws2.cell(row=row, column=10, value=notes)
    ws2.cell(row=row, column=13).value = f'=IF(K{row}="","",K{row}-E{row})'
    for c in range(1, 14):
        ws2.cell(row=row, column=c).font = body_font
        ws2.cell(row=row, column=c).border = thin_border
        ws2.cell(row=row, column=c).alignment = wrap_align
    for c in [11, 12, 13]:
        ws2.cell(row=row, column=c).fill = actuals_fill
    for c in [7, 8, 9]:
        ws2.cell(row=row, column=c).number_format = currency_fmt
    row += 1

# Phase 2 subtotal
p2_total = row
ws2.cell(row=row, column=1, value='PHASE 2 SUBTOTAL').font = total_font
for c in range(1, 14):
    ws2.cell(row=row, column=c).fill = total_fill
    ws2.cell(row=row, column=c).border = thin_border
for col in [3, 4, 5]:
    cl = get_column_letter(col)
    ws2.cell(row=row, column=col).value = f'=SUM({cl}{p2_start}:{cl}{row - 1})'
    ws2.cell(row=row, column=col).font = total_font
ws2.cell(row=row, column=6).value = '=$B$1'
ws2.cell(row=row, column=6).font = total_font
for col in [7, 8, 9]:
    cl = get_column_letter(col)
    ws2.cell(row=row, column=col).value = f'=SUM({cl}{p2_start}:{cl}{row - 1})'
    ws2.cell(row=row, column=col).font = total_font
    ws2.cell(row=row, column=col).number_format = currency_fmt
for col in [11, 12]:
    cl = get_column_letter(col)
    ws2.cell(row=row, column=col).value = f'=SUM({cl}{p2_start}:{cl}{row - 1})'
row += 2

# Grand total
gt_row = row
ws2.cell(row=row, column=1, value='GRAND TOTAL').font = Font(name='Arial', size=12, bold=True, color='1F4E79')
for c in range(1, 14):
    ws2.cell(row=row, column=c).fill = grand_total_fill
    ws2.cell(row=row, column=c).border = thin_border
for col in [3, 4, 5]:
    cl = get_column_letter(col)
    ws2.cell(row=row, column=col).value = f'={cl}{p1_total}+{cl}{p2_total}'
    ws2.cell(row=row, column=col).font = Font(name='Arial', size=12, bold=True)
ws2.cell(row=row, column=6).value = '=$B$1'
for col in [7, 8, 9]:
    cl = get_column_letter(col)
    ws2.cell(row=row, column=col).value = f'={cl}{p1_total}+{cl}{p2_total}'
    ws2.cell(row=row, column=col).font = Font(name='Arial', size=12, bold=True)
    ws2.cell(row=row, column=col).number_format = currency_fmt
for col in [11, 12]:
    cl = get_column_letter(col)
    ws2.cell(row=row, column=col).value = f'={cl}{p1_total}+{cl}{p2_total}'
row += 2

# Fee rows
ws2.cell(row=row, column=1, value='PHASE 1 FIXED FEE:').font = Font(name='Arial', size=11, bold=True, color='1F4E79')
ws2.cell(row=row, column=9, value=30000).font = Font(name='Arial', size=11, bold=True, color='1F4E79')
ws2.cell(row=row, column=9).number_format = currency_fmt
ws2.cell(row=row, column=10, value='Fixed fee. Client sees this number only -- no hours breakdown in proposal.').font = body_font
row += 1
ws2.cell(row=row, column=1, value='PHASE 2 FIXED FEE:').font = Font(name='Arial', size=11, bold=True, color='1F4E79')
ws2.cell(row=row, column=9, value=11500).font = Font(name='Arial', size=11, bold=True, color='1F4E79')
ws2.cell(row=row, column=9).number_format = currency_fmt
ws2.cell(row=row, column=10, value='Fixed fee. Separate SOW or Phase 2 addendum.').font = body_font
row += 1
ws2.cell(row=row, column=1, value='COMBINED FIXED FEE:').font = Font(name='Arial', size=12, bold=True, color='1F4E79')
ws2.cell(row=row, column=9).value = f'=I{row - 2}+I{row - 1}'
ws2.cell(row=row, column=9).font = Font(name='Arial', size=12, bold=True, color='1F4E79')
ws2.cell(row=row, column=9).number_format = currency_fmt

widths = {1: 30, 2: 55, 3: 12, 4: 12, 5: 14, 6: 10, 7: 12, 8: 12, 9: 14, 10: 55, 11: 14, 12: 14, 13: 12}
for col, w in widths.items():
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.freeze_panes = 'B4'

# ============================================================
# SHEET 3: RISK REGISTER
# ============================================================
ws3 = wb.create_sheet('Risk Register')

headers3 = ['#', 'Risk', 'Severity', 'Likelihood', 'Impact', 'Mitigation', 'Owner', 'Status']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header_row(ws3, 1, 8)

risks = [
    (1, 'HubSpot custom code 20-second execution timeout for complex multi-line quotes',
     'MEDIUM', 'MEDIUM',
     'Multi-product quotes requiring weight calc + Kuebix API call may approach or exceed 20s limit. Quote fails to generate freight line item.',
     'Test with largest known quote configurations in Phase 1. Manual fallback for edge cases. Phase 2 eliminates this risk entirely by migrating to Railway API (no execution limits).',
     'Consultant', 'Open -- Phase 1 risk'),

    (2, 'Excel weight calculator edge cases -- "Special Case Check with Delbert" in Dorm Mattress',
     'MEDIUM', 'LOW',
     'Undocumented business logic in Dorm Mattress calculator. #VALUE! errors noted. Edge cases may produce incorrect weights for specific configurations.',
     'Clarify with Caleb/Delbert at kickoff. Build Dorm Mattress calculator first. Sampling strategy for 180+ paths (~20 representative test cases). Validate against known weights.',
     'Consultant + Caleb', 'Open'),

    (3, 'NetSuite data quality -- SKU duplicates, stale pricing, missing attributes from Jan 2026 launch',
     'HIGH', 'HIGH',
     'Product sync imports bad data into HubSpot. Quotes show wrong prices or descriptions. Customer trust eroded.',
     'NetSuite item data audit during Week 1. Define cleanup scope before building sync. SuiteQL queries validate data quality before bulk import. Flag items missing required fields.',
     'Consultant + Sarah-Beth', 'Open'),

    (4, 'Scope creep via Vesco/Kaitlynn relationship -- government RFP work creeps into Phase 1',
     'HIGH', 'MEDIUM',
     'Informal requests bypass SOW. Government-specific requirements consume hours budgeted for CPQ. Don Reynolds enthusiasm for future phases pulls scope.',
     'SOW signed by Mike Taylor or Patrick (budget authority). Phase boundaries iron-clad in SOW. All change requests documented. Weekly hours tracking visible to client.',
     'Kyle + Mike', 'Open'),

    (5, 'NetSuite still stabilizing (launched Jan 1, 2026) -- API schema changes, TPI updates',
     'HIGH', 'MEDIUM',
     'Integration breaks after TPI makes changes. Data model shifts invalidate SuiteQL queries. API access disrupted during maintenance.',
     'Build against documented REST API (not custom SuiteScript). Version-lock API endpoints. Establish communication channel with TPI for change notifications.',
     'Consultant', 'Open'),

    (6, 'Multiple shipping origins -- wrong origin = wrong freight = wrong price',
     'MEDIUM', 'HIGH',
     'Incorrect origin address sent to Kuebix. Freight quotes are wrong. Customer receives inaccurate pricing. Trust and margin impact.',
     'Map all warehouse addresses at kickoff. Define explicit routing rules. Test each origin with known Kuebix quotes. Default origin fallback if routing is ambiguous.',
     'Consultant + Mike', 'Open'),

    (7, 'HubSpot CPQ limitations -- no calculated fields, no product variants, no native tax engine',
     'MEDIUM', 'HIGH',
     'Workarounds needed for freight as line item (must be pre-computed). Product selection UX is less polished than dedicated CPQ. Tax calculation requires external logic.',
     'Pre-compute all calculated values in custom code before quote is published. Use custom properties and filtered views for product selection. Tax logic in Ops Hub custom code.',
     'Consultant', 'Open'),

    (8, 'Kuebix API rate limits undocumented -- high quote volume could hit throttling',
     'MEDIUM', 'LOW',
     '65 quotes/week plus revisions could exceed undocumented rate limits. Freight quotes fail or are throttled.',
     'Confirm rate limits with Kuebix support. Implement response caching for repeat shipping lanes (Phase 2). Graceful fallback to manual entry.',
     'Consultant', 'Open'),

    (9, 'Adoption risk -- sales reps switching from NetSuite quoting to HubSpot quoting',
     'MEDIUM', 'MEDIUM',
     'Caleb and Don revert to old process. New system goes unused. ROI unrealized.',
     'Executive mandate from Mike. Show concrete time savings in training. 2-week parallel run. Monitor adoption metrics. Caleb is motivated (described the pain).',
     'Mike + Consultant', 'Open'),

    (10, 'Freight class determination per product is unclear or inconsistent',
     'MEDIUM', 'MEDIUM',
     'Wrong freight class to Kuebix. Shipping costs systematically over or under estimated. Margin impact across all quotes.',
     'Confirm freight class source at kickoff. Document per product line. Validate against known Kuebix quotes.',
     'Consultant + Caleb', 'Open'),

    (11, 'Budget authority anchored on Billy $25K verbal benchmark -- two-phase total of ~$41.5K creates sticker shock',
     'MEDIUM', 'MEDIUM',
     'Patrick (CFO) expects $25K. Phase 1 at ~$30K may trigger budget negotiation.',
     'Frame $25K as pre-discovery estimate. Phase 1 at ~$30K is closer to anchor than previous $34.5K. Two-phase reduces upfront commitment. Value-based ROI framing: Phase 1 pays back in 6-8 months.',
     'Kyle + Billy', 'Open'),
]

for i, (num, risk, sev, lik, impact, mit, owner, status) in enumerate(risks, 2):
    ws3.cell(row=i, column=1, value=num)
    ws3.cell(row=i, column=2, value=risk)
    ws3.cell(row=i, column=3, value=sev)
    ws3.cell(row=i, column=4, value=lik)
    ws3.cell(row=i, column=5, value=impact)
    ws3.cell(row=i, column=6, value=mit)
    ws3.cell(row=i, column=7, value=owner)
    ws3.cell(row=i, column=8, value=status)

apply_data_styles(ws3, 2, len(risks) + 1, 8)
for i, (_, _, sev, *_) in enumerate(risks, 2):
    c = ws3.cell(row=i, column=3)
    if sev == 'HIGH':
        c.fill = sev_high_fill
    elif sev == 'MEDIUM':
        c.fill = sev_med_fill
    else:
        c.fill = sev_low_fill

# Resolved risk
res_row = len(risks) + 3
ws3.cell(row=res_row, column=1, value='RESOLVED').font = Font(name='Arial', size=10, bold=True, color='008000')
res_row += 1
ws3.cell(row=res_row, column=1, value='--').font = resolved_font
ws3.cell(row=res_row, column=2, value='Excel weight calculator logic more complex than analyzed -- VBA macros, external references, or circular formulas').font = resolved_font
ws3.cell(row=res_row, column=3, value='HIGH').font = resolved_font
ws3.cell(row=res_row, column=4, value='MEDIUM').font = resolved_font
ws3.cell(row=res_row, column=5, value='Dynamic weight engine blows past 44 hrs. Fixed fee untenable. Workstream 5 unbounded.').font = resolved_font
ws3.cell(row=res_row, column=6, value='RESOLVED 2026-04-04: All 5 Excel files received and validated. No VBA, no external refs, no circular formulas. Dorm Mattress confirmed as highest complexity (9-level nesting, 180+ paths) but within estimate range.').font = resolved_font
ws3.cell(row=res_row, column=7, value='Kyle').font = resolved_font
ws3.cell(row=res_row, column=8, value='Closed').font = resolved_font
for c in range(1, 9):
    ws3.cell(row=res_row, column=c).border = thin_border
    ws3.cell(row=res_row, column=c).alignment = wrap_align

widths3 = {1: 5, 2: 50, 3: 12, 4: 12, 5: 45, 6: 55, 7: 25, 8: 18}
for col, w in widths3.items():
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.freeze_panes = 'A2'

# ============================================================
# SHEET 4: ASSUMPTIONS & EXCLUSIONS
# ============================================================
ws4 = wb.create_sheet('Assumptions & Exclusions')

ws4.cell(row=1, column=1, value='Scope Assumptions').font = section_font
assumptions = [
    'HubSpot Sales Hub Professional + Commerce Hub Professional (existing)',
    'Operations Hub Professional required (separate license ~$800/mo -- enables custom code actions)',
    'Up to 700 products in HubSpot product library (flat catalog, one record per SKU)',
    'NetSuite REST API access with OAuth 2.0 M2M credentials provided by client',
    'Kuebix API credentials (username, API key, 15-character Client ID) provided by client',
    'Railway cloud platform for production API hosting (~$5-20/mo -- Phase 2)',
    'Multiple shipping origin addresses -- client provides complete warehouse list with routing rules at kickoff',
    'Freight class stored as a property per product in NetSuite or as a static lookup per product line',
    'Up to 30 custom properties across deals, quotes, and line items (overages are change orders)',
    'Phase 1: 10-12 weeks | Phase 2: 4-6 weeks (18 weeks total)',
    'All work conducted remotely',
    'Rate: $150/hr (internal tracking -- client sees fixed fee only)',
    'Phase 1 fixed fee: ~$30,000 -- 5 payments of $6,300 ($6,000 + 5% tech/admin) every 15 days, Net-15',
    'Phase 2 fixed fee: ~$11,500 -- 3 payments of ~$4,027 (~$3,835 + 5% tech/admin) every 15 days, Net-15',
    'Dynamic weight calculation covers 5 product lines (Dorm, Camp, Dura-Last, SoFlux OX, Vinyl)',
    'Excel calculators validated: no VBA, no external refs, no circular formulas (confirmed 2026-04-04)',
    'Weight engine uses composition architecture: shared utilities + per-product modules (not a single template)',
    'Quote template matches current NetSuite format (5 PDF examples as reference)',
    'SOW signed by Mike Taylor (VP) or Patrick (CFO) -- budget authority',
    'Production code delivered as version-controlled codebase (not n8n workflows)',
]
for i, a in enumerate(assumptions, 2):
    ws4.cell(row=i, column=1, value=i - 1)
    ws4.cell(row=i, column=2, value=a)

gap = len(assumptions) + 3
ws4.cell(row=gap, column=1, value='Out of Scope').font = section_font
exclusions = [
    'Government RFP automation, compliance tracking, bid management (future phase)',
    'Customer-facing product configurator / guided selling decision tree (future phase)',
    'External storefront build (Shopify or HubSpot Commerce -- future phase)',
    'VoIP / call capture system selection or integration',
    'AI lead enrichment from call transcripts',
    'NetSuite reimplementation, new module deployment, or SuiteScript development',
    'Ongoing Kuebix TMS administration or carrier management',
    'HubSpot Marketing Hub configuration or marketing automation',
    'Custom self-service pricing portal',
    'Ongoing managed services beyond 2-week hypercare per phase',
    'HubSpot or NetSuite license procurement or tier changes (recommendations only)',
    'Data migration from external systems beyond NetSuite product catalog sync',
    'TPI (NetSuite partner) coordination or management -- Sayer works directly in NetSuite',
    'n8n as production runtime (may be used for prototyping only)',
]
for i, e in enumerate(exclusions, gap + 1):
    ws4.cell(row=i, column=1, value=i - gap)
    ws4.cell(row=i, column=2, value=e)

gap2 = gap + len(exclusions) + 2
ws4.cell(row=gap2, column=1, value='Client Responsibilities').font = section_font
responsibilities = [
    'Provide Kuebix API credentials (username, API key, 15-character Client ID)',
    'Provide NetSuite admin access or create OAuth 2.0 M2M integration record with API permissions',
    'Excel shipping calculator files -- RECEIVED AND VALIDATED (2026-04-04)',
    'Provide complete warehouse address list with product/region routing rules',
    'Confirm freight class per product line (or provide lookup reference)',
    'Designate primary point of contact (Sarah-Beth recommended)',
    'Ensure Caleb and Don availability for 2+ UAT sessions during weeks 9-10',
    'SOW signed by Mike Taylor or Patrick (CFO) -- budget authority',
    'Validate 5-10 sample quotes against current NetSuite output during UAT',
    'Provide 5-10 recent quotes with known freight costs for validation testing',
    'Clarify "Special Case Check with Delbert" for Dorm Mattress calculator',
]
for i, r in enumerate(responsibilities, gap2 + 1):
    ws4.cell(row=i, column=1, value=i - gap2)
    ws4.cell(row=i, column=2, value=r)

gap3 = gap2 + len(responsibilities) + 2
ws4.cell(row=gap3, column=1, value='Open Items (Confirm During Kickoff)').font = section_font
open_items = [
    'NetSuite item types in use (InventoryItem vs. AssemblyItem vs. NonInventoryItem)',
    'NetSuite price level structure (single vs. multiple price levels per customer tier)',
    'Exact number and full street addresses of all shipping warehouses',
    'Routing rules: which products or regions ship from which warehouse',
    'Discount approval thresholds -- is 10% always automatic? Do higher discounts need approval?',
    'Credit hold workflow details -- what triggers a hold, who can override, block vs. flag?',
    'Quote validity period (30 days per current NetSuite quotes -- confirm)',
    'Dorm Mattress: "Special Case Check with Delbert" -- undocumented business logic for edge cases',
    'Dimensional data future-proofing -- carriers may require L x W x H per item',
    'Kuebix rate limits and acceptable API call volume (confirm with Kuebix support)',
    'Operations Hub Professional -- client decision on adding this tier (required for Phase 1)',
    'Railway hosting -- confirm client or Sayer hosts the production API (Phase 2)',
]
for i, o in enumerate(open_items, gap3 + 1):
    ws4.cell(row=i, column=1, value=i - gap3)
    ws4.cell(row=i, column=2, value=o)

for row_cells in ws4.iter_rows(min_row=2, max_col=3):
    for cell in row_cells:
        cell.font = body_font
        cell.alignment = wrap_align

ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 90

# SAVE
output_path = '/Users/harbuckconsulting/projects/project-scoping-tool/American Bedding - CPQ Discussion/American_Bedding_CPQ_Estimate.xlsx'
wb.save(output_path)
print(f'Saved to: {output_path}')
